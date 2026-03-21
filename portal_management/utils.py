"""
portal_management/utils.py
════════════════
Excel export utilities for student and session result reports.

All data is loaded upfront in bulk queries before any Excel writing begins.
No database queries inside loops.
"""

import logging
from collections import defaultdict
from decimal import Decimal
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import GradingScale, Student, StudentEnrollment, StudentExamMetrics, StudentExamPosition, StudentSubjectResult, Subject

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Shared style helpers
# ─────────────────────────────────────────────────────────────────────────────

_THIN = Side(style='thin')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
BOLD = Font(bold=True)
BOLD_14 = Font(bold=True, size=14)


def _hcell(ws, row, col, value, merge_end_col=None):
    """Write a styled header cell, optionally merging columns."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_end_col,
        )
    return cell


def _cell(ws, row, col, value, align=CENTER, bold=False):
    """Write a plain bordered cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = align
    c.border = BORDER
    if bold:
        c.font = BOLD
    return c


# ─────────────────────────────────────────────────────────────────────────────
# Student result report — one student, multiple sessions
# ─────────────────────────────────────────────────────────────────────────────

def export_student_report(student, exam_sessions) -> openpyxl.Workbook:
    """
    Build an Excel workbook showing a student's results across one or more
    exam sessions, with per-subject aggregates and session-level summaries.

    All DB queries are issued before any Excel writing begins.

    Args:
        student:       Student model instance
        exam_sessions: iterable of ExamSession instances

    Returns:
        openpyxl.Workbook
    """


    sessions = list(exam_sessions)
    session_ids = [s.id for s in sessions]
    session_by_id = {s.id: s for s in sessions}

    if not sessions:
        wb = openpyxl.Workbook()
        wb.active.title = 'No Data'
        wb.active['A1'] = 'No exam sessions found.'
        return wb

    # ── Determine education level from current enrollment ─────────────────
    enrollment = (
        StudentEnrollment.objects
        .filter(student=student)
        .select_related('class_level__educational_level')
        .order_by('-academic_year__start_date')
        .first()
    )
    education_level = (
        enrollment.class_level.educational_level if enrollment else None
    )
    level_type = education_level.level_type if education_level else ''

    # ── Load grading scale once ────────────────────────────────────────────
    grading_scale = []
    if education_level:
        grading_scale = list(
            GradingScale.objects
            .filter(education_level=education_level)
            .order_by('-min_mark')
            .values('grade', 'min_mark', 'max_mark', 'points')
        )

    # ── Load ALL subject results for this student × sessions in ONE query ──
    results_qs = list(
        StudentSubjectResult.objects
        .filter(student=student, exam_session_id__in=session_ids)
        .select_related('subject')
        .values(
            'subject_id', 'subject__name', 'subject__short_name',
            'exam_session_id', 'total_marks', 'grade', 'points',
        )
    )

    # ── Load metrics and positions in ONE query each ───────────────────────
    metrics_by_session = {
        m.exam_session_id: m
        for m in StudentExamMetrics.objects.filter(
            student=student, exam_session_id__in=session_ids
        )
    }
    positions_by_session = {
        p.exam_session_id: p
        for p in StudentExamPosition.objects.filter(
            student=student, exam_session_id__in=session_ids
        )
    }

    # ── Organise results by subject ────────────────────────────────────────
    # subject_id → {name, session_id → {marks, grade, points}}
    subjects: dict[int, dict] = {}
    for row in results_qs:
        sid = row['subject_id']
        if sid not in subjects:
            subjects[sid] = {
                'name': row['subject__name'],
                'short_name': row['subject__short_name'] or row['subject__name'],
                'sessions': {},
                'total_marks': Decimal('0'),
                'total_points': Decimal('0'),
                'count': 0,
            }
        if row['total_marks'] is not None:
            subjects[sid]['sessions'][row['exam_session_id']] = {
                'marks': row['total_marks'],
                'grade': row['grade'] or '-',
                'points': row['points'],
            }
            subjects[sid]['total_marks'] += row['total_marks']
            subjects[sid]['total_points'] += (row['points'] or Decimal('0'))
            subjects[sid]['count'] += 1

    # ── Compute subject averages and overall grades ────────────────────────
    for sid, subj in subjects.items():
        if subj['count'] > 0:
            avg = subj['total_marks'] / subj['count']
            subj['average_marks'] = avg.quantize(Decimal('0.01'))
            # Resolve overall grade from grading scale
            subj['overall_grade'] = '-'
            for band in grading_scale:
                if band['min_mark'] <= avg <= band['max_mark']:
                    subj['overall_grade'] = band['grade']
                    break
        else:
            subj['average_marks'] = None
            subj['overall_grade'] = '-'

    subjects_list = sorted(subjects.values(), key=lambda s: s['name'])

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Report'

    # Title
    ws.merge_cells('A1:B1')
    last_col = get_column_letter(3 + len(sessions) * 2 + 5)
    ws.merge_cells(f'A1:{last_col}1')
    t = ws['A1']
    t.value = 'STUDENT EXAM RESULTS REPORT'
    t.font = BOLD_14
    t.alignment = CENTER

    # Student info row
    ws['A2'] = f"Name: {student.full_name}"
    ws['A2'].font = BOLD
    ws['D2'] = f"Reg No: {student.registration_number or 'N/A'}"
    ws['F2'] = f"Gender: {student.get_gender_display() if student.gender else 'N/A'}"

    # ── Headers ────────────────────────────────────────────────────────────
    HEADER_ROW = 4
    DATA_ROW = HEADER_ROW + 2

    # S/N and Subject — span 2 header rows
    for col, label in [(1, 'S/N'), (2, 'SUBJECT')]:
        _hcell(ws, HEADER_ROW, col, label)
        ws.merge_cells(
            start_row=HEADER_ROW, start_column=col,
            end_row=HEADER_ROW + 1, end_column=col,
        )

    col = 3
    session_col_map: dict[int, int] = {}
    for session in sessions:
        _hcell(ws, HEADER_ROW, col, session.name, merge_end_col=col + 1)
        _hcell(ws, HEADER_ROW + 1, col, 'MARKS')
        _hcell(ws, HEADER_ROW + 1, col + 1, 'GRADE')
        session_col_map[session.id] = col
        col += 2

    AGGS = ['TOTAL', 'AVERAGE', 'GRADE', 'REMARK']
    agg_col: dict[str, int] = {}
    for label in AGGS:
        _hcell(ws, HEADER_ROW, col, label)
        ws.merge_cells(
            start_row=HEADER_ROW, start_column=col,
            end_row=HEADER_ROW + 1, end_column=col,
        )
        agg_col[label] = col
        col += 1

    total_cols = col - 1

    # ── Subject rows ───────────────────────────────────────────────────────
    for idx, subj in enumerate(subjects_list, start=1):
        row = DATA_ROW + idx - 1
        _cell(ws, row, 1, idx)
        _cell(ws, row, 2, subj['name'], align=LEFT)

        for sess_id, start_col in session_col_map.items():
            sess_data = subj['sessions'].get(sess_id, {})
            marks = sess_data.get('marks')
            grade = sess_data.get('grade', '-')
            _cell(ws, row, start_col, float(marks) if marks is not None else '-')
            _cell(ws, row, start_col + 1, grade)

        avg = subj['average_marks']
        total = subj['total_marks'] if subj['count'] > 0 else None
        remark = ('Pass' if avg is not None and avg >= 40 else 'Fail') if avg is not None else '-'

        _cell(ws, row, agg_col['TOTAL'], float(total) if total else '-')
        _cell(ws, row, agg_col['AVERAGE'], float(avg) if avg is not None else '-')
        _cell(ws, row, agg_col['GRADE'], subj['overall_grade'])
        _cell(ws, row, agg_col['REMARK'], remark)

    current_row = DATA_ROW + len(subjects_list) + 1

    # ── Session summary footer rows ────────────────────────────────────────
    FOOTER_ROWS = [
        ('Total Marks',   'total_marks'),
        ('Average Marks', 'average_marks'),
        ('Division',      'division'),
        ('Class Position','class_position'),
        ('Stream Position','stream_position'),
    ]

    for label, key in FOOTER_ROWS:
        _cell(ws, current_row, 1, '', align=LEFT)
        _cell(ws, current_row, 2, label, align=LEFT, bold=True)

        for sess_id, start_col in session_col_map.items():
            metrics = metrics_by_session.get(sess_id)
            position = positions_by_session.get(sess_id)
            val = '-'
            if key == 'total_marks' and metrics:
                val = float(metrics.total_marks) if metrics.total_marks else '-'
            elif key == 'average_marks' and metrics:
                val = float(metrics.average_marks) if metrics.average_marks else '-'
            elif key == 'division' and metrics:
                val = metrics.division or '-'
            elif key == 'class_position' and position:
                val = position.class_position or '-'
            elif key == 'stream_position' and position:
                val = position.stream_position or '-'

            # Merge MARKS and GRADE columns for footer rows
            merged = ws.cell(row=current_row, column=start_col, value=val)
            merged.alignment = CENTER
            merged.border = BORDER
            ws.merge_cells(
                start_row=current_row, start_column=start_col,
                end_row=current_row, end_column=start_col + 1,
            )

        # Border remaining aggregate columns
        for c in range(agg_col['TOTAL'], total_cols + 1):
            ws.cell(row=current_row, column=c).border = BORDER

        current_row += 1

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for i in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # Timestamp
    current_row += 1
    ws.cell(
        row=current_row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ).font = Font(italic=True, size=9)

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Full session report — all students in one exam session
# ─────────────────────────────────────────────────────────────────────────────

def export_session_report(exam_session) -> openpyxl.Workbook:
    """
    Build an Excel workbook with all students in an exam session,
    one row per student showing their subject results, metrics, and position.

    All DB queries issued before any Excel writing.
    """


    # ── Load everything in bulk ────────────────────────────────────────────
    subjects = list(
        Subject.objects
        .filter(
            student_results__exam_session=exam_session
        )
        .distinct()
        .order_by('name')
        .values('id', 'name', 'short_name')
    )
    subject_ids = [s['id'] for s in subjects]
    subject_by_id = {s['id']: s for s in subjects}

    # All results for this session in ONE query
    results_qs = list(
        StudentSubjectResult.objects
        .filter(exam_session=exam_session)
        .values('student_id', 'subject_id', 'total_marks', 'grade', 'points')
    )

    # Group: student_id → {subject_id → result_row}
    results_by_student: dict[int, dict] = defaultdict(dict)
    for row in results_qs:
        results_by_student[row['student_id']][row['subject_id']] = row

    # Metrics in ONE query
    metrics_by_student = {
        m.student_id: m
        for m in StudentExamMetrics.objects.filter(exam_session=exam_session)
        .select_related('student')
    }

    # Positions in ONE query
    positions_by_student = {
        p.student_id: p
        for p in StudentExamPosition.objects.filter(exam_session=exam_session)
    }


    student_info = {
        s['id']: s
        for s in Student.objects
        .filter(id__in=metrics_by_student.keys())
        .values('id', 'first_name', 'last_name', 'registration_number')
    }

    # Sort students by class position
    sorted_student_ids = sorted(
        metrics_by_student.keys(),
        key=lambda sid: positions_by_student.get(sid) and
            positions_by_student[sid].class_position or 9999,
    )

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Session Results'

    # Title
    ws.merge_cells(f'A1:{get_column_letter(5 + len(subjects) * 2)}1')
    t = ws['A1']
    t.value = f'EXAM RESULTS — {exam_session.name}'
    t.font = BOLD_14
    t.alignment = CENTER

    ws['A2'] = f"Class: {exam_session.class_level}"
    ws['D2'] = f"Term: {exam_session.term}"
    ws['G2'] = f"Year: {exam_session.academic_year}"

    # ── Headers ────────────────────────────────────────────────────────────
    HR = 4
    static_headers = ['POS', 'STREAM\nPOS', 'REG NO', 'STUDENT NAME']
    for ci, h in enumerate(static_headers, start=1):
        _hcell(ws, HR, ci, h)
        ws.merge_cells(start_row=HR, start_column=ci, end_row=HR + 1, end_column=ci)

    col = len(static_headers) + 1
    subj_col_map: dict[int, int] = {}
    for subj in subjects:
        label = subj['short_name'] or subj['name']
        _hcell(ws, HR, col, label, merge_end_col=col + 1)
        _hcell(ws, HR + 1, col, 'MK')
        _hcell(ws, HR + 1, col + 1, 'GR')
        subj_col_map[subj['id']] = col
        col += 2

    summary_headers = ['TOTAL\nMARKS', 'AVERAGE', 'POINTS', 'DIVISION', 'REMARK']
    sum_col: dict[str, int] = {}
    for h in summary_headers:
        _hcell(ws, HR, col, h)
        ws.merge_cells(start_row=HR, start_column=col, end_row=HR + 1, end_column=col)
        sum_col[h] = col
        col += 1

    # ── Data rows ──────────────────────────────────────────────────────────
    data_row = HR + 2
    for student_id in sorted_student_ids:
        info = student_info.get(student_id, {})
        metrics = metrics_by_student.get(student_id)
        position = positions_by_student.get(student_id)
        subj_results = results_by_student.get(student_id, {})

        full_name = f"{info.get('first_name', '')} {info.get('last_name', '')}".strip()
        class_pos = position.class_position if position else '-'
        stream_pos = position.stream_position if position else '-'

        _cell(ws, data_row, 1, class_pos)
        _cell(ws, data_row, 2, stream_pos)
        _cell(ws, data_row, 3, info.get('registration_number', '-'))
        _cell(ws, data_row, 4, full_name, align=LEFT)

        for subj_id, start_col in subj_col_map.items():
            res = subj_results.get(subj_id, {})
            marks = res.get('total_marks')
            grade = res.get('grade', '-')
            _cell(ws, data_row, start_col, float(marks) if marks is not None else '-')
            _cell(ws, data_row, start_col + 1, grade)

        if metrics:
            remark = 'Pass' if metrics.average_marks and metrics.average_marks >= 40 else 'Fail'
            _cell(ws, data_row, sum_col['TOTAL\nMARKS'],
                  float(metrics.total_marks) if metrics.total_marks else '-')
            _cell(ws, data_row, sum_col['AVERAGE'],
                  float(metrics.average_marks) if metrics.average_marks else '-')
            _cell(ws, data_row, sum_col['POINTS'],
                  float(metrics.total_points) if metrics.total_points else '-')
            _cell(ws, data_row, sum_col['DIVISION'], metrics.division or '-')
            _cell(ws, data_row, sum_col['REMARK'], remark)
        else:
            for h in summary_headers:
                _cell(ws, data_row, sum_col[h], '-')

        data_row += 1

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 28
    for i in range(5, col):
        ws.column_dimensions[get_column_letter(i)].width = 10

    data_row += 1
    ws.cell(
        row=data_row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ).font = Font(italic=True, size=9)

    return wb
