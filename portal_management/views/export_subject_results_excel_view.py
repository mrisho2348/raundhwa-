from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import date

from core.mixins import ManagementRequiredMixin
from core.models import ExamSession, GradingScale, StudentPaperScore, Subject, SubjectExamPaper
from portal_management.views.exam_views import _enrolled_students



# ──────────────────────────────────────────────────────────────
#  Style helpers
# ──────────────────────────────────────────────────────────────

def _thin_border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=10, color='000000', italic=False, name='Arial'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ──────────────────────────────────────────────────────────────
#  Palette constants
# ──────────────────────────────────────────────────────────────

C_BRAND_DARK = '0D3349'
C_BRAND_MID  = '1A5276'
C_ACCENT     = '148F77'
C_HEADER_BG  = '1B4F72'
C_SUBHDR_BG  = '2980B9'
C_STATS_BG   = '1A5276'
C_STATS_ROW  = 'D6EAF8'
C_WHITE      = 'FFFFFF'
C_GOLD       = 'F0B429'

# Cycling colour palette for grade badges and row backgrounds.
# Ordered best → worst so A/highest grade always gets the green entry.
# These are applied in _build_grade_style_maps() based on the order
# returned by GradingScale, not hardcoded per grade letter.
_BADGE_PALETTE = [
    '1E7E34',  # deep green   — best grade
    '1565C0',  # deep blue
    'E65100',  # deep orange
    '6A1E6E',  # deep purple
    'B71C1C',  # deep red
    '004D40',  # teal
    '37474F',  # slate
    '880E4F',  # deep pink
    '4A148C',  # deep violet
    '3E2723',  # deep brown   — worst grade / overflow
]
_ROW_BG_PALETTE = [
    'E8F5E9',  # light green
    'E3F2FD',  # light blue
    'FFF3E0',  # light orange
    'F3E5F5',  # light purple
    'FFEBEE',  # light red
    'E0F2F1',  # light teal
    'ECEFF1',  # light slate
    'FCE4EC',  # light pink
    'EDE7F6',  # light violet
    'EFEBE9',  # light brown
]


# ──────────────────────────────────────────────────────────────
#  Build grade-style maps from GradingScale queryset
# ──────────────────────────────────────────────────────────────

def _build_grade_style_maps(grading_scale_qs):
    """
    Returns three dicts keyed by grade value (e.g. 'A', 'B', 'C' …).
    Everything is derived from the GradingScale queryset — nothing is
    hardcoded.  The queryset should be ordered by -min_mark so the best
    grade (highest threshold) receives the first palette entry.

    Returns:
        grade_badge_fill  : {grade: PatternFill}  — badge cell background
        grade_text_color  : {grade: str}           — badge font hex (always white)
        grade_row_bg      : {grade: str}           — full data-row bg hex
    """
    grade_badge_fill = {}
    grade_text_color = {}
    grade_row_bg     = {}

    for idx, gs in enumerate(grading_scale_qs):
        badge_hex  = _BADGE_PALETTE[idx % len(_BADGE_PALETTE)]
        row_bg_hex = _ROW_BG_PALETTE[idx % len(_ROW_BG_PALETTE)]
        grade_badge_fill[gs.grade] = _fill(badge_hex)
        grade_text_color[gs.grade] = C_WHITE   # white text always readable on dark badge
        grade_row_bg[gs.grade]     = row_bg_hex

    return grade_badge_fill, grade_text_color, grade_row_bg


# ──────────────────────────────────────────────────────────────
#  View
# ──────────────────────────────────────────────────────────────

class ExportSubjectResultsExcelView(ManagementRequiredMixin, View):
    """Export subject results to a richly-formatted Excel report."""

    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)

        papers = SubjectExamPaper.objects.filter(
            exam_session=session, subject=subject
        ).order_by('paper_number')

        if not papers.exists():
            messages.warning(request, f'No papers found for {subject.name}.')
            return redirect(
                'management:subject_results_summary',
                session_pk=session_pk, subject_pk=subject_pk,
            )

        students  = _enrolled_students(session)
        scores    = StudentPaperScore.objects.filter(
            exam_paper__in=papers, student__in=students
        ).select_related('student', 'exam_paper')

        total_max   = sum(float(p.max_marks) for p in papers)
        papers_list = list(papers)   # evaluate queryset once

        # ── Grading scale — single source of truth for ALL grade data ─────
        #
        # Ordered by -min_mark so best grade (highest threshold) comes first.
        # Every grade-related decision in this view — grade calculation,
        # style colours, distribution table rows, pass/fail logic, the
        # reference table at the bottom — is driven entirely from this
        # queryset.  Nothing is hardcoded.
        ed_level      = session.class_level.educational_level
        grading_scale = GradingScale.objects.filter(
            education_level=ed_level
        ).order_by('-min_mark')

        # Evaluate once so we can safely iterate multiple times
        grading_scale_list = list(grading_scale)

        # Style maps keyed by grade value — all derived from GradingScale
        grade_badge_fill, grade_text_color, grade_row_bg = \
            _build_grade_style_maps(grading_scale_list)

        # Grade → description  (from GradingScale.description; blank if unset)
        grade_description = {gs.grade: gs.description or '' for gs in grading_scale_list}

        # Grade → points  (for display in distribution & reference tables)
        grade_points = {gs.grade: gs.points for gs in grading_scale_list}

        # "Pass" grades are those whose GradingScale.points > 0.
        # This replaces any hardcoded set like {'A','B','C','D','S'} and
        # automatically works for any grading system the school configures.
        pass_grades = {gs.grade for gs in grading_scale_list if gs.points > 0}

        # ── Grade calculator ──────────────────────────────────────────────
        def calculate_grade(pct):
            """
            Maps a percentage to a grade using the live GradingScale data.
            Returns the last (worst) grade as a safety fallback if no band
            matches — never raises an exception.
            """
            for gs in grading_scale_list:
                if float(gs.min_mark) <= pct <= float(gs.max_mark):
                    return gs.grade
            # Fallback: lowest-ranked grade (last in desc-min_mark order)
            return grading_scale_list[-1].grade if grading_scale_list else 'F'

        # ── Group scores by student ───────────────────────────────────────
        student_scores = {}
        for score in scores:
            sid = score.student_id
            if sid not in student_scores:
                student_scores[sid] = {
                    'student': score.student, 'scores': {}, 'total': 0,
                }
            student_scores[sid]['scores'][score.exam_paper_id] = float(score.marks)
            student_scores[sid]['total'] += float(score.marks)

        student_results = []
        for data in student_scores.values():
            pct = (data['total'] / total_max * 100) if total_max else 0
            student_results.append({
                'student':     data['student'],
                'total_marks': data['total'],
                'percentage':  round(pct, 1),
                'grade':       calculate_grade(pct),
                'scores':      data['scores'],
            })
        student_results.sort(key=lambda x: x['total_marks'], reverse=True)

        # ── Statistics ────────────────────────────────────────────────────
        marks_list = [r['total_marks'] for r in student_results]
        stats = {
            'total_students':       len(students),
            'students_with_scores': len(student_results),
            'without_scores':       len(students) - len(student_results),
            'average':   sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest':   max(marks_list) if marks_list else 0,
            'lowest':    min(marks_list) if marks_list else 0,
            'completion': (len(student_results) / len(students) * 100) if students else 0,
        }

        from collections import Counter
        grade_dist = Counter(r['grade'] for r in student_results)
        pass_count = sum(1 for r in student_results if r['grade'] in pass_grades)
        pass_rate  = (pass_count / len(student_results) * 100) if student_results else 0

        # ── School meta ───────────────────────────────────────────────────
        school         = getattr(session.class_level, 'school', None)
        school_name    = getattr(settings, 'SCHOOL_NAME', 'SCHOOL NAME NOT SET'),
        school_address = getattr(settings, 'SCHOOL_ADDRESS', ''),
        school_phone   = getattr(settings, 'SCHOOL_PHONE', ''),
        school_email   = getattr(settings, 'SCHOOL_EMAIL', ''),                
        school_motto   = getattr(settings, 'SCHOOL_MOTTO', ''),
        school_reg_no  = getattr(settings, 'SCHOOL_REGISTRATION_NUMBER', ''),

        # ══════════════════════════════════════════════════════════════════
        #  Workbook scaffold
        # ══════════════════════════════════════════════════════════════════
        wb          = openpyxl.Workbook()
        ws          = wb.active
        ws.title    = f'{subject.name[:28]} Results'
        total_cols  = 7 + len(papers_list)
        last_col    = total_cols
        last_letter = get_column_letter(last_col)

        def band(row, value, fill_hex, font_obj, align_obj=None, height=None):
            """Merge the full row width, apply fill/font/alignment, set height."""
            ws.merge_cells(f'A{row}:{last_letter}{row}')
            c           = ws.cell(row=row, column=1, value=value)
            c.fill      = _fill(fill_hex)
            c.font      = font_obj
            c.alignment = align_obj or _align('center', 'center')
            if height:
                ws.row_dimensions[row].height = height
            return c

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 1 – SCHOOL HEADER  (rows 1-7)
        # ══════════════════════════════════════════════════════════════════

        band(1, '', C_BRAND_DARK, _font(), height=8)

        band(2, school_name.upper(), C_BRAND_DARK,
             _font(bold=True, size=18, color=C_GOLD),
             _align('center', 'center'), height=38)

        contact_parts = [p for p in [school_address, school_phone, school_email] if p]
        band(3, '   |   '.join(contact_parts) if contact_parts else '',
             C_BRAND_MID, _font(size=9, color='D6EAF8', italic=True),
             _align('center', 'center'), height=18)

        motto_line = ''
        if school_motto:  motto_line += f'"{school_motto}"'
        if school_reg_no: motto_line += f'   Reg No: {school_reg_no}'
        band(4, motto_line, C_BRAND_MID,
             _font(size=9, color='AED6F1', italic=bool(school_motto)),
             _align('center', 'center'), height=18)

        band(5, '', C_GOLD, _font(), height=4)

        band(6, 'SUBJECT EXAMINATION RESULTS REPORT', C_ACCENT,
             _font(bold=True, size=13, color=C_WHITE),
             _align('center', 'center'), height=28)

        band(7, f'{subject.name.upper()}  ({subject.code})', '0E6655',
             _font(bold=True, size=11, color='A9DFBF'),
             _align('center', 'center'), height=22)

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 2 – SESSION INFO  (rows 8-10)
        # ══════════════════════════════════════════════════════════════════

        band(8,
             f'Session: {session.name}   │   Class: {session.class_level.name}'
             f'   │   Term: {session.term}   │   Academic Year: {session.academic_year}',
             'D6EAF8', _font(bold=True, size=9, color='1A5276'),
             _align('center', 'center'), height=20)

        band(9,
             f'Education Level: {ed_level}'
             f'   │   Max Marks per Student: {total_max:.0f}'
             f'   │   Papers: {len(papers_list)}',
             'EBF5FB', _font(size=9, color='1A5276'),
             _align('center', 'center'), height=18)

        generated_by = request.user.get_full_name() or request.user.username
        band(10,
             f'Generated: {timezone.now().strftime("%A, %d %B %Y  %H:%M")}   │   By: {generated_by}',
             'F2F3F4', _font(size=8, color='555555', italic=True),
             _align('center', 'center'), height=16)

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 3 – QUICK-STATS BAR  (rows 11-12)
        # ══════════════════════════════════════════════════════════════════

        band(11, 'QUICK STATISTICS', C_HEADER_BG,
             _font(bold=True, size=9, color='AED6F1'),
             _align('center', 'center'), height=16)

        ws.row_dimensions[12].height = 22
        qs_items = [
            ('STUDENTS',   str(stats['total_students'])),
            ('SCORED',     str(stats['students_with_scores'])),
            ('AVG MARK',   f"{stats['average']:.1f}"),
            ('HIGHEST',    f"{stats['highest']:.0f}"),
            ('LOWEST',     f"{stats['lowest']:.0f}"),
            ('PASS RATE',  f"{pass_rate:.1f}%"),
            ('COMPLETION', f"{stats['completion']:.1f}%"),
        ]
        chunk = max(1, total_cols // len(qs_items))
        for i, (label, val) in enumerate(qs_items):
            col_s = i * chunk + 1
            col_e = (col_s + chunk - 1) if i < len(qs_items) - 1 else last_col
            col_e = min(col_e, last_col)
            if col_s > last_col:
                break
            if col_s < col_e:
                ws.merge_cells(start_row=12, start_column=col_s,
                               end_row=12,   end_column=col_e)
            c           = ws.cell(row=12, column=col_s, value=f'{label}: {val}')
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_SUBHDR_BG if i % 2 == 0 else C_ACCENT)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border(C_BRAND_DARK)

        band(13, '', 'D5D8DC', _font(), height=4)

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 4 – COLUMN HEADERS  (rows 14-15)
        # ══════════════════════════════════════════════════════════════════

        ws.row_dimensions[14].height = 16
        for col in range(1, 8):
            c        = ws.cell(row=14, column=col, value='')
            c.fill   = _fill(C_BRAND_DARK)
            c.border = _thin_border(C_BRAND_MID)

        for idx, paper in enumerate(papers_list, 8):
            c           = ws.cell(row=14, column=idx,
                                   value=f'Max: {paper.max_marks:.0f}')
            c.font      = _font(size=8, color='AED6F1', italic=True)
            c.fill      = _fill(C_BRAND_DARK)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border(C_BRAND_MID)

        headers = ['#', 'REG. NUMBER', 'STUDENT NAME', 'GENDER',
                   'TOTAL MARKS', 'PERCENTAGE', 'GRADE']
        for paper in papers_list:
            headers.append(f'PAPER {paper.paper_number}')

        ws.row_dimensions[15].height = 28
        for col, header in enumerate(headers, 1):
            c           = ws.cell(row=15, column=col, value=header)
            c.font      = _font(bold=True, size=10, color=C_WHITE)
            c.fill      = _fill(C_HEADER_BG)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border(C_BRAND_DARK)

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 5 – DATA ROWS  (from row 16)
        # ══════════════════════════════════════════════════════════════════

        data_start_row = 16
        for row_idx, result in enumerate(student_results, data_start_row):
            student = result['student']
            grade   = result['grade']
            row_bg  = grade_row_bg.get(grade, 'FAFCFE')
            ws.row_dimensions[row_idx].height = 20

            row_data = [
                row_idx - data_start_row + 1,
                student.registration_number or 'N/A',
                student.full_name,
                student.get_gender_display() if student.gender else 'N/A',
                result['total_marks'],
                result['percentage'],
                grade,
            ]
            for paper in papers_list:
                row_data.append(result['scores'].get(paper.pk))

            col_aligns = ['center', 'left', 'left', 'center',
                          'center', 'center', 'center']
            col_aligns += ['center'] * len(papers_list)

            for col, (val, align) in enumerate(zip(row_data, col_aligns), 1):
                c = ws.cell(row=row_idx, column=col)

                if col == 7:
                    # Grade badge — colours come from grade_badge_fill which
                    # was built entirely from GradingScale, not hardcoded
                    c.value      = val
                    c.font       = _font(bold=True, size=10,
                                         color=grade_text_color.get(grade, C_WHITE))
                    c.fill       = grade_badge_fill.get(grade, _fill('888888'))
                    c.alignment  = _align('center', 'center')

                elif col >= 8:
                    # Paper score — colour by % of that paper's max marks
                    paper = papers_list[col - 8]
                    if val is None:
                        c.value      = '–'
                        c.font       = _font(size=9, color='AAAAAA', italic=True)
                        c.fill       = _fill('F0F0F0')
                        c.alignment  = _align('center', 'center')
                    else:
                        pmax      = float(paper.max_marks)
                        pct_paper = (val / pmax * 100) if pmax else 0
                        if pct_paper >= 75:
                            sc = '1B5E20'
                        elif pct_paper >= 50:
                            sc = '1A237E'
                        elif pct_paper >= 30:
                            sc = 'E65100'
                        else:
                            sc = 'B71C1C'
                        c.value         = val
                        c.number_format = '0.0'
                        c.font          = _font(bold=True, size=9, color=sc)
                        c.fill          = _fill(row_bg)
                        c.alignment     = _align('center', 'center')

                elif col == 5:
                    c.value         = val
                    c.number_format = '0.0'
                    c.font          = _font(bold=True, size=10, color='1A5276')
                    c.fill          = _fill(row_bg)
                    c.alignment     = _align('center', 'center')

                elif col == 6:
                    c.value         = val
                    c.number_format = '0.0'
                    c.font          = _font(size=9, color='1A5276')
                    c.fill          = _fill(row_bg)
                    c.alignment     = _align('center', 'center')

                else:
                    c.value      = val
                    c.font       = _font(size=9)
                    c.fill       = _fill(row_bg)
                    c.alignment  = _align(align, 'center')

                c.border = _thin_border()

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 6 – GRADE DISTRIBUTION TABLE
        #  Rows generated by iterating grading_scale_list — no grade
        #  list hardcoded anywhere.
        # ══════════════════════════════════════════════════════════════════

        last_data_row = data_start_row + len(student_results) - 1
        dist_start    = last_data_row + 3

        ws.merge_cells(f'A{dist_start}:{last_letter}{dist_start}')
        c           = ws.cell(row=dist_start, column=1, value='GRADE DISTRIBUTION')
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_STATS_BG)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[dist_start].height = 24

        dist_hdr_row = dist_start + 1
        ws.row_dimensions[dist_hdr_row].height = 20
        for ci, lbl in enumerate(
            ['GRADE', 'RANGE (%)', 'POINTS', 'COUNT', 'PERCENTAGE', 'DESCRIPTION', 'BAR'], 1
        ):
            c           = ws.cell(row=dist_hdr_row, column=ci, value=lbl)
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_ACCENT)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border()

        n_scored = len(student_results) or 1
        for ri, gs in enumerate(grading_scale_list, dist_hdr_row + 1):
            g     = gs.grade
            cnt   = grade_dist.get(g, 0)
            pct_g = cnt / n_scored * 100
            bar   = '█' * max(1, int(pct_g / 5)) if cnt else ''

            row_vals = [
                g,
                f'{gs.min_mark:.0f} – {gs.max_mark:.0f}',
                f'{gs.points:.1f}',
                cnt,
                f'{pct_g:.1f}%',
                grade_description.get(g, ''),
                bar,
            ]
            for ci, val in enumerate(row_vals, 1):
                c           = ws.cell(row=ri, column=ci, value=val)
                c.fill      = grade_badge_fill.get(g, _fill('DDDDDD')) if ci == 1 \
                               else _fill('F7F9FC')
                c.font      = _font(
                    bold=(ci == 1), size=9,
                    color=grade_text_color.get(g, '000000') if ci == 1 else '1A252F',
                )
                c.alignment = _align('center', 'center')
                c.border    = _thin_border()
            ws.row_dimensions[ri].height = 18

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 7 – COMPREHENSIVE SUMMARY STATISTICS
        # ══════════════════════════════════════════════════════════════════

        # Number of data rows written = len(grading_scale_list)
        sum_start = dist_hdr_row + len(grading_scale_list) + 3

        ws.merge_cells(f'A{sum_start}:{last_letter}{sum_start}')
        c           = ws.cell(row=sum_start, column=1,
                               value='COMPREHENSIVE SUMMARY STATISTICS')
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_STATS_BG)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[sum_start].height = 24

        # Build the pass-grades label from the live set — not hardcoded
        pass_grade_label = ', '.join(sorted(pass_grades)) if pass_grades else 'N/A'

        stat_sections = [
            ('ENROLMENT', [
                ('Total Enrolled Students',  stats['total_students']),
                ('Students with Scores',     stats['students_with_scores']),
                ('Students without Scores',  stats['without_scores']),
                ('Completion Rate',          f"{stats['completion']:.1f}%"),
            ]),
            ('PERFORMANCE', [
                ('Average Score',            f"{stats['average']:.1f}"),
                ('Highest Score',            f"{stats['highest']:.0f}"),
                ('Lowest Score',             f"{stats['lowest']:.0f}"),
                ('Score Range',              f"{stats['highest'] - stats['lowest']:.0f}"),
            ]),
            ('PASS / FAIL', [
                ('Total Passed',             pass_count),
                ('Total Failed',             len(student_results) - pass_count),
                ('Overall Pass Rate',        f"{pass_rate:.1f}%"),
                (f'Pass Grades ({pass_grade_label})', 'Grades with Points > 0'),
            ]),
            ('EXAM CONFIGURATION', [
                ('Max Marks per Student',    f"{total_max:.0f}"),
                ('Number of Papers',         len(papers_list)),
                ('Grading System',           str(ed_level)),
                ('Grade Bands Configured',   len(grading_scale_list)),
            ]),
        ]

        cur_row = sum_start + 1
        half    = last_col // 2

        for section_label, rows in stat_sections:
            ws.merge_cells(f'A{cur_row}:{last_letter}{cur_row}')
            c           = ws.cell(row=cur_row, column=1, value=section_label)
            c.font      = _font(bold=True, size=9, color=C_BRAND_DARK)
            c.fill      = _fill('AED6F1')
            c.alignment = _align('left', 'center')
            c.border    = _thin_border()
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

            for label, value in rows:
                ws.merge_cells(f'A{cur_row}:{get_column_letter(half)}{cur_row}')
                cl           = ws.cell(row=cur_row, column=1, value=label)
                cl.font      = _font(size=9, color='1A252F')
                cl.fill      = _fill(C_STATS_ROW if cur_row % 2 == 0 else 'EBF5FB')
                cl.alignment = _align('left', 'center')
                cl.border    = _thin_border()

                ws.merge_cells(
                    f'{get_column_letter(half+1)}{cur_row}:{last_letter}{cur_row}'
                )
                cv           = ws.cell(row=cur_row, column=half + 1, value=value)
                cv.font      = _font(bold=True, size=9, color=C_BRAND_DARK)
                cv.fill      = _fill(C_STATS_ROW if cur_row % 2 == 0 else 'EBF5FB')
                cv.alignment = _align('center', 'center')
                cv.border    = _thin_border()
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 8 – GRADING SCALE REFERENCE TABLE
        #  Fully generated from GradingScale — zero hardcoded values
        # ══════════════════════════════════════════════════════════════════

        ref_start = cur_row + 2
        ws.merge_cells(f'A{ref_start}:{last_letter}{ref_start}')
        c           = ws.cell(row=ref_start, column=1,
                               value=f'GRADING SCALE REFERENCE  –  {ed_level}')
        c.font      = _font(bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_BRAND_MID)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[ref_start].height = 22

        ref_hdr = ref_start + 1
        for ci, lbl in enumerate(
            ['GRADE', 'MIN %', 'MAX %', 'POINTS', 'DESCRIPTION'], 1
        ):
            c           = ws.cell(row=ref_hdr, column=ci, value=lbl)
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_HEADER_BG)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border()
        ws.row_dimensions[ref_hdr].height = 18

        for ri, gs in enumerate(grading_scale_list, ref_hdr + 1):
            for ci, val in enumerate(
                [gs.grade, f'{gs.min_mark:.0f}', f'{gs.max_mark:.0f}',
                 f'{gs.points:.1f}', gs.description or ''], 1
            ):
                c           = ws.cell(row=ri, column=ci, value=val)
                c.fill      = grade_badge_fill.get(gs.grade, _fill('EEEEEE')) if ci == 1 \
                               else _fill('F7F9FC')
                c.font      = _font(
                    bold=(ci == 1), size=9,
                    color=grade_text_color.get(gs.grade, '000000') if ci == 1 else '1A252F',
                )
                c.alignment = _align('center', 'center')
                c.border    = _thin_border()
            ws.row_dimensions[ri].height = 16

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 9 – FOOTER
        # ══════════════════════════════════════════════════════════════════

        footer_row = ref_hdr + len(grading_scale_list) + 2
        ws.merge_cells(f'A{footer_row}:{last_letter}{footer_row}')
        c           = ws.cell(
            row=footer_row, column=1,
            value=(
                f'This report is computer-generated and is valid without a signature.  '
                f'© {date.today().year} {school_name}  │  '
                f'Printed: {timezone.now().strftime("%d %b %Y %H:%M")}'
            )
        )
        c.font      = _font(size=8, color='888888', italic=True)
        c.fill      = _fill('F2F3F4')
        c.alignment = _align('center', 'center')
        ws.row_dimensions[footer_row].height = 16

        ws.merge_cells(f'A{footer_row+1}:{last_letter}{footer_row+1}')
        ws.cell(row=footer_row + 1, column=1).fill = _fill(C_GOLD)
        ws.row_dimensions[footer_row + 1].height   = 5

        # ══════════════════════════════════════════════════════════════════
        #  COLUMN WIDTHS
        # ══════════════════════════════════════════════════════════════════

        for letter, width in [('A', 5), ('B', 20), ('C', 28), ('D', 10),
                               ('E', 14), ('F', 13), ('G', 9)]:
            ws.column_dimensions[letter].width = width
        for idx in range(8, 8 + len(papers_list)):
            ws.column_dimensions[get_column_letter(idx)].width = 13

        # ══════════════════════════════════════════════════════════════════
        #  FREEZE PANES & PRINT SETTINGS
        # ══════════════════════════════════════════════════════════════════

        ws.freeze_panes                               = f'A{data_start_row}'
        ws.print_title_rows                           = f'1:{data_start_row - 1}'
        ws.page_setup.orientation                     = 'landscape'
        ws.page_setup.paperSize                       = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage                       = True
        ws.page_setup.fitToWidth                      = 1
        ws.page_setup.fitToHeight                     = 0
        ws.sheet_properties.pageSetUpPr.fitToPage     = True

        # ══════════════════════════════════════════════════════════════════
        #  HTTP RESPONSE
        # ══════════════════════════════════════════════════════════════════

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_name = (
            f'{subject.code}_{session.name}_Results_{date.today()}'
            .replace(' ', '_').replace('/', '-')
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
        wb.save(response)
        return response
