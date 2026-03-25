"""
portal_management/views/exams.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Complete Examination & Result System views.

Sections:
  ── EXAM TYPE CRUD ─────── ExamTypeListView, ExamTypeCreateView,
                             ExamTypeUpdateView, ExamTypeDeleteView,
                             ExamTypeSearchView
  ── EXAM SESSION CRUD ──── ExamSessionListView, ExamSessionCreateView,
                             ExamSessionDetailView, ExamSessionUpdateView,
                             ExamSessionDeleteView
  ── EXAM PAPERS ──────────  SubjectExamPaperCreateView, SubjectExamPaperUpdateView,
                             SubjectExamPaperDeleteView, GetPapersForSubjectView
  ── RESULT ENTRY ────────── SessionResultsView (student × subject grid),
                             StudentResultEntryView (per-student per-subject marks),
                             SavePaperScoreView (AJAX single score save),
                             BulkSaveScoresView (AJAX save all scores for a student)
  ── EXCEL UPLOAD/DOWNLOAD ─ DownloadResultTemplateView (subject-wise prefilled),
                             DownloadFullTemplateView (all subjects),
                             UploadResultsView (parse + save uploaded Excel)
  ── CALCULATION ────────── CalculateSubjectResultsView, CalculateMetricsView,
                             CalculatePositionsView, CalculateFullResultsView
  ── WORKFLOW ─────────────  SubmitSessionView, VerifySessionView,
                             PublishSessionView, UnpublishSessionView
  ── EXPORT ───────────────  ExportSessionReportView, ExportSubjectReportView
  ── GRADING ──────────────  GradingScaleListView, GradingScaleCreateView,
                             GradingScaleDeleteView, DivisionScaleCreateView,
                             DivisionScaleDeleteView
  ── AJAX ─────────────────  TermsByAcademicYearView, StreamsByClassLevelView,
                             SubjectsByClassLevelView
"""

import io
import json
import logging
from datetime import date
from decimal import Decimal, InvalidOperation
from statistics import mean, median

from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from django.template.loader import render_to_string
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import View

from core.mixins import ManagementRequiredMixin
from core.models import (
    AcademicYear, ClassLevel, DivisionScale, EducationalLevel,
    ExamSession, ExamType, GradingScale, SchoolProfile, StreamClass, Student,
    StudentEnrollment, StudentExamMetrics, StudentExamPosition,
    StudentPaperScore, StudentSubjectResult, Subject,
    SubjectExamPaper, Term,
)

logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _is_ajax(request):
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def _terms_by_year():
    """Return dict of academic_year_id -> list of term dicts (for JS)."""
    result = {}
    for t in Term.objects.select_related('academic_year').order_by('term_number'):
        result.setdefault(t.academic_year_id, []).append({
            'id': t.pk,
            'name': f'Term {t.term_number}',
            'term_number': t.term_number,
            'start_date': t.start_date.strftime('%Y-%m-%d'),
            'end_date': t.end_date.strftime('%Y-%m-%d'),
        })
    return result


def _session_form_context():
    """Common context for create/edit session forms."""
    return {
        'exam_types': ExamType.objects.order_by('name'),
        'academic_years': AcademicYear.objects.order_by('-start_date'),
        'class_levels': ClassLevel.objects.select_related(
            'educational_level'
        ).order_by('educational_level', 'order'),
        'stream_classes': StreamClass.objects.select_related(
            'class_level'
        ).order_by('class_level', 'stream_letter'),
        'status_choices': ExamSession.STATUS_CHOICES,
        'terms_data': json.dumps(_terms_by_year()),
    }


def _enrolled_students(session):
    """Return students enrolled for this session's class level & year."""
    qs = StudentEnrollment.objects.filter(
        academic_year=session.academic_year,
        class_level=session.class_level,
        status='active',
    ).select_related('student')
    if session.stream_class_id:
        qs = qs.filter(
            stream_assignment__stream_class=session.stream_class
        )
    return [e.student for e in qs.order_by('student__last_name', 'student__first_name')]


def _xl_header_style(ws, row, col, value, bg='1a73e8', fg='FFFFFF', bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return cell


def _xl_data_style(ws, row, col, value='', locked=False, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if locked:
        cell.fill = PatternFill('solid', fgColor='F0F0F0')
    return cell


# ════════════════════════════════════════════════════════════════════════════
# EXAM TYPE CRUD
# ════════════════════════════════════════════════════════════════════════════

class ExamTypeListView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_type_list.html'

    def get(self, request):
        qs = ExamType.objects.annotate(
            session_count=Count('exam_sessions')
        ).order_by('name')
        search = request.GET.get('search', '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(code__icontains=search))
        return render(request, self.template_name, {
            'exam_types': qs,
            'search_query': search,
        })


class ExamTypeCreateView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_type_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'title': 'Create Exam Type', 'is_edit': False,
        })

    def post(self, request):
        ajax = _is_ajax(request)
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        errors = {}

        if not name:
            errors['name'] = ['Name is required.']
        if not code:
            errors['code'] = ['Code is required.']
        elif ExamType.objects.filter(code=code).exists():
            errors['code'] = [f'Code "{code}" already exists.']

        try:
            weight = Decimal(request.POST.get('weight', '0'))
            if not (0 <= weight <= 100):
                errors['weight'] = ['Weight must be 0–100.']
        except InvalidOperation:
            errors['weight'] = ['Invalid weight.']
            weight = Decimal('0')

        try:
            max_score = Decimal(request.POST.get('max_score', '100'))
            if max_score < 1:
                errors['max_score'] = ['Max score must be ≥ 1.']
        except InvalidOperation:
            errors['max_score'] = ['Invalid max score.']
            max_score = Decimal('100')

        if errors:
            if ajax:
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            return render(request, self.template_name, {
                'title': 'Create Exam Type', 'is_edit': False,
                'form_data': request.POST, 'errors': errors,
            })

        try:
            with transaction.atomic():
                et = ExamType.objects.create(
                    name=name, code=code, weight=weight,
                    max_score=max_score,
                    description=request.POST.get('description', '').strip(),
                )
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Exam type "{name}" created.',
                    'redirect_url': reverse('management:exam_type_list'),
                })
            messages.success(request, f'Exam type "{name}" created.')
            return redirect('management:exam_type_list')
        except Exception as e:
            logger.error('ExamTypeCreate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return render(request, self.template_name, {
                'title': 'Create Exam Type', 'is_edit': False,
                'form_data': request.POST,
            })


class ExamTypeUpdateView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_type_form.html'

    def get(self, request, pk):
        et = get_object_or_404(ExamType, pk=pk)
        return render(request, self.template_name, {
            'title': f'Edit — {et.name}', 'is_edit': True, 'exam_type': et,
        })

    def post(self, request, pk):
        ajax = _is_ajax(request)
        et = get_object_or_404(ExamType, pk=pk)
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        errors = {}

        if not name:
            errors['name'] = ['Name is required.']
        if not code:
            errors['code'] = ['Code is required.']
        elif code != et.code and ExamType.objects.filter(code=code).exists():
            errors['code'] = [f'Code "{code}" already exists.']

        try:
            weight = Decimal(request.POST.get('weight', '0'))
        except InvalidOperation:
            errors['weight'] = ['Invalid weight.']
            weight = Decimal('0')
        try:
            max_score = Decimal(request.POST.get('max_score', '100'))
        except InvalidOperation:
            errors['max_score'] = ['Invalid max score.']
            max_score = Decimal('100')

        if errors:
            if ajax:
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            return render(request, self.template_name, {
                'title': f'Edit — {et.name}', 'is_edit': True, 'exam_type': et,
                'form_data': request.POST, 'errors': errors,
            })

        try:
            with transaction.atomic():
                et.name = name
                et.code = code
                et.weight = weight
                et.max_score = max_score
                et.description = request.POST.get('description', '').strip()
                et.save()
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Exam type "{name}" updated.',
                    'redirect_url': reverse('management:exam_type_list'),
                })
            messages.success(request, f'Exam type "{name}" updated.')
            return redirect('management:exam_type_list')
        except Exception as e:
            logger.error('ExamTypeUpdate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return render(request, self.template_name, {
                'title': f'Edit — {et.name}', 'is_edit': True, 'exam_type': et,
                'form_data': request.POST,
            })


class ExamTypeDeleteView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        et = get_object_or_404(ExamType, pk=pk)
        count = ExamSession.objects.filter(exam_type=et).count()
        if count:
            msg = f'Cannot delete — {count} exam session(s) use this type.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_type_list')
        name = et.name
        et.delete()
        if ajax:
            return JsonResponse({'success': True, 'message': f'"{name}" deleted.'})
        messages.success(request, f'Exam type "{name}" deleted.')
        return redirect('management:exam_type_list')


class ExamTypeSearchView(ManagementRequiredMixin, View):
    def get(self, request):
        term = request.GET.get('term', '').strip()
        qs = ExamType.objects.order_by('name')
        if term:
            qs = qs.filter(Q(name__icontains=term) | Q(code__icontains=term))
        return JsonResponse({
            'results': [
                {'id': et.pk, 'text': f'{et.name} ({et.code})',
                 'name': et.name, 'code': et.code,
                 'weight': float(et.weight), 'max_score': float(et.max_score)}
                for et in qs[:20]
            ]
        })


# ════════════════════════════════════════════════════════════════════════════
# EXAM SESSION CRUD
# ════════════════════════════════════════════════════════════════════════════

class ExamSessionListView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_session_list.html'

    def get(self, request):
        qs = ExamSession.objects.select_related(
            'exam_type', 'academic_year', 'term', 'class_level', 'stream_class'
        ).order_by('-exam_date')

        if s := request.GET.get('search', '').strip():
            qs = qs.filter(
                Q(name__icontains=s) | Q(exam_type__name__icontains=s) |
                Q(class_level__name__icontains=s)
            )
        if v := request.GET.get('status'):
            qs = qs.filter(status=v)
        if v := request.GET.get('academic_year'):
            qs = qs.filter(academic_year_id=v)
        if v := request.GET.get('class_level'):
            qs = qs.filter(class_level_id=v)

        return render(request, self.template_name, {
            'sessions': qs,
            'academic_years': AcademicYear.objects.order_by('-start_date'),
            'class_levels': ClassLevel.objects.select_related(
                'educational_level'
            ).order_by('educational_level', 'order'),
            'status_choices': ExamSession.STATUS_CHOICES,
            'search_query': request.GET.get('search', ''),
            'selected_status': request.GET.get('status', ''),
            'selected_academic_year': request.GET.get('academic_year', ''),
            'selected_class_level': request.GET.get('class_level', ''),
            # Status counts for summary pills
            'draft_count':     ExamSession.objects.filter(status='draft').count(),
            'submitted_count': ExamSession.objects.filter(status='submitted').count(),
            'verified_count':  ExamSession.objects.filter(status='verified').count(),
            'published_count': ExamSession.objects.filter(status='published').count(),
        })


class ExamSessionCreateView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_session_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'title': 'Create Exam Session', 'is_edit': False,
            **_session_form_context(),
        })

    def post(self, request):
        ajax = _is_ajax(request)
        try:
            with transaction.atomic():
                session = ExamSession(
                    name=request.POST.get('name', '').strip(),
                    exam_type_id=request.POST.get('exam_type'),
                    academic_year_id=request.POST.get('academic_year'),
                    term_id=request.POST.get('term'),
                    class_level_id=request.POST.get('class_level'),
                    stream_class_id=request.POST.get('stream_class') or None,
                    exam_date=request.POST.get('exam_date'),
                    status=request.POST.get('status', 'draft'),
                )
                session.full_clean()
                session.save()
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Session "{session.name}" created.',
                    'redirect_url': reverse('management:exam_session_detail',
                                            args=[session.pk]),
                })
            messages.success(request, f'Session "{session.name}" created.')
            return redirect('management:exam_session_detail', pk=session.pk)
        except ValidationError as e:
            errs = e.message_dict if hasattr(e, 'message_dict') else {'__all__': [str(e)]}
            if ajax:
                return JsonResponse({'success': False, 'errors': errs}, status=400)
            return render(request, self.template_name, {
                'title': 'Create Exam Session', 'is_edit': False,
                'form_data': request.POST, 'errors': errs,
                **_session_form_context(),
            })
        except Exception as e:
            logger.error('ExamSessionCreate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return render(request, self.template_name, {
                'title': 'Create Exam Session', 'is_edit': False,
                'form_data': request.POST, **_session_form_context(),
            })


class ExamSessionDetailView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_session_detail.html'

    def get(self, request, pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type', 'academic_year', 'term',
                'class_level__educational_level', 'stream_class'
            ), pk=pk
        )
        
        # Get papers
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        # Group papers by subject
        papers_by_subject = {}
        for p in papers:
            papers_by_subject.setdefault(p.subject, []).append(p)
        
        # Get enrolled students
        enrolled_students = _enrolled_students(session)
        enrolled_count = len(enrolled_students)
        
        # Get students with scores
        students_with_scores = StudentSubjectResult.objects.filter(
            exam_session=session
        ).values_list('student_id', flat=True).distinct()
        scored_count = len(set(students_with_scores))
        
        # Count metrics
        metrics_count = StudentExamMetrics.objects.filter(exam_session=session).count()
        results_count = StudentSubjectResult.objects.filter(exam_session=session).count()
        
        # Check status
        has_results = StudentSubjectResult.objects.filter(exam_session=session).exists()
        has_metrics = StudentExamMetrics.objects.filter(exam_session=session).exists()
        has_positions = StudentExamPosition.objects.filter(exam_session=session).exists()
        
        # Determine if upload is allowed
        can_upload = session.status != 'published' and papers.exists()
        can_edit = session.status != 'published'
        
        # Get available subjects for adding papers
        available_subjects = Subject.objects.filter(
            educational_level=session.class_level.educational_level
        ).order_by('name')
        
        # Calculate progress percentage
        progress_percent = 0
        if enrolled_count > 0:
            progress_percent = int((scored_count / enrolled_count) * 100)
        
        return render(request, self.template_name, {
            'session': session,
            'papers': papers,
            'papers_by_subject': papers_by_subject,
            'enrolled_count': enrolled_count,
            'scored_count': scored_count,
            'paper_count': papers.count(),
            'subject_count': len(papers_by_subject),
            'results_count': results_count,
            'metrics_count': metrics_count,
            'has_results': has_results,
            'has_metrics': has_metrics,
            'has_positions': has_positions,
            'status_choices': ExamSession.STATUS_CHOICES,
            'available_subjects': available_subjects,
            'can_upload': can_upload,
            'can_edit': can_edit,
            'progress_percent': progress_percent,
        })
    

class ExamSessionUpdateView(ManagementRequiredMixin, View):
    template_name = 'portal_management/exams/exam_session_form.html'

    def get(self, request, pk):
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            messages.warning(request, 'Published sessions cannot be edited.')
            return redirect('management:exam_session_detail', pk=pk)
        return render(request, self.template_name, {
            'title': f'Edit — {session.name}', 'is_edit': True,
            'session': session, **_session_form_context(),
        })

    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            msg = 'Published sessions cannot be edited.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        try:
            with transaction.atomic():
                session.name = request.POST.get('name', '').strip()
                session.exam_type_id = request.POST.get('exam_type')
                session.academic_year_id = request.POST.get('academic_year')
                session.term_id = request.POST.get('term')
                session.class_level_id = request.POST.get('class_level')
                session.stream_class_id = request.POST.get('stream_class') or None
                session.exam_date = request.POST.get('exam_date')
                session.status = request.POST.get('status', session.status)
                session.full_clean()
                session.save()
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Session "{session.name}" updated.',
                    'redirect_url': reverse('management:exam_session_detail',
                                            args=[session.pk]),
                })
            messages.success(request, f'Session updated.')
            return redirect('management:exam_session_detail', pk=pk)
        except ValidationError as e:
            errs = e.message_dict if hasattr(e, 'message_dict') else {'__all__': [str(e)]}
            if ajax:
                return JsonResponse({'success': False, 'errors': errs}, status=400)
            return render(request, self.template_name, {
                'title': f'Edit — {session.name}', 'is_edit': True,
                'session': session, 'form_data': request.POST,
                'errors': errs, **_session_form_context(),
            })
        except Exception as e:
            logger.error('ExamSessionUpdate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return redirect('management:exam_session_detail', pk=pk)


class ExamSessionDeleteView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        deps = []
        pc = SubjectExamPaper.objects.filter(exam_session=session).count()
        rc = StudentSubjectResult.objects.filter(exam_session=session).count()
        if pc:
            deps.append(f'{pc} exam paper(s)')
        if rc:
            deps.append(f'{rc} result(s)')
        if deps:
            msg = f'Cannot delete — has {", ".join(deps)}.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        name = session.name
        session.delete()
        if ajax:
            return JsonResponse({'success': True, 'message': f'"{name}" deleted.'})
        messages.success(request, f'Session "{name}" deleted.')
        return redirect('management:exam_session_list')


# ════════════════════════════════════════════════════════════════════════════
# EXAM PAPERS (SubjectExamPaper CRUD)
# ════════════════════════════════════════════════════════════════════════════

class SubjectExamPaperCreateView(ManagementRequiredMixin, View):
    """Add a paper to an exam session. A subject can have multiple papers (P1, P2, P3)."""

    def post(self, request, session_pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=session_pk)
        if session.status == 'published':
            msg = 'Cannot add papers to a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        try:
            with transaction.atomic():
                paper = SubjectExamPaper(
                    exam_session=session,
                    subject_id=request.POST.get('subject'),
                    paper_number=request.POST.get('paper_number', 1),
                    paper_name=request.POST.get('paper_name', '').strip(),
                    max_marks=Decimal(request.POST.get('max_marks', '100')),
                    exam_date=request.POST.get('exam_date') or None,
                    duration_minutes=request.POST.get('duration_minutes') or None,
                )
                paper.full_clean()
                paper.save()
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Paper added: {paper}',
                    'paper': {
                        'id': paper.pk,
                        'subject': str(paper.subject),
                        'paper_number': paper.paper_number,
                        'paper_name': paper.paper_name,
                        'max_marks': float(paper.max_marks),
                    }
                })
            messages.success(request, f'Paper "{paper}" added.')
        except ValidationError as e:
            msg = '; '.join(
                sum(e.message_dict.values(), [])
                if hasattr(e, 'message_dict') else [str(e)]
            )
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
        except Exception as e:
            logger.error('ExamPaperCreate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=session_pk)


class SubjectExamPaperUpdateView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        paper = get_object_or_404(SubjectExamPaper, pk=pk)
        session = paper.exam_session
        if session.status == 'published':
            msg = 'Cannot edit papers in a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session.pk)
        try:
            with transaction.atomic():
                paper.paper_name = request.POST.get('paper_name', '').strip()
                paper.max_marks = Decimal(request.POST.get('max_marks', paper.max_marks))
                paper.exam_date = request.POST.get('exam_date') or None
                paper.duration_minutes = request.POST.get('duration_minutes') or None
                paper.full_clean()
                paper.save()
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Paper updated.',
                    'paper': {
                        'id': paper.pk,
                        'paper_name': paper.paper_name,
                        'max_marks': float(paper.max_marks),
                    }
                })
            messages.success(request, 'Paper updated.')
        except Exception as e:
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=session.pk)


class SubjectExamPaperDeleteView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        paper = get_object_or_404(SubjectExamPaper, pk=pk)
        session_pk = paper.exam_session_id
        if paper.exam_session.status == 'published':
            msg = 'Cannot delete papers in a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        # Check for existing scores
        score_count = StudentPaperScore.objects.filter(exam_paper=paper).count()
        if score_count:
            msg = f'Cannot delete — {score_count} score(s) exist for this paper.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        paper.delete()
        if ajax:
            return JsonResponse({'success': True, 'message': 'Paper deleted.'})
        messages.success(request, 'Paper deleted.')
        return redirect('management:exam_session_detail', pk=session_pk)


class GetPapersForSubjectView(ManagementRequiredMixin, View):
    """AJAX — return papers for a subject within a session."""

    def get(self, request, session_pk, subject_pk):
        papers = SubjectExamPaper.objects.filter(
            exam_session_id=session_pk, subject_id=subject_pk
        ).order_by('paper_number')
        return JsonResponse({
            'papers': [
                {
                    'id': p.pk,
                    'paper_number': p.paper_number,
                    'paper_name': p.paper_name,
                    'max_marks': float(p.max_marks),
                }
                for p in papers
            ]
        })


# ════════════════════════════════════════════════════════════════════════════
# RESULT ENTRY — 3 methods:
#   Method 1: Manual entry via the web grid
#   Method 2 & 3: Excel upload/download (see below)
# ════════════════════════════════════════════════════════════════════════════

class SessionResultsView(ManagementRequiredMixin, View):
    """
    Main result entry page for a session.

    Shows a grid:  Rows = students  |  Columns = subject papers
    Each cell is an editable marks input.
    Calculated grade/points are shown as read-only once saved.

    Also shows the summary table of StudentSubjectResult per student.
    """
    template_name = 'portal_management/exams/session_results.html'

    def get(self, request, pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level', 'academic_year', 'term'
            ), pk=pk
        )
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')

        students = _enrolled_students(session)

        # Load all existing scores into a dict for fast lookup
        scores = {
            (sc.student_id, sc.exam_paper_id): sc
            for sc in StudentPaperScore.objects.filter(
                exam_paper__exam_session=session
            ).select_related('exam_paper')
        }

        # Load subject results for summary column
        subject_results = {
            (r.student_id, r.subject_id): r
            for r in StudentSubjectResult.objects.filter(
                exam_session=session
            ).select_related('subject')
        }

        # Build grid: [ {student, papers: [{paper, score, result}] } ]
        grid = []
        for student in students:
            row = {'student': student, 'papers': []}
            for paper in papers:
                score_obj = scores.get((student.pk, paper.pk))
                result_obj = subject_results.get((student.pk, paper.subject_id))
                row['papers'].append({
                    'paper': paper,
                    'score': score_obj,
                    'marks': score_obj.marks if score_obj else '',
                    'grade': result_obj.grade if result_obj else '',
                    'points': result_obj.points if result_obj else '',
                })
            grid.append(row)

        # Group papers by subject for the header
        papers_by_subject = {}
        for p in papers:
            papers_by_subject.setdefault(p.subject, []).append(p)

        return render(request, self.template_name, {
            'session': session,
            'papers': papers,
            'papers_by_subject': papers_by_subject,
            'grid': grid,
            'students': students,
            'can_edit': session.status not in ('published',),
        })


class StudentResultEntryView(ManagementRequiredMixin, View):
    """
    Per-student result detail page.
    Shows all subjects + papers for this student in this session.
    Allows editing individual paper marks.
    Shows auto-calculated total, grade, points per subject.
    """
    template_name = 'portal_management/exams/student_result_entry.html'

    def get(self, request, session_pk, student_pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level'
            ), pk=session_pk
        )
        student = get_object_or_404(Student, pk=student_pk)

        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')

        scores = {
            sc.exam_paper_id: sc
            for sc in StudentPaperScore.objects.filter(
                student=student,
                exam_paper__exam_session=session
            )
        }
        subject_results = {
            r.subject_id: r
            for r in StudentSubjectResult.objects.filter(
                student=student, exam_session=session
            )
        }

        # Build subject groups with paper scores
        subjects_data = {}
        for paper in papers:
            sub = paper.subject
            if sub not in subjects_data:
                subjects_data[sub] = {
                    'subject': sub,
                    'papers': [],
                    'result': subject_results.get(sub.pk),
                }
            score_obj = scores.get(paper.pk)
            subjects_data[sub]['papers'].append({
                'paper': paper,
                'score': score_obj,
                'marks': score_obj.marks if score_obj else None,
            })

        return render(request, self.template_name, {
            'session': session,
            'student': student,
            'subjects_data': list(subjects_data.values()),
            'can_edit': session.status not in ('published',),
        })


class SavePaperScoreView(ManagementRequiredMixin, View):
    """
    AJAX — save a single paper score for one student.
    Creates or updates StudentPaperScore.
    Returns the updated score + computed subject total if all papers scored.
    """

    def post(self, request):
        try:
            student_id  = request.POST.get('student_id')
            paper_id    = request.POST.get('paper_id')
            marks_raw   = request.POST.get('marks', '').strip()

            if marks_raw == '':
                # Delete score if marks cleared
                StudentPaperScore.objects.filter(
                    student_id=student_id, exam_paper_id=paper_id
                ).delete()
                return JsonResponse({'success': True, 'cleared': True})

            marks = Decimal(marks_raw)
            paper = get_object_or_404(SubjectExamPaper, pk=paper_id)

            with transaction.atomic():
                score, _ = StudentPaperScore.objects.update_or_create(
                    student_id=student_id,
                    exam_paper=paper,
                    defaults={'marks': marks},
                )

            # Compute running total for this subject across all papers
            all_papers = SubjectExamPaper.objects.filter(
                exam_session=paper.exam_session,
                subject=paper.subject,
            )
            all_scores = StudentPaperScore.objects.filter(
                student_id=student_id,
                exam_paper__in=all_papers,
            )
            total_max  = sum(p.max_marks for p in all_papers)
            total_scored = sum(s.marks for s in all_scores)
            all_scored = all_scores.count() == all_papers.count()

            # Percentage for grade lookup
            pct = (total_scored / total_max * 100) if total_max > 0 else 0

            return JsonResponse({
                'success': True,
                'marks': float(marks),
                'paper_max': float(paper.max_marks),
                'subject_total': float(total_scored),
                'subject_max': float(total_max),
                'subject_pct': round(float(pct), 2),
                'all_scored': all_scored,
            })

        except (InvalidOperation, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid marks value.'}, status=400)
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            logger.error('SavePaperScore error: %s', e, exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class BulkSaveScoresView(ManagementRequiredMixin, View):
    """
    AJAX — save all paper scores for one student in a session at once.
    Payload: {student_id, session_id, scores: [{paper_id, marks}]}
    """

    def post(self, request):
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            scores_data = data.get('scores', [])

            saved = 0
            errors = []

            with transaction.atomic():
                for item in scores_data:
                    paper_id = item.get('paper_id')
                    marks_raw = item.get('marks')

                    if marks_raw is None or str(marks_raw).strip() == '':
                        StudentPaperScore.objects.filter(
                            student_id=student_id, exam_paper_id=paper_id
                        ).delete()
                        continue

                    try:
                        marks = Decimal(str(marks_raw))
                        paper = SubjectExamPaper.objects.get(pk=paper_id)
                        score, _ = StudentPaperScore.objects.update_or_create(
                            student_id=student_id,
                            exam_paper=paper,
                            defaults={'marks': marks},
                        )
                        saved += 1
                    except (InvalidOperation, ValueError):
                        errors.append(f'Paper {paper_id}: invalid marks.')
                    except ValidationError as e:
                        errors.append(f'Paper {paper_id}: {e}')

            return JsonResponse({
                'success': True,
                'saved': saved,
                'errors': errors,
                'message': f'Saved {saved} score(s).' + (
                    f' {len(errors)} error(s).' if errors else ''
                ),
            })

        except Exception as e:
            logger.error('BulkSaveScores error: %s', e, exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)





# ════════════════════════════════════════════════════════════════════════════
# CALCULATION VIEWS
# ════════════════════════════════════════════════════════════════════════════

class CalculateSubjectResultsView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            msg = 'Published sessions cannot be recalculated.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        try:
            from results.services import calculate_subject_results
            result = calculate_subject_results(pk)
            msg = (
                f'Subject results: {result["created"]} created, '
                f'{result["updated"]} updated, {result["skipped"]} skipped.'
            )
            if ajax:
                return JsonResponse({'success': True, 'message': msg, 'result': result})
            messages.success(request, msg)
        except Exception as e:
            logger.error('CalculateSubjectResults error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=pk)


class CalculateMetricsView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            msg = 'Published sessions cannot be recalculated.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        try:
            from results.services import calculate_metrics
            result = calculate_metrics(pk)
            msg = (
                f'Metrics: {result["created"]} created, '
                f'{result["updated"]} updated, {result["skipped"]} skipped.'
            )
            if ajax:
                return JsonResponse({'success': True, 'message': msg, 'result': result})
            messages.success(request, msg)
        except Exception as e:
            logger.error('CalculateMetrics error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=pk)


class CalculatePositionsView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            msg = 'Published sessions cannot be recalculated.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        try:
            from results.services import calculate_positions
            result = calculate_positions(pk)
            msg = (
                f'Positions: {result["class_positions"]} class, '
                f'{result["stream_positions"]} stream.'
            )
            if ajax:
                return JsonResponse({'success': True, 'message': msg, 'result': result})
            messages.success(request, msg)
        except Exception as e:
            logger.error('CalculatePositions error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=pk)


class CalculateFullResultsView(ManagementRequiredMixin, View):
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status == 'published':
            msg = 'Published sessions cannot be recalculated.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        try:
            from results.services import calculate_session_results
            result = calculate_session_results(pk)
            sr = result['subject_results']
            mr = result['metrics']
            pr = result['positions']
            msg = (
                f'Full calculation complete — '
                f'Subject results: {sr["created"]}+{sr["updated"]}, '
                f'Metrics: {mr["created"]}+{mr["updated"]}, '
                f'Positions: {pr["class_positions"]} class / {pr["stream_positions"]} stream.'
            )
            if ajax:
                return JsonResponse({'success': True, 'message': msg, 'result': result})
            messages.success(request, msg)
        except Exception as e:
            logger.error('CalculateFullResults error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
        return redirect('management:exam_session_detail', pk=pk)


# ════════════════════════════════════════════════════════════════════════════
# WORKFLOW — SUBMIT / VERIFY / PUBLISH / UNPUBLISH
# ════════════════════════════════════════════════════════════════════════════

class SubmitSessionView(ManagementRequiredMixin, View):
    """Move session from draft → submitted."""
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status != 'draft':
            msg = f'Session is "{session.status}", not draft.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        has_scores = StudentPaperScore.objects.filter(
            exam_paper__exam_session=session
        ).exists()
        if not has_scores:
            msg = 'Cannot submit — no scores have been entered yet.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        session.status = 'submitted'
        session.save(update_fields=['status'])
        msg = f'Session "{session.name}" submitted for verification.'
        if ajax:
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)
        return redirect('management:exam_session_detail', pk=pk)


class VerifySessionView(ManagementRequiredMixin, View):
    """Move session from submitted → verified."""
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status != 'submitted':
            msg = f'Session must be submitted before verification.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        has_metrics = StudentExamMetrics.objects.filter(exam_session=session).exists()
        if not has_metrics:
            msg = 'Cannot verify — metrics have not been calculated yet.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        session.status = 'verified'
        session.save(update_fields=['status'])
        msg = f'Session "{session.name}" verified.'
        if ajax:
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)
        return redirect('management:exam_session_detail', pk=pk)


class PublishSessionView(ManagementRequiredMixin, View):
    """Move session from verified → published. Students can now view results."""
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status != 'verified':
            msg = 'Session must be verified before publishing.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        session.status = 'published'
        session.save(update_fields=['status'])
        msg = f'Session "{session.name}" published. Students can now view results.'
        if ajax:
            return JsonResponse({'success': True, 'message': msg})
        messages.success(request, msg)
        return redirect('management:exam_session_detail', pk=pk)


class UnpublishSessionView(ManagementRequiredMixin, View):
    """Revert published session back to verified for corrections."""
    def post(self, request, pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=pk)
        if session.status != 'published':
            msg = 'Session is not published.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        reason = request.POST.get('reason', '').strip()
        if not reason:
            msg = 'A reason is required to unpublish.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=pk)
        session.status = 'verified'
        session.save(update_fields=['status'])
        msg = f'Session "{session.name}" unpublished. Reason: {reason}'
        if ajax:
            return JsonResponse({'success': True, 'message': msg})
        messages.warning(request, msg)
        return redirect('management:exam_session_detail', pk=pk)


# ════════════════════════════════════════════════════════════════════════════
# EXPORT
# ════════════════════════════════════════════════════════════════════════════

class ExportSubjectReportView(ManagementRequiredMixin, View):
    """Export results for one subject in a session."""

    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)

        papers = SubjectExamPaper.objects.filter(
            exam_session=session, subject=subject
        ).order_by('paper_number')
        students = _enrolled_students(session)
        scores = {
            sc.student_id: sc
            for sc in StudentPaperScore.objects.filter(
                student__in=students,
                exam_paper__exam_session=session,
                exam_paper__subject=subject,
            ).select_related('exam_paper')
        }
        results = {
            r.student_id: r
            for r in StudentSubjectResult.objects.filter(
                exam_session=session, subject=subject
            )
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{subject.short_name or subject.name}'

        # Header
        headers = ['#', 'Reg No', 'Full Name']
        for p in papers:
            headers.append(f'P{p.paper_number} (/{p.max_marks})')
        headers += ['Total', 'Grade', 'Points']

        for col, h in enumerate(headers, 1):
            _xl_header_style(ws, 1, col, h)

        # Data rows
        for r_idx, student in enumerate(students, 2):
            ws.cell(row=r_idx, column=1, value=r_idx - 1)
            ws.cell(row=r_idx, column=2, value=student.registration_number or '')
            ws.cell(row=r_idx, column=3, value=student.full_name)
            for p_idx, paper in enumerate(papers):
                sc = StudentPaperScore.objects.filter(
                    student=student, exam_paper=paper
                ).first()
                ws.cell(row=r_idx, column=4 + p_idx, value=float(sc.marks) if sc else '')
            result = results.get(student.pk)
            ws.cell(row=r_idx, column=4 + len(papers), value=float(result.total_marks) if result else '')
            ws.cell(row=r_idx, column=5 + len(papers), value=result.grade if result else '')
            ws.cell(row=r_idx, column=6 + len(papers), value=float(result.points) if result and result.points else '')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fname = f'results_{subject.short_name or subject.name}_{session.name[:20]}_{date.today()}.xlsx'.replace(' ', '_')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════════════════════════════════
# AJAX HELPERS
# ════════════════════════════════════════════════════════════════════════════

class TermsByAcademicYearView(ManagementRequiredMixin, View):
    def get(self, request):
        year_id = request.GET.get('academic_year_id')
        if not year_id:
            return JsonResponse({'terms': []})
        terms = Term.objects.filter(
            academic_year_id=year_id
        ).order_by('term_number')
        return JsonResponse({'terms': [
            {
                'id': t.pk,
                'name': f'Term {t.term_number}',
                'term_number': t.term_number,
                'start_date': t.start_date.strftime('%Y-%m-%d'),
                'end_date': t.end_date.strftime('%Y-%m-%d'),
            }
            for t in terms
        ]})


class StreamsByClassLevelView(ManagementRequiredMixin, View):
    def get(self, request):
        level_id = request.GET.get('class_level_id')
        if not level_id:
            return JsonResponse({'streams': []})
        streams = StreamClass.objects.filter(
            class_level_id=level_id
        ).order_by('stream_letter')
        return JsonResponse({'streams': [
            {'id': s.pk, 'name': str(s), 'capacity': s.capacity}
            for s in streams
        ]})


class SubjectsByClassLevelView(ManagementRequiredMixin, View):
    def get(self, request):
        level_id = request.GET.get('class_level_id')
        if not level_id:
            return JsonResponse({'subjects': []})
        cl = get_object_or_404(ClassLevel, pk=level_id)
        subjects = Subject.objects.filter(
            educational_level=cl.educational_level
        ).order_by('name')
        return JsonResponse({'subjects': [
            {'id': s.pk, 'name': s.name, 'code': s.code,
             'short_name': s.short_name or s.name}
            for s in subjects
        ]})


# Add these to your portal_management/views/exams.py file

# ════════════════════════════════════════════════════════════════════════════
# EXAM PAPER MANAGEMENT VIEWS
# ════════════════════════════════════════════════════════════════════════════

class SubjectExamPaperListView(ManagementRequiredMixin, View):
    """
    List all papers for a specific exam session.
    Displays papers grouped by subject.
    """
    template_name = 'portal_management/exams/exam_paper_list.html'

    def get(self, request, session_pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level', 'academic_year', 'term'
            ), pk=session_pk
        )
        
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        # Group papers by subject
        papers_by_subject = {}
        for paper in papers:
            papers_by_subject.setdefault(paper.subject, []).append(paper)
        
        # Get available subjects for this educational level
        available_subjects = Subject.objects.filter(
            educational_level=session.class_level.educational_level
        ).order_by('name')
        
        return render(request, self.template_name, {
            'session': session,
            'papers_by_subject': papers_by_subject,
            'papers': papers,
            'available_subjects': available_subjects,
            'can_edit': session.status != 'published',
        })


class SubjectExamPaperCreateView(ManagementRequiredMixin, View):
    """
    Create a new exam paper for a session.
    Supports both AJAX and traditional form submission.
    """
    template_name = 'portal_management/exams/exam_paper_form.html'

    def get(self, request, session_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        
        if session.status == 'published':
            messages.error(request, 'Cannot add papers to a published session.')
            return redirect('management:exam_session_detail', pk=session_pk)
        
        available_subjects = Subject.objects.filter(
            educational_level=session.class_level.educational_level
        ).order_by('name')
        
        return render(request, self.template_name, {
            'session': session,
            'available_subjects': available_subjects,
            'title': f'Add Paper - {session.name}',
            'is_edit': False,
        })

    def post(self, request, session_pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=session_pk)
        
        if session.status == 'published':
            msg = 'Cannot add papers to a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        
        errors = {}
        
        # Validate required fields
        subject_id = request.POST.get('subject')
        if not subject_id:
            errors['subject'] = ['Subject is required.']
        
        paper_number = request.POST.get('paper_number')
        if not paper_number:
            errors['paper_number'] = ['Paper number is required.']
        else:
            try:
                paper_number = int(paper_number)
                if paper_number < 1:
                    errors['paper_number'] = ['Paper number must be at least 1.']
            except ValueError:
                errors['paper_number'] = ['Invalid paper number.']
        
        max_marks = request.POST.get('max_marks')
        if not max_marks:
            errors['max_marks'] = ['Maximum marks is required.']
        else:
            try:
                max_marks = Decimal(max_marks)
                if max_marks <= 0:
                    errors['max_marks'] = ['Maximum marks must be greater than 0.']
            except InvalidOperation:
                errors['max_marks'] = ['Invalid maximum marks value.']
        
        # Check for duplicate paper number per subject
        if subject_id and paper_number and not errors:
            exists = SubjectExamPaper.objects.filter(
                exam_session=session,
                subject_id=subject_id,
                paper_number=paper_number
            ).exists()
            if exists:
                errors['paper_number'] = [
                    f'Paper {paper_number} already exists for this subject.'
                ]
        
        if errors:
            if ajax:
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            
            available_subjects = Subject.objects.filter(
                educational_level=session.class_level.educational_level
            ).order_by('name')
            
            return render(request, self.template_name, {
                'session': session,
                'available_subjects': available_subjects,
                'title': f'Add Paper - {session.name}',
                'is_edit': False,
                'form_data': request.POST,
                'errors': errors,
            })
        
        try:
            with transaction.atomic():
                paper = SubjectExamPaper(
                    exam_session=session,
                    subject_id=subject_id,
                    paper_number=paper_number,
                    paper_name=request.POST.get('paper_name', '').strip(),
                    max_marks=max_marks,
                    exam_date=request.POST.get('exam_date') or None,
                    duration_minutes=request.POST.get('duration_minutes') or None,
                )
                paper.full_clean()
                paper.save()
            
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Paper added successfully.',
                    'paper': {
                        'id': paper.pk,
                        'subject': str(paper.subject),
                        'subject_id': paper.subject_id,
                        'paper_number': paper.paper_number,
                        'paper_name': paper.paper_name,
                        'max_marks': float(paper.max_marks),
                        'duration_minutes': paper.duration_minutes,
                        'exam_date': paper.exam_date.strftime('%Y-%m-%d') if paper.exam_date else None,
                    }
                })
            
            messages.success(request, f'Paper {paper.paper_number} for {paper.subject} added successfully.')
            return redirect('management:exam_session_detail', pk=session_pk)
            
        except ValidationError as e:
            if ajax:
                msg = '; '.join(sum(e.message_dict.values(), [])) if hasattr(e, 'message_dict') else str(e)
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, str(e))
            return redirect('management:exam_paper_list', session_pk=session_pk)
            
        except Exception as e:
            logger.error('ExamPaperCreate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return redirect('management:exam_paper_list', session_pk=session_pk)


class SubjectExamPaperUpdateView(ManagementRequiredMixin, View):
    """
    Update an existing exam paper.
    Supports both AJAX and traditional form submission.
    """
    template_name = 'portal_management/exams/exam_paper_form.html'

    def get(self, request, pk):
        paper = get_object_or_404(
            SubjectExamPaper.objects.select_related(
                'exam_session', 'subject'
            ), pk=pk
        )
        
        if paper.exam_session.status == 'published':
            messages.error(request, 'Cannot edit papers in a published session.')
            return redirect('management:exam_session_detail', pk=paper.exam_session.pk)
        
        available_subjects = Subject.objects.filter(
            educational_level=paper.exam_session.class_level.educational_level
        ).order_by('name')
        
        return render(request, self.template_name, {
            'paper': paper,
            'session': paper.exam_session,
            'available_subjects': available_subjects,
            'title': f'Edit Paper - {paper}',
            'is_edit': True,
        })

    def post(self, request, pk):
        ajax = _is_ajax(request)
        paper = get_object_or_404(SubjectExamPaper, pk=pk)
        session = paper.exam_session
        
        if session.status == 'published':
            msg = 'Cannot edit papers in a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session.pk)
        
        errors = {}
        
        # Validate fields
        subject_id = request.POST.get('subject')
        if not subject_id:
            errors['subject'] = ['Subject is required.']
        
        paper_number = request.POST.get('paper_number')
        if not paper_number:
            errors['paper_number'] = ['Paper number is required.']
        else:
            try:
                paper_number = int(paper_number)
                if paper_number < 1:
                    errors['paper_number'] = ['Paper number must be at least 1.']
            except ValueError:
                errors['paper_number'] = ['Invalid paper number.']
        
        max_marks = request.POST.get('max_marks')
        if not max_marks:
            errors['max_marks'] = ['Maximum marks is required.']
        else:
            try:
                max_marks = Decimal(max_marks)
                if max_marks <= 0:
                    errors['max_marks'] = ['Maximum marks must be greater than 0.']
            except InvalidOperation:
                errors['max_marks'] = ['Invalid maximum marks value.']
        
        # Check for duplicate paper number if subject or number changed
        if (subject_id and paper_number and 
            (subject_id != str(paper.subject_id) or paper_number != paper.paper_number) and
            not errors):
            exists = SubjectExamPaper.objects.filter(
                exam_session=session,
                subject_id=subject_id,
                paper_number=paper_number
            ).exclude(pk=paper.pk).exists()
            if exists:
                errors['paper_number'] = [
                    f'Paper {paper_number} already exists for this subject.'
                ]
        
        if errors:
            if ajax:
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            
            available_subjects = Subject.objects.filter(
                educational_level=session.class_level.educational_level
            ).order_by('name')
            
            return render(request, self.template_name, {
                'paper': paper,
                'session': session,
                'available_subjects': available_subjects,
                'title': f'Edit Paper - {paper}',
                'is_edit': True,
                'form_data': request.POST,
                'errors': errors,
            })
        
        try:
            with transaction.atomic():
                paper.subject_id = subject_id
                paper.paper_number = paper_number
                paper.paper_name = request.POST.get('paper_name', '').strip()
                paper.max_marks = max_marks
                paper.exam_date = request.POST.get('exam_date') or None
                paper.duration_minutes = request.POST.get('duration_minutes') or None
                paper.full_clean()
                paper.save()
            
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Paper updated successfully.',
                    'paper': {
                        'id': paper.pk,
                        'subject': str(paper.subject),
                        'subject_id': paper.subject_id,
                        'paper_number': paper.paper_number,
                        'paper_name': paper.paper_name,
                        'max_marks': float(paper.max_marks),
                        'duration_minutes': paper.duration_minutes,
                        'exam_date': paper.exam_date.strftime('%Y-%m-%d') if paper.exam_date else None,
                    }
                })
            
            messages.success(request, f'Paper updated successfully.')
            return redirect('management:exam_session_detail', pk=session.pk)
            
        except ValidationError as e:
            if ajax:
                msg = '; '.join(sum(e.message_dict.values(), [])) if hasattr(e, 'message_dict') else str(e)
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, str(e))
            return redirect('management:exam_paper_list', session_pk=session.pk)
            
        except Exception as e:
            logger.error('ExamPaperUpdate error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return redirect('management:exam_paper_list', session_pk=session.pk)


class SubjectExamPaperDeleteView(ManagementRequiredMixin, View):
    """
    Delete an exam paper.
    Checks for existing scores before deletion.
    Returns JSON response for AJAX requests, redirects for regular requests.
    """
    
    def post(self, request, pk):
        ajax = _is_ajax(request)
        paper = get_object_or_404(SubjectExamPaper, pk=pk)
        session_pk = paper.exam_session_id
        
        # Check if session is published - cannot delete papers from published sessions
        if paper.exam_session.status == 'published':
            msg = 'Cannot delete papers in a published session.'
            if ajax:
                return JsonResponse({
                    'success': False, 
                    'message': msg,
                    'error_type': 'published_session'
                }, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        
        # Check for existing scores
        scores = StudentPaperScore.objects.filter(exam_paper=paper).select_related('student')
        score_count = scores.count()
        
        if score_count > 0:
            # Get up to 5 student names to show which students have scores
            student_names = [score.student.full_name for score in scores[:5]]
            dependencies = student_names
            if score_count > 5:
                dependencies.append(f'and {score_count - 5} more student(s)')
            
            msg = f'Cannot delete paper - {score_count} student score(s) exist for this paper.'
            
            if ajax:
                return JsonResponse({
                    'success': False, 
                    'message': msg,
                    'has_dependencies': True,
                    'dependencies': dependencies,
                    'score_count': score_count,
                    'error_type': 'has_scores'
                }, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        
        # Check if paper has subject results (should be deleted by cascade, but double-check)
        subject_results_count = StudentSubjectResult.objects.filter(
            exam_session=paper.exam_session,
            subject=paper.subject
        ).count()
        
        # Store paper details before deletion for response
        paper_name = str(paper)
        paper_id = paper.pk
        subject_name = paper.subject.name
        paper_number = paper.paper_number
        
        try:
            with transaction.atomic():
                # Delete the paper (cascades to StudentPaperScore if any)
                paper.delete()
                
                # After deleting the paper, we should check if the subject results need recalculation
                # This is handled by the signal or can be triggered separately
                
            if ajax:
                return JsonResponse({
                    'success': True, 
                    'message': f'Paper "{paper_name}" deleted successfully.',
                    'paper_id': paper_id,
                    'subject_id': paper.subject_id,
                    'subject_name': subject_name,
                    'paper_number': paper_number,
                    'redirect_url': reverse('management:exam_paper_list', args=[session_pk])
                })
            
            messages.success(request, f'Paper "{paper_name}" deleted.')
            return redirect('management:exam_paper_list', session_pk=session_pk)
            
        except Exception as e:
            logger.error(f'Error deleting paper {paper_id}: {str(e)}', exc_info=True)
            msg = f'An error occurred while deleting the paper: {str(e)}'
            
            if ajax:
                return JsonResponse({
                    'success': False,
                    'message': msg,
                    'error_type': 'server_error'
                }, status=500)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)


class SubjectExamPaperDetailView(ManagementRequiredMixin, View):
    """
    View details of a single exam paper including all student scores.
    """
    template_name = 'portal_management/exams/exam_paper_detail.html'

    def get(self, request, pk):
        paper = get_object_or_404(
            SubjectExamPaper.objects.select_related(
                'exam_session__class_level__educational_level',
                'exam_session__academic_year',
                'subject'
            ), pk=pk
        )
        
        # Get all students enrolled in this session
        students = _enrolled_students(paper.exam_session)
        
        # Get scores for this paper
        scores = {
            sc.student_id: sc
            for sc in StudentPaperScore.objects.filter(
                exam_paper=paper,
                student__in=students
            ).select_related('student')
        }
        
        # Build student list with their scores
        student_scores = []
        for student in students:
            score = scores.get(student.pk)
            student_scores.append({
                'student': student,
                'score': score,
                'marks': score.marks if score else None,
                'has_score': score is not None,
            })
        
        # Calculate statistics
        marks_list = [s['marks'] for s in student_scores if s['has_score']]
        stats = {
            'total_students': len(students),
            'students_with_scores': len(marks_list),
            'students_without_scores': len(students) - len(marks_list),
            'average': sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
        }
        
        return render(request, self.template_name, {
            'paper': paper,
            'session': paper.exam_session,
            'student_scores': student_scores,
            'stats': stats,
            'can_edit': paper.exam_session.status != 'published',
        })


# ════════════════════════════════════════════════════════════════════════════
# BULK EXAM PAPER CREATION - MULTIPLE SUBJECTS SUPPORT
# ════════════════════════════════════════════════════════════════════════════

class BulkExamPaperCreateView(ManagementRequiredMixin, View):
    """
    Render template for bulk creating exam papers across multiple subjects.
    Allows adding multiple papers for multiple subjects in a single form.
    
    Features:
        - Select multiple subjects
        - Add multiple papers per subject
        - Paper number auto-assignment (next available) if not specified
        - Default max marks = 100 if not specified
        - Validation to prevent duplicate paper numbers per subject
    """
    template_name = 'portal_management/exams/bulk_exam_paper_create.html'

    def get(self, request, session_pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level',
                'academic_year',
                'term'
            ), pk=session_pk
        )
        
        # Check if session is published
        if session.status == 'published':
            messages.error(request, 'Cannot add papers to a published session.')
            return redirect('management:exam_session_detail', pk=session_pk)
        
        # Get available subjects for this educational level
        available_subjects = Subject.objects.filter(
            educational_level=session.class_level.educational_level
        ).order_by('name')
        
        # Get existing papers to calculate next available paper numbers
        existing_papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject')
        
        # Build a map of subject_id -> list of existing paper numbers
        existing_paper_numbers = {}
        for paper in existing_papers:
            existing_paper_numbers.setdefault(paper.subject_id, []).append(paper.paper_number)
        
        # For each subject, calculate the next available paper number
        subject_next_paper = {}
        for subject in available_subjects:
            existing_numbers = existing_paper_numbers.get(subject.pk, [])
            if existing_numbers:
                next_number = max(existing_numbers) + 1
            else:
                next_number = 1
            subject_next_paper[subject.pk] = next_number
        
        return render(request, self.template_name, {
            'session': session,
            'available_subjects': available_subjects,
            'subject_next_paper': subject_next_paper,
            'existing_paper_numbers': existing_paper_numbers,
            'can_edit': session.status != 'published',
            'title': f'Bulk Create Exam Papers - {session.name}',
        })


class BulkExamPaperCreateSubmitView(ManagementRequiredMixin, View):
    """
    Process bulk exam paper creation submission.
    Accepts JSON payload with multiple subjects and multiple papers per subject.
    
    Expected JSON payload:
    {
        "papers": [
            {
                "subject_id": 24,
                "papers": [
                    {"paper_number": 1, "paper_name": "Theory", "max_marks": 100},
                    {"paper_number": 2, "paper_name": "Practical", "max_marks": 100}
                ]
            },
            {
                "subject_id": 25,
                "papers": [
                    {"paper_number": null, "paper_name": "Paper 1", "max_marks": null},
                    {"paper_number": null, "paper_name": "Paper 2", "max_marks": null}
                ]
            }
        ]
    }
    
    Paper number is optional - if not provided, automatically assigns next available number.
    Max marks is optional - defaults to 100 if not provided.
    """
    
    def post(self, request, session_pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=session_pk)
        
        # Check if session is published
        if session.status == 'published':
            msg = 'Cannot add papers to a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_paper_list', session_pk=session_pk)
        
        try:
            # Parse request data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                # Handle form data
                data = request.POST.dict()
                if 'papers' in data:
                    data['papers'] = json.loads(data['papers'])
            
            # Extract papers data
            subjects_data = data.get('papers', [])
            if not subjects_data:
                error_msg = 'At least one subject with papers is required.'
                if ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('management:exam_paper_list', session_pk=session_pk)
            
            # Validate structure
            if not isinstance(subjects_data, list):
                error_msg = 'Invalid data format. Expected a list of subjects with papers.'
                if ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('management:exam_paper_list', session_pk=session_pk)
            
            # Get existing papers to check for duplicates and calculate next numbers
            existing_papers = SubjectExamPaper.objects.filter(
                exam_session=session
            ).select_related('subject')
            
            existing_paper_numbers = {}
            for paper in existing_papers:
                existing_paper_numbers.setdefault(paper.subject_id, []).append(paper.paper_number)
            
            # Track validation results
            created = 0
            errors = []
            created_papers = []
            subject_results = {}
            
            with transaction.atomic():
                for subject_data in subjects_data:
                    subject_id = subject_data.get('subject_id')
                    papers_list = subject_data.get('papers', [])
                    
                    # Validate subject
                    if not subject_id:
                        errors.append('Subject ID is missing for one of the entries.')
                        continue
                    
                    try:
                        subject = Subject.objects.get(pk=subject_id)
                    except Subject.DoesNotExist:
                        errors.append(f'Subject with ID {subject_id} does not exist.')
                        continue
                    
                    # Validate subject belongs to the correct educational level
                    if subject.educational_level != session.class_level.educational_level:
                        errors.append(
                            f'Subject "{subject.name}" belongs to {subject.educational_level.name} '
                            f'but this exam session is for {session.class_level.educational_level.name}. '
                            f'Please select a subject from the correct educational level.'
                        )
                        continue
                    
                    if not papers_list:
                        errors.append(f'No papers provided for subject "{subject.name}".')
                        continue
                    
                    # Validate papers list structure
                    if not isinstance(papers_list, list):
                        errors.append(f'Invalid papers format for subject "{subject.name}". Expected a list.')
                        continue
                    
                    # Track paper numbers for this subject to detect duplicates in request
                    request_paper_numbers = []
                    subject_created = 0
                    subject_errors = []
                    
                    # Get existing paper numbers for this subject
                    existing_numbers = set(existing_paper_numbers.get(subject_id, []))
                    next_available = max(existing_numbers) + 1 if existing_numbers else 1
                    
                    for idx, paper_data in enumerate(papers_list):
                        # Extract paper data with defaults
                        paper_number = paper_data.get('paper_number')
                        max_marks = paper_data.get('max_marks')
                        paper_name = paper_data.get('paper_name', '').strip()
                        exam_date = paper_data.get('exam_date')
                        duration_minutes = paper_data.get('duration_minutes')
                        
                        # Auto-assign paper number if not provided
                        if paper_number is None or paper_number == '':
                            paper_number = next_available
                            # Increment for next paper in this subject
                            next_available += 1
                        else:
                            try:
                                paper_number = int(paper_number)
                            except (ValueError, TypeError):
                                subject_errors.append(
                                    f'Paper {idx + 1}: Invalid paper number format. '
                                    f'Using auto-assigned number instead.'
                                )
                                paper_number = next_available
                                next_available += 1
                        
                        # Auto-assign max marks if not provided
                        if max_marks is None or max_marks == '':
                            max_marks = Decimal('100')
                        else:
                            try:
                                max_marks = Decimal(str(max_marks))
                            except (InvalidOperation, ValueError, TypeError):
                                subject_errors.append(
                                    f'Paper {paper_number}: Invalid max marks value. '
                                    f'Using default value 100.'
                                )
                                max_marks = Decimal('100')
                        
                        # Validate paper number is positive
                        if paper_number < 1:
                            subject_errors.append(
                                f'Paper {paper_number}: Paper number must be at least 1. '
                                f'Using auto-assigned number instead.'
                            )
                            paper_number = next_available
                            next_available += 1
                        
                        # Validate max marks
                        if max_marks <= 0:
                            subject_errors.append(
                                f'Paper {paper_number}: Maximum marks must be greater than 0. '
                                f'Using default value 100.'
                            )
                            max_marks = Decimal('100')
                        if max_marks > 1000:
                            subject_errors.append(
                                f'Paper {paper_number}: Maximum marks cannot exceed 1000. '
                                f'Capping at 1000.'
                            )
                            max_marks = Decimal('1000')
                        
                        # Check for duplicate paper numbers in request
                        if paper_number in request_paper_numbers:
                            subject_errors.append(
                                f'Paper {paper_number}: Duplicate paper number found in request. '
                                f'Using auto-assigned number instead.'
                            )
                            paper_number = next_available
                            next_available += 1
                        
                        request_paper_numbers.append(paper_number)
                        
                        # Check for duplicate in database
                        if paper_number in existing_numbers:
                            # If duplicate, find next available number
                            while paper_number in existing_numbers or paper_number in request_paper_numbers:
                                paper_number = next_available
                                next_available += 1
                            subject_errors.append(
                                f'Paper {paper_number}: A paper with number {paper_data.get("paper_number", "auto")} '
                                f'already exists. Using auto-assigned number {paper_number} instead.'
                            )
                        
                        # Validate exam date if provided
                        if exam_date:
                            try:
                                from datetime import datetime
                                exam_date_obj = datetime.strptime(exam_date, '%Y-%m-%d').date()
                            except (ValueError, TypeError):
                                subject_errors.append(
                                    f'Paper {paper_number}: Invalid exam date format. Use YYYY-MM-DD. '
                                    f'Skipping date.'
                                )
                                exam_date = None
                        
                        # Validate duration if provided
                        if duration_minutes:
                            try:
                                duration_minutes = int(duration_minutes)
                                if duration_minutes <= 0:
                                    subject_errors.append(
                                        f'Paper {paper_number}: Duration must be positive. Skipping duration.'
                                    )
                                    duration_minutes = None
                                elif duration_minutes > 480:
                                    subject_errors.append(
                                        f'Paper {paper_number}: Duration cannot exceed 480 minutes (8 hours). '
                                        f'Skipping duration.'
                                    )
                                    duration_minutes = None
                            except (ValueError, TypeError):
                                subject_errors.append(
                                    f'Paper {paper_number}: Invalid duration format. Skipping duration.'
                                )
                                duration_minutes = None
                        
                        # Create the paper
                        try:
                            paper = SubjectExamPaper(
                                exam_session=session,
                                subject=subject,
                                paper_number=paper_number,
                                paper_name=paper_name,
                                max_marks=max_marks,
                                exam_date=exam_date if exam_date else None,
                                duration_minutes=duration_minutes,
                            )
                            paper.full_clean()
                            paper.save()
                            created += 1
                            subject_created += 1
                            created_papers.append({
                                'id': paper.pk,
                                'subject_id': subject.pk,
                                'subject_name': subject.name,
                                'paper_number': paper.paper_number,
                                'paper_name': paper.paper_name,
                                'max_marks': float(paper.max_marks),
                                'exam_date': paper.exam_date.strftime('%Y-%m-%d') if paper.exam_date else None,
                                'duration_minutes': paper.duration_minutes,
                            })
                            
                            # Add to existing numbers set to prevent duplicates within transaction
                            existing_numbers.add(paper_number)
                            
                        except ValidationError as e:
                            error_messages = []
                            if hasattr(e, 'message_dict'):
                                for field, msgs in e.message_dict.items():
                                    error_messages.extend(msgs)
                            else:
                                error_messages.append(str(e))
                            subject_errors.append(
                                f'Paper {paper_number}: {"; ".join(error_messages)}'
                            )
                        except Exception as e:
                            subject_errors.append(f'Paper {paper_number}: {str(e)}')
                    
                    # Store results for this subject
                    if subject_created > 0 or subject_errors:
                        subject_results[subject.name] = {
                            'created': subject_created,
                            'total': len(papers_list),
                            'errors': subject_errors,
                        }
                    
                    if subject_errors:
                        errors.extend(subject_errors)
            
            # Prepare response
            total_papers = sum(len(s.get('papers', [])) for s in subjects_data)
            success_message = f'{created} paper(s) created successfully across {len([r for r in subject_results.values() if r["created"] > 0])} subject(s).'
            
            if errors:
                success_message += f' {len(errors)} error(s) occurred.'
            
            if ajax:
                return JsonResponse({
                    'success': created > 0,
                    'message': success_message,
                    'created': created,
                    'total': total_papers,
                    'errors': errors[:50],  # Limit errors in response
                    'subject_results': subject_results,
                    'created_papers': created_papers[:100],  # Limit created papers in response
                })
            
            if created > 0:
                messages.success(request, success_message)
            if errors:
                messages.warning(request, f'Some papers could not be created. Details: {" ".join(errors[:5])}')
            
            return redirect('management:exam_paper_list', session_pk=session_pk)
            
        except json.JSONDecodeError as e:
            error_msg = f'Invalid JSON data: {str(e)}'
            logger.error('BulkExamPaperCreateSubmit JSON decode error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect('management:exam_paper_list', session_pk=session_pk)
            
        except Exception as e:
            logger.error('BulkExamPaperCreateSubmit error: %s', e, exc_info=True)
            error_msg = f'An unexpected error occurred: {str(e)}'
            if ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('management:exam_paper_list', session_pk=session_pk)
        


class SubjectExamPaperReorderView(ManagementRequiredMixin, View):
    """
    Reorder papers within a subject (change paper numbers).
    Useful when papers need to be renumbered.
    """
    
    def post(self, request, session_pk, subject_pk):
        ajax = _is_ajax(request)
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)
        
        if session.status == 'published':
            msg = 'Cannot reorder papers in a published session.'
            if ajax:
                return JsonResponse({'success': False, 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('management:exam_session_detail', pk=session_pk)
        
        try:
            data = json.loads(request.body) if ajax else request.POST
            paper_order = data.get('paper_order', [])
            
            if not paper_order:
                return JsonResponse({'success': False, 'message': 'Paper order is required.'}, status=400)
            
            with transaction.atomic():
                for item in paper_order:
                    paper_id = item.get('paper_id')
                    new_number = item.get('paper_number')
                    
                    if not paper_id or not new_number:
                        continue
                    
                    paper = SubjectExamPaper.objects.get(
                        pk=paper_id,
                        exam_session=session,
                        subject=subject
                    )
                    paper.paper_number = new_number
                    paper.save(update_fields=['paper_number'])
            
            if ajax:
                return JsonResponse({
                    'success': True,
                    'message': 'Papers reordered successfully.',
                })
            
            messages.success(request, 'Papers reordered successfully.')
            return redirect('management:exam_session_detail', pk=session_pk)
            
        except Exception as e:
            logger.error('ExamPaperReorder error: %s', e, exc_info=True)
            if ajax:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
            messages.error(request, str(e))
            return redirect('management:exam_session_detail', pk=session_pk)


# Add this to portal_management/views/exams.py

# ════════════════════════════════════════════════════════════════════════════
# BULK RESULT ENTRY FOR EXAM PAPER
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# BULK RESULT ENTRY FOR EXAM PAPER - SPLIT VIEWS
# ════════════════════════════════════════════════════════════════════════════

class BulkPaperResultEntryView(ManagementRequiredMixin, View):
    """
    Bulk result entry page for a single exam paper.
    Renders the template with student data and scores.
    """
    template_name = 'portal_management/exams/bulk_paper_result_entry.html'

    def get(self, request, paper_pk):
        paper = get_object_or_404(
            SubjectExamPaper.objects.select_related(
                'exam_session__class_level__educational_level',
                'exam_session__academic_year',
                'subject'
            ), pk=paper_pk
        )
        
        session = paper.exam_session
        
        # Check if session is published
        if session.status == 'published':
            messages.error(request, 'Cannot enter results for a published session.')
            return redirect('management:exam_paper_detail', pk=paper_pk)
        
        # Get all enrolled students
        students = _enrolled_students(session)
        
        # Get existing scores
        scores = {
            sc.student_id: sc
            for sc in StudentPaperScore.objects.filter(
                exam_paper=paper,
                student__in=students
            ).select_related('student')
        }
        
        # Get grading scale for this educational level
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # Build student list with scores and grades
        student_scores = []
        for student in students:
            score = scores.get(student.pk)
            marks = score.marks if score else None
            grade = None
            percentage = None
            
            if marks is not None:
                percentage = (float(marks) / float(paper.max_marks)) * 100
                # Find grade based on percentage
                for gs in grading_scale:
                    if gs.min_mark <= percentage <= gs.max_mark:
                        grade = gs.grade
                        break
            
            student_scores.append({
                'student': student,
                'score': score,
                'marks': marks,
                'has_score': score is not None,
                'grade': grade,
                'percentage': percentage,
            })
        
        # Calculate statistics
        marks_list = [s['marks'] for s in student_scores if s['has_score']]
        stats = {
            'total_students': len(students),
            'students_with_scores': len(marks_list),
            'students_without_scores': len(students) - len(marks_list),
            'average': sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'completion_percentage': (len(marks_list) / len(students) * 100) if students else 0,
        }
        
        # Calculate grade distribution
        grade_distribution = []
        for gs in grading_scale:
            count = sum(1 for s in student_scores if s['grade'] == gs.grade)
            percentage = (count / len(students) * 100) if students else 0
            grade_distribution.append({
                'grade': gs.grade,
                'count': count,
                'percentage': round(percentage, 1),
            })
        
        return render(request, self.template_name, {
            'paper': paper,
            'session': session,
            'student_scores': student_scores,
            'stats': stats,
            'grade_distribution': grade_distribution,
            'can_edit': session.status != 'published',
        })


class BulkPaperResultSaveView(ManagementRequiredMixin, View):
    """
    AJAX view for saving bulk results.
    Handles the actual save operation and returns JSON response.
    """
    
    def post(self, request, paper_pk):
        """Save bulk results via AJAX"""
        try:
            # Check if it's an AJAX request OR if the request expects JSON
            is_ajax = _is_ajax(request) or request.headers.get('Accept') == 'application/json'
            
            paper = get_object_or_404(SubjectExamPaper, pk=paper_pk)
            
            # Check if session is published
            if paper.exam_session.status == 'published':
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot save results for a published session.'
                }, status=400)
            
            # Parse request data - handle both JSON and form data
            if request.content_type == 'application/json':
                try:
                    data = json.loads(request.body)
                except json.JSONDecodeError as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid JSON data: {str(e)}'
                    }, status=400)
            else:
                # Handle form data
                data = request.POST.dict()
                if 'scores' in data:
                    try:
                        data['scores'] = json.loads(data['scores'])
                    except:
                        pass
            
            scores_data = data.get('scores', [])
            
            if not scores_data:
                return JsonResponse({
                    'success': False,
                    'message': 'No score data provided.'
                }, status=400)
            
            saved = 0
            errors = []
            updated_scores = []
            
            # Get grading scale for grade calculation
            grading_scale = GradingScale.objects.filter(
                education_level=paper.exam_session.class_level.educational_level
            ).order_by('-min_mark')
            
            def calculate_grade(percentage):
                for gs in grading_scale:
                    if gs.min_mark <= percentage <= gs.max_mark:
                        return gs.grade
                return 'F'
            
            with transaction.atomic():
                for item in scores_data:
                    student_id = item.get('student_id')
                    marks_raw = item.get('marks')
                    
                    if not student_id:
                        errors.append('Missing student ID')
                        continue
                    
                    # Validate student exists and is enrolled
                    try:
                        student = Student.objects.get(pk=student_id)
                    except Student.DoesNotExist:
                        errors.append(f'Student ID {student_id} not found.')
                        continue
                    
                    is_enrolled = StudentEnrollment.objects.filter(
                        student=student,
                        academic_year=paper.exam_session.academic_year,
                        class_level=paper.exam_session.class_level,
                        status='active'
                    ).exists()
                    
                    if not is_enrolled:
                        errors.append(f'Student {student.full_name} is not enrolled in this session.')
                        continue
                    
                    # Handle empty marks (clear score)
                    if marks_raw is None or str(marks_raw).strip() == '':
                        deleted = StudentPaperScore.objects.filter(
                            student_id=student_id,
                            exam_paper=paper
                        ).delete()[0]
                        if deleted:
                            saved += 1
                            updated_scores.append({
                                'student_id': student_id,
                                'student_name': student.full_name,
                                'marks': None,
                                'grade': None,
                                'percentage': None,
                                'cleared': True
                            })
                        continue
                    
                    # Validate and save marks
                    try:
                        marks = Decimal(str(marks_raw))
                        
                        # Validate marks range
                        if marks < 0:
                            errors.append(f'{student.full_name}: Marks cannot be negative.')
                            continue
                        if marks > paper.max_marks:
                            errors.append(
                                f'{student.full_name}: Marks ({marks}) exceed maximum ({paper.max_marks}).'
                            )
                            continue
                        
                        # Create or update score
                        score, created = StudentPaperScore.objects.update_or_create(
                            student_id=student_id,
                            exam_paper=paper,
                            defaults={'marks': marks}
                        )
                        
                        # Calculate percentage and grade
                        percentage = (float(marks) / float(paper.max_marks)) * 100
                        grade = calculate_grade(percentage)
                        
                        saved += 1
                        updated_scores.append({
                            'student_id': student_id,
                            'student_name': student.full_name,
                            'marks': float(marks),
                            'percentage': round(percentage, 1),
                            'grade': grade,
                            'created': created
                        })
                        
                    except (InvalidOperation, ValueError):
                        errors.append(f'{student.full_name}: Invalid marks value "{marks_raw}".')
                    except ValidationError as e:
                        errors.append(f'{student.full_name}: {str(e)}')
                    except Exception as e:
                        errors.append(f'{student.full_name}: Unexpected error: {str(e)}')
            
            # Recalculate statistics
            all_students = _enrolled_students(paper.exam_session)
            all_scores = StudentPaperScore.objects.filter(
                exam_paper=paper,
                student__in=all_students
            )
            marks_list = [float(s.marks) for s in all_scores]
            
            # Calculate grade distribution for statistics
            grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'S': 0}
            for score in all_scores:
                percentage = (float(score.marks) / float(paper.max_marks)) * 100
                grade = calculate_grade(percentage)
                if grade in grade_counts:
                    grade_counts[grade] += 1
            
            stats = {
                'total_students': len(all_students),
                'students_with_scores': len(marks_list),
                'students_without_scores': len(all_students) - len(marks_list),
                'average': sum(marks_list) / len(marks_list) if marks_list else 0,
                'highest': max(marks_list) if marks_list else 0,
                'lowest': min(marks_list) if marks_list else 0,
                'completion_percentage': (len(marks_list) / len(all_students) * 100) if all_students else 0,
                'grade_distribution': grade_counts,
            }
            
            # Prepare response
            response_data = {
                'success': True,
                'message': f'Saved {saved} score(s).',
                'saved': saved,
                'total': len(scores_data),
                'errors': errors[:20],
                'updated_scores': updated_scores[:100],
                'stats': stats,
            }
            
            if errors:
                response_data['message'] += f' {len(errors)} error(s) occurred.'
            
            return JsonResponse(response_data)
            
        except Exception as e:
            logger.error('BulkPaperResultSave error: %s', e, exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            }, status=500)


class QuickScoreSaveView(ManagementRequiredMixin, View):
    """
    Quick AJAX save for a single student's score.
    Used for inline editing in the bulk entry table.
    """
    
    def post(self, request):
        try:
            # Check if it's an AJAX request (more flexible)
            is_ajax = (
                request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
                request.headers.get('Accept') == 'application/json' or
                request.content_type == 'application/json'
            )
            
            student_id = request.POST.get('student_id') or request.POST.get('studentId')
            paper_id = request.POST.get('paper_id') or request.POST.get('paperId')
            marks_raw = request.POST.get('marks', '').strip()
            
            # Try to get from JSON if it's JSON request
            if request.content_type == 'application/json':
                try:
                    data = json.loads(request.body)
                    student_id = data.get('student_id') or data.get('studentId')
                    paper_id = data.get('paper_id') or data.get('paperId')
                    marks_raw = data.get('marks', '').strip()
                except:
                    pass
            
            if not student_id or not paper_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing student ID or paper ID.'
                }, status=400)
            
            paper = get_object_or_404(SubjectExamPaper, pk=paper_id)
            
            if paper.exam_session.status == 'published':
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot save results for a published session.'
                }, status=400)
            
            # Validate student enrollment
            student = get_object_or_404(Student, pk=student_id)
            is_enrolled = StudentEnrollment.objects.filter(
                student=student,
                academic_year=paper.exam_session.academic_year,
                class_level=paper.exam_session.class_level,
                status='active'
            ).exists()
            
            if not is_enrolled:
                return JsonResponse({
                    'success': False,
                    'message': f'Student {student.full_name} is not enrolled in this session.'
                }, status=400)
            
            # Get grading scale
            grading_scale = GradingScale.objects.filter(
                education_level=paper.exam_session.class_level.educational_level
            ).order_by('-min_mark')
            
            def calculate_grade(percentage):
                for gs in grading_scale:
                    if gs.min_mark <= percentage <= gs.max_mark:
                        return gs.grade
                return 'F'
            
            # Handle empty marks (clear score)
            if marks_raw == '':
                deleted = StudentPaperScore.objects.filter(
                    student_id=student_id,
                    exam_paper=paper
                ).delete()[0]
                
                # Recalculate statistics and grade distribution
                all_students = _enrolled_students(paper.exam_session)
                all_scores = StudentPaperScore.objects.filter(
                    exam_paper=paper,
                    student__in=all_students
                )
                marks_list = [float(s.marks) for s in all_scores]
                
                # Calculate grade distribution
                grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'S': 0}
                for score in all_scores:
                    percentage = (float(score.marks) / float(paper.max_marks)) * 100
                    grade = calculate_grade(percentage)
                    if grade in grade_counts:
                        grade_counts[grade] += 1
                
                stats = {
                    'total_students': len(all_students),
                    'students_with_scores': len(marks_list),
                    'students_without_scores': len(all_students) - len(marks_list),
                    'average': sum(marks_list) / len(marks_list) if marks_list else 0,
                    'highest': max(marks_list) if marks_list else 0,
                    'lowest': min(marks_list) if marks_list else 0,
                    'completion_percentage': (len(marks_list) / len(all_students) * 100) if all_students else 0,
                    'grade_distribution': grade_counts,  # Add grade distribution
                }
                
                return JsonResponse({
                    'success': True,
                    'cleared': True,
                    'message': 'Score cleared successfully.',
                    'stats': stats
                })
            
            # Validate and save marks
            try:
                marks = Decimal(str(marks_raw))
                
                if marks < 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'Marks cannot be negative.'
                    }, status=400)
                    
                if marks > paper.max_marks:
                    return JsonResponse({
                        'success': False,
                        'message': f'Marks cannot exceed {paper.max_marks}.'
                    }, status=400)
                
                # Save score
                score, created = StudentPaperScore.objects.update_or_create(
                    student_id=student_id,
                    exam_paper=paper,
                    defaults={'marks': marks}
                )
                
                # Calculate percentage and grade
                percentage = (float(marks) / float(paper.max_marks)) * 100
                grade = calculate_grade(percentage)
                
                # Recalculate statistics and grade distribution
                all_students = _enrolled_students(paper.exam_session)
                all_scores = StudentPaperScore.objects.filter(
                    exam_paper=paper,
                    student__in=all_students
                )
                marks_list = [float(s.marks) for s in all_scores]
                
                # Calculate grade distribution
                grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0, 'S': 0}
                for s in all_scores:
                    pct = (float(s.marks) / float(paper.max_marks)) * 100
                    g = calculate_grade(pct)
                    if g in grade_counts:
                        grade_counts[g] += 1
                
                stats = {
                    'total_students': len(all_students),
                    'students_with_scores': len(marks_list),
                    'students_without_scores': len(all_students) - len(marks_list),
                    'average': sum(marks_list) / len(marks_list) if marks_list else 0,
                    'highest': max(marks_list) if marks_list else 0,
                    'lowest': min(marks_list) if marks_list else 0,
                    'completion_percentage': (len(marks_list) / len(all_students) * 100) if all_students else 0,
                    'grade_distribution': grade_counts,  # Add grade distribution
                }
                
                return JsonResponse({
                    'success': True,
                    'marks': float(marks),
                    'percentage': round(percentage, 1),
                    'grade': grade,
                    'created': created,
                    'stats': stats,
                    'message': 'Score saved successfully.'
                })
                
            except (InvalidOperation, ValueError):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid marks value. Please enter a valid number.'
                }, status=400)
            except ValidationError as e:
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
                
        except Exception as e:
            logger.error('QuickScoreSave error: %s', e, exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            }, status=500)
        



# ════════════════════════════════════════════════════════════════════════════
# SUBJECT RESULTS SUMMARY VIEW
# ════════════════════════════════════════════════════════════════════════════

class SubjectResultsSummaryView(ManagementRequiredMixin, View):
    """
    Display aggregated results for a subject across all papers in a session.
    Shows:
        - Subject information
        - List of papers with their statistics
        - Student results with total marks per student
        - Average, highest, lowest scores
        - Grade × Gender Cross-Analysis Matrix
    """
    template_name = 'portal_management/exams/subject_results_summary.html'

    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)
        
        # Get all papers for this subject in the session
        papers = SubjectExamPaper.objects.filter(
            exam_session=session,
            subject=subject
        ).order_by('paper_number')
        
        if not papers.exists():
            messages.warning(request, f'No papers found for {subject.name} in this session.')
            return redirect('management:exam_paper_list', session_pk=session_pk)
        
        # Get enrolled students
        students = _enrolled_students(session)
        
        # Get all scores for these papers
        scores = StudentPaperScore.objects.filter(
            exam_paper__in=papers,
            student__in=students
        ).select_related('student', 'exam_paper')
        
        # Get grading scale for this educational level
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # Extract unique grades from grading scale
        matrix_grades = [gs.grade for gs in grading_scale]
        
        # Build data structures
        student_totals = {}
        paper_scores = {paper.pk: [] for paper in papers}
        
        for score in scores:
            student_id = score.student_id
            paper_id = score.exam_paper_id
            marks = float(score.marks)
            
            # Track per paper scores
            if paper_id in paper_scores:
                paper_scores[paper_id].append(marks)
            
            # Track per student totals
            if student_id not in student_totals:
                student_totals[student_id] = {
                    'student': score.student,
                    'total_marks': 0,
                    'max_possible': sum(float(p.max_marks) for p in papers),
                    'paper_scores': {}
                }
            student_totals[student_id]['total_marks'] += marks
            student_totals[student_id]['paper_scores'][paper_id] = marks
        
        # Calculate statistics for each paper
        paper_stats = []
        total_max_possible = sum(float(p.max_marks) for p in papers)
        
        for paper in papers:
            marks_list = paper_scores.get(paper.pk, [])
            paper_stats.append({
                'paper': paper,
                'students_with_scores': len(marks_list),
                'average': sum(marks_list) / len(marks_list) if marks_list else 0,
                'highest': max(marks_list) if marks_list else 0,
                'lowest': min(marks_list) if marks_list else 0,
                'max_marks': float(paper.max_marks),
            })
        
        # Calculate student results with percentages and grades
        student_results = []
        marks_list = []
        
        def calculate_grade(percentage):
            for gs in grading_scale:
                if float(gs.min_mark) <= percentage <= float(gs.max_mark):
                    return gs.grade
            return 'F'
        
        for student_id, data in student_totals.items():
            percentage = (data['total_marks'] / total_max_possible * 100) if total_max_possible > 0 else 0
            grade = calculate_grade(percentage)
            marks_list.append(data['total_marks'])
            
            student_results.append({
                'student': data['student'],
                'total_marks': data['total_marks'],
                'percentage': round(percentage, 1),
                'grade': grade,
                'paper_scores': data['paper_scores'],
            })
        
        # Sort by total marks (highest first)
        student_results.sort(key=lambda x: x['total_marks'], reverse=True)
        
        # Calculate overall statistics
        overall_stats = {
            'total_students': len(students),
            'students_with_scores': len(student_results),
            'students_without_scores': len(students) - len(student_results),
            'average': sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'total_max_possible': total_max_possible,
            'completion_percentage': (len(student_results) / len(students) * 100) if students else 0,
        }
        
        # ============================================
        # GRADE × GENDER CROSS-ANALYSIS MATRIX
        # ============================================
        # Initialize matrix with all grades from grading scale
        grade_gender_matrix = {}
        gender_totals = {'Male': 0, 'Female': 0, 'Other': 0}
        grade_totals = {grade: 0 for grade in matrix_grades}
        grand_total = 0
        
        # Get unique genders from students
        genders = {'Male', 'Female', 'Other'}
        
        # Initialize matrix for each gender
        for gender in genders:
            grade_gender_matrix[gender] = {grade: 0 for grade in matrix_grades}
        
        # Populate matrix
        for result in student_results:
            # Determine gender key
            student_gender = result['student'].gender
            if student_gender == 'male':
                gender_key = 'Male'
            elif student_gender == 'female':
                gender_key = 'Female'
            else:
                gender_key = 'Other'
            
            grade = result['grade']
            
            # Update matrix if grade exists in our grading scale
            if grade in grade_gender_matrix[gender_key]:
                grade_gender_matrix[gender_key][grade] += 1
                grade_totals[grade] += 1
                gender_totals[gender_key] += 1
                grand_total += 1
        
        # Remove empty gender entries if they have no students
        grade_gender_matrix = {k: v for k, v in grade_gender_matrix.items() if gender_totals[k] > 0}
        
        # Determine if results can be edited (session not published)
        can_edit = session.status != 'published'
        
        return render(request, self.template_name, {
            'session': session,
            'subject': subject,
            'papers': papers,
            'paper_stats': paper_stats,
            'student_results': student_results,
            'overall_stats': overall_stats,
            'total_max_possible': total_max_possible,
            'can_edit': can_edit,
            # Matrix data
            'matrix_grades': matrix_grades,
            'grade_gender_matrix': grade_gender_matrix,
            'gender_totals': gender_totals,
            'grade_totals': grade_totals,
            'grand_total': grand_total,
        })


# ════════════════════════════════════════════════════════════════════════════
# SUBJECT RESULTS EXPORT PDF VIEW (WeasyPrint)
# ════════════════════════════════════════════════════════════════════════════

class ExportSubjectResultsPDFView(ManagementRequiredMixin, View):
    """Export subject results to PDF format using WeasyPrint."""

    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)

        # Get papers for this subject
        papers = SubjectExamPaper.objects.filter(
            exam_session=session,
            subject=subject
        ).order_by('paper_number')

        if not papers.exists():
            messages.warning(request, f'No papers found for {subject.name}.')
            return redirect(
                'management:subject_results_summary',
                session_pk=session_pk, subject_pk=subject_pk,
            )

        # Get enrolled students and their scores
        students = _enrolled_students(session)
        scores = StudentPaperScore.objects.filter(
            exam_paper__in=papers,
            student__in=students
        ).select_related('student', 'exam_paper')

        total_max_possible = sum(float(p.max_marks) for p in papers)

        # Get grading scale
        ed_level = session.class_level.educational_level
        grading_scale = GradingScale.objects.filter(
            education_level=ed_level
        ).order_by('-min_mark')
        grading_scale_list = list(grading_scale)

        def calculate_grade(percentage):
            for gs in grading_scale_list:
                if float(gs.min_mark) <= percentage <= float(gs.max_mark):
                    return gs.grade
            return grading_scale_list[-1].grade if grading_scale_list else 'F'

        # Group scores by student
        student_scores = {}
        for score in scores:
            sid = score.student_id
            if sid not in student_scores:
                student_scores[sid] = {
                    'student': score.student,
                    'scores': {},
                    'total': 0,
                }
            student_scores[sid]['scores'][score.exam_paper_id] = float(score.marks)
            student_scores[sid]['total'] += float(score.marks)

        # Build student results
        student_results = []
        for data in student_scores.values():
            pct = (data['total'] / total_max_possible * 100) if total_max_possible > 0 else 0
            student_results.append({
                'student': data['student'],
                'total_marks': data['total'],
                'percentage': round(pct, 1),
                'grade': calculate_grade(pct),
                'scores': data['scores'],
            })
        student_results.sort(key=lambda x: x['total_marks'], reverse=True)

        # Calculate overall statistics
        marks_list = [r['total_marks'] for r in student_results]
        overall_stats = {
            'total_students': len(students),
            'students_with_scores': len(student_results),
            'students_without_scores': len(students) - len(student_results),
            'average': sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'completion_percentage': (len(student_results) / len(students) * 100) if students else 0,
        }

        # Calculate paper statistics
        paper_stats = []
        for paper in papers:
            paper_scores = [
                s['scores'][paper.pk]
                for s in student_scores.values()
                if paper.pk in s['scores']
            ]
            paper_stats.append({
                'paper': paper,
                'students_with_scores': len(paper_scores),
                'average': sum(paper_scores) / len(paper_scores) if paper_scores else 0,
                'highest': max(paper_scores) if paper_scores else 0,
                'lowest': min(paper_scores) if paper_scores else 0,
                'max_marks': float(paper.max_marks),
            })

        # Calculate Grade × Gender Matrix
        matrix_grades = [gs.grade for gs in grading_scale_list]
        grade_gender_matrix = {
            'Male': {g: 0 for g in matrix_grades},
            'Female': {g: 0 for g in matrix_grades},
            'Other': {g: 0 for g in matrix_grades},
        }
        gender_totals = {'Male': 0, 'Female': 0, 'Other': 0}
        grade_totals = {g: 0 for g in matrix_grades}
        grand_total = 0

        for result in student_results:
            gender_key = {
                'male': 'Male',
                'female': 'Female',
            }.get(result['student'].gender, 'Other')
            
            g = result['grade']
            if g in grade_gender_matrix[gender_key]:
                grade_gender_matrix[gender_key][g] += 1
                grade_totals[g] += 1
                gender_totals[gender_key] += 1
                grand_total += 1

        # Remove gender rows with zero students
        grade_gender_matrix = {
            k: v for k, v in grade_gender_matrix.items() if gender_totals[k] > 0
        }

        # Pass/fail counts based on points > 0
        pass_grades = {gs.grade for gs in grading_scale_list if gs.points > 0}
        pass_count = sum(1 for r in student_results if r['grade'] in pass_grades)
        pass_rate = (pass_count / len(student_results) * 100) if student_results else 0

        # Get school info (if available)
        school_info = {
            'name': 'SCHOOL NAME NOT SET',
            'address': '',
            'phone': '',
            'email': '',
            'motto': '',
            'reg_no': '',
        }
        
        # Try to get school from class level or session
        if hasattr(session.class_level, 'school'):
            school = session.class_level.school
            if school:
                school_info = {
                    'name': getattr(school, 'name', school_info['name']),
                    'address': getattr(school, 'address', ''),
                    'phone': getattr(school, 'phone', ''),
                    'email': getattr(school, 'email', ''),
                    'motto': getattr(school, 'motto', ''),
                    'reg_no': getattr(school, 'registration_number', ''),
                }

        # Define grade colors for styling
        grade_colors = {
            'A': '#10b981',
            'B': '#34d399',
            'C': '#fbbf24',
            'D': '#f59e0b',
            'E': '#ef4444',
            'F': '#dc2626',
            'S': '#8b5cf6',
        }

        context = {
            'session': session,
            'subject': subject,
            'papers': papers,
            'paper_stats': paper_stats,
            'student_results': student_results,
            'overall_stats': overall_stats,
            'grade_gender_matrix': grade_gender_matrix,
            'gender_totals': gender_totals,
            'grade_totals': grade_totals,
            'grand_total': grand_total,
            'grading_scale': grading_scale_list,
            'total_max_possible': total_max_possible,
            'pass_count': pass_count,
            'pass_rate': round(pass_rate, 1),
            'pass_grades': pass_grades,
            'generated_date': timezone.now(),
            'school_info': school_info,
            'grade_colors': grade_colors,
            'request': request,
        }

        # Render HTML
        html_string = render_to_string(
            'portal_management/exams/subject_results_pdf.html', 
            context
        )
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f'{subject.code}_{session.name}_results_{date.today()}.pdf'.replace(' ', '_').replace('/', '-')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        try:
            from weasyprint import HTML, CSS
            from weasyprint.text.fonts import FontConfiguration
            
            font_config = FontConfiguration()
            
            # Generate PDF - removed optimize_size option
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config,
                presentational_hints=True,  # Keep this option - it's valid
            )
            
            # Log successful PDF generation
            logger.info(f'PDF generated successfully: {filename}')
            
            return response
            
        except ImportError as e:
            logger.error(f'WeasyPrint not installed: {e}')
            messages.error(request, 'PDF generation is not available. Please contact system administrator.')
            return redirect(
                'management:subject_results_summary',
                session_pk=session_pk, subject_pk=subject_pk,
            )
            
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            
            # If PDF generation fails, return the HTML as a fallback for debugging
            if request.GET.get('debug') == '1':
                return HttpResponse(html_string, content_type='text/html')
            
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect(
                'management:subject_results_summary',
                session_pk=session_pk, subject_pk=subject_pk,
            )
        
# ════════════════════════════════════════════════════════════════════════════
# SUBJECT ANALYTICS VIEW
# ════════════════════════════════════════════════════════════════════════════

class SubjectAnalyticsView(ManagementRequiredMixin, View):
    """
    Comprehensive subject analytics for an exam session.
    """
    template_name = 'portal_management/exams/subject_analytics.html'

    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)
        
        # Get parameters with proper type conversion
        top_n = request.GET.get('top_n', 10)
        bottom_n = request.GET.get('bottom_n', 10)
        range_min = request.GET.get('range_min')
        range_max = request.GET.get('range_max')
        grade_filter = request.GET.get('grade')
        
        # Validate and convert parameters
        try:
            top_n = int(top_n)
            if top_n < 1:
                top_n = 10
            if top_n > 500:
                top_n = 500
        except (ValueError, TypeError):
            top_n = 10
        
        try:
            bottom_n = int(bottom_n)
            if bottom_n < 1:
                bottom_n = 10
            if bottom_n > 500:
                bottom_n = 500
        except (ValueError, TypeError):
            bottom_n = 10
        
        if range_min and range_min.strip():
            try:
                range_min = float(range_min)
                if range_min < 0:
                    range_min = 0
                if range_min > 100:
                    range_min = 100
            except (ValueError, TypeError):
                range_min = None
        else:
            range_min = None
        
        if range_max and range_max.strip():
            try:
                range_max = float(range_max)
                if range_max < 0:
                    range_max = 0
                if range_max > 100:
                    range_max = 100
            except (ValueError, TypeError):
                range_max = None
        else:
            range_max = None
        
        # Ensure min <= max
        if range_min is not None and range_max is not None and range_min > range_max:
            range_min, range_max = range_max, range_min
        
        # Get all papers for this subject in the session
        papers = SubjectExamPaper.objects.filter(
            exam_session=session,
            subject=subject
        ).order_by('paper_number')
        
        if not papers.exists():
            messages.warning(request, f'No papers found for {subject.name} in this session.')
            return redirect('management:exam_paper_list', session_pk=session_pk)
        
        # Get enrolled students
        students = _enrolled_students(session)
        
        # Get all scores for these papers
        scores = StudentPaperScore.objects.filter(
            exam_paper__in=papers,
            student__in=students
        ).select_related('student', 'exam_paper')
        
        # Get grading scale for this educational level
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # ============================================
        # DYNAMICALLY DETERMINE PASSING GRADES
        # ============================================
        # Convert Decimal to float for comparison
        passing_threshold_percentage = 50.0  # Configurable
        
        passing_grades = []
        failing_grades = []
        
        for gs in grading_scale:
            # Convert Decimal to float for comparison
            min_mark_float = float(gs.min_mark)
            
            # Check if this grade represents a passing mark
            if min_mark_float >= passing_threshold_percentage:
                passing_grades.append(gs.grade)
            else:
                failing_grades.append(gs.grade)
        
        # If no passing grades were identified, fall back to reasonable defaults
        if not passing_grades:
            # Common passing grades by educational level
            if session.class_level.educational_level.level_type in ['O_LEVEL', 'A_LEVEL']:
                passing_grades = ['A', 'B', 'C', 'D']
                failing_grades = ['E', 'F', 'S']
            else:
                passing_grades = ['A', 'B', 'C']
                failing_grades = ['D', 'E', 'F']
        
        # ============================================
        # Helper function to check if a grade is passing
        # ============================================
        def is_passing_grade(grade):
            """Determine if a grade is considered passing."""
            return grade in passing_grades
        
        # Calculate total maximum possible marks
        total_max_possible = sum(float(p.max_marks) for p in papers)
        
        # Build student result data
        student_scores = {}
        paper_scores = {paper.pk: [] for paper in papers}
        
        for score in scores:
            student_id = score.student_id
            paper_id = score.exam_paper_id
            marks = float(score.marks)
            
            if student_id not in student_scores:
                student_scores[student_id] = {
                    'student': score.student,
                    'scores': {},
                    'total': 0,
                    'paper_counts': set()
                }
            
            student_scores[student_id]['scores'][paper_id] = marks
            student_scores[student_id]['total'] += marks
            student_scores[student_id]['paper_counts'].add(paper_id)
            
            # Track per paper scores for statistics
            if paper_id in paper_scores:
                paper_scores[paper_id].append(marks)
        
        # Calculate percentage and grade for each student
        def calculate_grade(percentage):
            for gs in grading_scale:
                # Convert Decimal to float for comparison
                min_mark_float = float(gs.min_mark)
                max_mark_float = float(gs.max_mark)
                
                if min_mark_float <= percentage <= max_mark_float:
                    return {
                        'grade': gs.grade,
                        'description': gs.get_grade_display(),
                        'points': float(gs.points) if gs.points else 0,
                        'is_passing': is_passing_grade(gs.grade)
                    }
            return {
                'grade': 'F',
                'description': 'Fail',
                'points': 0,
                'is_passing': False
            }
        
        student_results = []
        for student_id, data in student_scores.items():
            total_marks = data['total']
            # Handle division by zero
            if total_max_possible > 0:
                percentage = (total_marks / total_max_possible * 100)
            else:
                percentage = 0
            grade_info = calculate_grade(percentage)
            
            student_results.append({
                'student': data['student'],
                'total_marks': total_marks,
                'percentage': round(percentage, 2),
                'grade': grade_info['grade'],
                'grade_description': grade_info['description'],
                'points': grade_info['points'],
                'is_passing': grade_info['is_passing'],
                'scores': data['scores'],
                'papers_completed': len(data['paper_counts']),
                'total_papers': len(papers),
            })
        
        # Apply filters
        filtered_results = student_results.copy()
        
        # Filter by grade
        if grade_filter:
            filtered_results = [r for r in filtered_results if r['grade'] == grade_filter.upper()]
        
        # Filter by score range
        if range_min is not None:
            filtered_results = [r for r in filtered_results if r['percentage'] >= range_min]
        if range_max is not None:
            filtered_results = [r for r in filtered_results if r['percentage'] <= range_max]
        
        # Sort by total marks for top/bottom
        sorted_by_marks = sorted(student_results, key=lambda x: x['total_marks'], reverse=True)
        
        # Top N performers
        top_performers = sorted_by_marks[:top_n]
        
        # Bottom N performers
        bottom_performers = sorted_by_marks[-bottom_n:] if bottom_n > 0 else []
        
        # Students in range
        students_in_range = filtered_results if (range_min is not None or range_max is not None) else []
        
        # Students by grade
        students_by_grade = {}
        for result in student_results:
            grade = result['grade']
            if grade not in students_by_grade:
                students_by_grade[grade] = []
            students_by_grade[grade].append(result)
        
        # ============================================
        # STATISTICAL ANALYSIS
        # ============================================
        marks_list = [r['total_marks'] for r in student_results]
        percentages = [r['percentage'] for r in student_results]
        
        from statistics import mean, median, mode, stdev
        
        stats = {
            'total_students': len(students),
            'students_with_scores': len(student_results),
            'students_without_scores': len(students) - len(student_results),
            'mean': mean(marks_list) if marks_list else 0,
            'median': median(marks_list) if marks_list else 0,
            'std_dev': stdev(marks_list) if len(marks_list) > 1 else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'total_max_possible': total_max_possible,
            'mean_percentage': mean(percentages) if percentages else 0,
            'median_percentage': median(percentages) if percentages else 0,
            'completion_rate': (len(student_results) / len(students) * 100) if students else 0,
        }
        
        # Calculate mode if possible
        try:
            stats['mode'] = mode(marks_list) if marks_list else 0
        except:
            stats['mode'] = 0
        
        # ============================================
        # GRADE DISTRIBUTION
        # ============================================
        grade_distribution = []
        total_with_grades = len(student_results)
        
        for gs in grading_scale:
            count = sum(1 for r in student_results if r['grade'] == gs.grade)
            percentage = (count / total_with_grades * 100) if total_with_grades > 0 else 0
            grade_distribution.append({
                'grade': gs.grade,
                'description': gs.get_grade_display(),
                'min_mark': float(gs.min_mark),
                'max_mark': float(gs.max_mark),
                'count': count,
                'percentage': round(percentage, 1),
                'points': float(gs.points) if gs.points else 0,
                'is_passing': is_passing_grade(gs.grade),
            })
        
        # ============================================
        # PAPER-WISE STATISTICS
        # ============================================
        paper_statistics = []
        for paper in papers:
            paper_marks = paper_scores.get(paper.pk, [])
            paper_stats = {
                'paper': paper,
                'students_with_scores': len(paper_marks),
                'max_marks': float(paper.max_marks),
                'mean': mean(paper_marks) if paper_marks else 0,
                'median': median(paper_marks) if paper_marks else 0,
                'highest': max(paper_marks) if paper_marks else 0,
                'lowest': min(paper_marks) if paper_marks else 0,
                'std_dev': stdev(paper_marks) if len(paper_marks) > 1 else 0,
                'pass_rate': 0,
            }
            
            # Calculate pass rate using dynamic passing grades
            if grading_scale and paper_marks:
                passed = 0
                for marks in paper_marks:
                    percentage = (marks / float(paper.max_marks)) * 100
                    grade = calculate_grade(percentage)
                    if grade['is_passing']:
                        passed += 1
                paper_stats['pass_rate'] = (passed / len(paper_marks) * 100) if paper_marks else 0
            
            paper_statistics.append(paper_stats)
        
        # ============================================
        # SCORE BINS FOR HISTOGRAM
        # ============================================
        max_marks_float = total_max_possible
        num_bins = 10
        bin_size = max_marks_float / num_bins if max_marks_float > 0 else 1
        
        score_bins = []
        for i in range(num_bins):
            bin_min = i * bin_size
            bin_max = (i + 1) * bin_size
            bin_count = sum(1 for m in marks_list if bin_min <= m < bin_max or (i == num_bins - 1 and m == bin_max))
            score_bins.append({
                'min': round(bin_min, 1),
                'max': round(bin_max, 1),
                'count': bin_count,
                'percentage': (bin_count / len(marks_list) * 100) if marks_list else 0,
            })
        
        # ============================================
        # GRADE × GENDER CROSS-ANALYSIS MATRIX
        # ============================================
        matrix_grades = [gs.grade for gs in grading_scale]
        grade_gender_matrix = {
            'Male': {g: 0 for g in matrix_grades},
            'Female': {g: 0 for g in matrix_grades},
            'Other': {g: 0 for g in matrix_grades}
        }
        gender_totals = {'Male': 0, 'Female': 0, 'Other': 0}
        grade_totals = {g: 0 for g in matrix_grades}
        grand_total = 0
        
        for result in student_results:
            student_gender = result['student'].gender
            if student_gender == 'male':
                gender_key = 'Male'
            elif student_gender == 'female':
                gender_key = 'Female'
            else:
                gender_key = 'Other'
            
            grade = result['grade']
            if grade in grade_gender_matrix[gender_key]:
                grade_gender_matrix[gender_key][grade] += 1
                grade_totals[grade] += 1
                gender_totals[gender_key] += 1
                grand_total += 1
        
        # Remove empty gender entries if they have no students
        grade_gender_matrix = {k: v for k, v in grade_gender_matrix.items() if gender_totals[k] > 0}
        
        # ============================================
        # SCORE RANGE ANALYSIS
        # ============================================
        score_ranges = [
            {'label': '90-100%', 'min': 90, 'max': 100, 'students': []},
            {'label': '80-89%', 'min': 80, 'max': 89, 'students': []},
            {'label': '70-79%', 'min': 70, 'max': 79, 'students': []},
            {'label': '60-69%', 'min': 60, 'max': 69, 'students': []},
            {'label': '50-59%', 'min': 50, 'max': 59, 'students': []},
            {'label': '40-49%', 'min': 40, 'max': 49, 'students': []},
            {'label': '30-39%', 'min': 30, 'max': 39, 'students': []},
            {'label': '0-29%', 'min': 0, 'max': 29, 'students': []},
        ]
        
        for result in student_results:
            for r in score_ranges:
                if r['min'] <= result['percentage'] <= r['max']:
                    r['students'].append(result)
        
        for r in score_ranges:
            r['count'] = len(r['students'])
            r['percentage'] = (r['count'] / len(student_results) * 100) if student_results else 0
        
        # ============================================
        # PERFORMANCE TRENDS ACROSS PAPERS
        # ============================================
        completed_all = [r for r in student_results if r['papers_completed'] == len(papers)]
        
        consistency_scores = []
        for result in completed_all:
            scores_list = [result['scores'].get(p.pk, 0) for p in papers]
            if scores_list and len(scores_list) > 1:
                from statistics import stdev
                variance = stdev(scores_list)
                consistency_scores.append({
                    'student': result['student'],
                    'scores': scores_list,
                    'variance': variance,
                    'average': result['total_marks'] / len(papers) if papers else 0,
                })
        
        consistent_performers = sorted(consistency_scores, key=lambda x: x['variance'])[:10] if consistency_scores else []
        inconsistent_performers = sorted(consistency_scores, key=lambda x: x['variance'], reverse=True)[:10] if consistency_scores else []
        
        # ============================================
        # PASS/FAIL ANALYSIS (Using dynamic passing grades)
        # ============================================
        passed = [r for r in student_results if r['is_passing']]
        failed = [r for r in student_results if not r['is_passing']]
        
        pass_fail_stats = {
            'passed_count': len(passed),
            'failed_count': len(failed),
            'pass_rate': (len(passed) / len(student_results) * 100) if student_results else 0,
            'fail_rate': (len(failed) / len(student_results) * 100) if student_results else 0,
            'passing_grades': passing_grades,
            'failing_grades': failing_grades,
        }
        
        # Pass/Fail by Gender
        pass_fail_gender = {
            'Male': {'passed': 0, 'failed': 0},
            'Female': {'passed': 0, 'failed': 0},
            'Other': {'passed': 0, 'failed': 0},
        }
        
        for result in student_results:
            gender_key = 'Male' if result['student'].gender == 'male' else 'Female' if result['student'].gender == 'female' else 'Other'
            if result['is_passing']:
                pass_fail_gender[gender_key]['passed'] += 1
            else:
                pass_fail_gender[gender_key]['failed'] += 1
        
        # ============================================
        # GENDER PERFORMANCE
        # ============================================
        gender_performance = {
            'Male': {'count': 0, 'total_marks': 0, 'average': 0, 'mean_percentage': 0, 'pass_count': 0, 'fail_count': 0, 'pass_rate': 0},
            'Female': {'count': 0, 'total_marks': 0, 'average': 0, 'mean_percentage': 0, 'pass_count': 0, 'fail_count': 0, 'pass_rate': 0},
            'Other': {'count': 0, 'total_marks': 0, 'average': 0, 'mean_percentage': 0, 'pass_count': 0, 'fail_count': 0, 'pass_rate': 0},
        }
        
        for result in student_results:
            gender_key = 'Male' if result['student'].gender == 'male' else 'Female' if result['student'].gender == 'female' else 'Other'
            gender_performance[gender_key]['count'] += 1
            gender_performance[gender_key]['total_marks'] += result['total_marks']
            if result['is_passing']:
                gender_performance[gender_key]['pass_count'] += 1
            else:
                gender_performance[gender_key]['fail_count'] += 1
        
        for gender in gender_performance:
            count = gender_performance[gender]['count']
            if count > 0:
                gender_performance[gender]['average'] = gender_performance[gender]['total_marks'] / count
                if total_max_possible > 0:
                    gender_performance[gender]['mean_percentage'] = (gender_performance[gender]['average'] / total_max_possible * 100)
                else:
                    gender_performance[gender]['mean_percentage'] = 0
                gender_performance[gender]['pass_rate'] = (gender_performance[gender]['pass_count'] / count * 100)
        
        return render(request, self.template_name, {
            'session': session,
            'subject': subject,
            'papers': papers,
            'student_results': student_results,
            'filtered_results': filtered_results,
            'top_performers': top_performers,
            'bottom_performers': bottom_performers,
            'students_in_range': students_in_range,
            'students_by_grade': students_by_grade,
            'stats': stats,
            'grade_distribution': grade_distribution,
            'paper_statistics': paper_statistics,
            'grade_gender_matrix': grade_gender_matrix,
            'gender_totals': gender_totals,
            'grade_totals': grade_totals,
            'grand_total': grand_total,
            'score_ranges': score_ranges,
            'score_bins': score_bins,
            'consistent_performers': consistent_performers,
            'inconsistent_performers': inconsistent_performers,
            'pass_fail_stats': pass_fail_stats,
            'pass_fail_gender': pass_fail_gender,
            'gender_performance': gender_performance,
            'matrix_grades': matrix_grades,
            'passing_grades': passing_grades,
            'failing_grades': failing_grades,
            'total_max_possible': total_max_possible,
            'top_n': top_n,
            'bottom_n': bottom_n,
            'range_min': range_min,
            'range_max': range_max,
            'grade_filter': grade_filter,
            'can_edit': session.status != 'published',
        })

# ════════════════════════════════════════════════════════════════════════════
# PAPER ANALYTICS VIEW
# ════════════════════════════════════════════════════════════════════════════

class PaperAnalyticsView(ManagementRequiredMixin, View):
    """
    Comprehensive analytics for a single exam paper.
    
    Features:
        - Paper details and statistics
        - Score distribution histogram
        - Grade distribution for the paper
        - Top and bottom performers on this paper
        - Statistical analysis (mean, median, mode, std deviation)
        - Score range analysis
        - Gender-based performance comparison
        - Comparison with overall subject performance
    """
    template_name = 'portal_management/exams/paper_analytics.html'

    def get(self, request, paper_pk):
        paper = get_object_or_404(
            SubjectExamPaper.objects.select_related(
                'exam_session__class_level__educational_level',
                'exam_session__academic_year',
                'subject'
            ), pk=paper_pk
        )
        
        session = paper.exam_session
        subject = paper.subject
        
        # Get enrolled students
        students = _enrolled_students(session)
        
        # Get all scores for this paper
        scores = StudentPaperScore.objects.filter(
            exam_paper=paper,
            student__in=students
        ).select_related('student')
        
        # Get grading scale for percentage-based grading
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # ============================================
        # DYNAMICALLY DETERMINE PASSING GRADES
        # ============================================
        passing_threshold_percentage = 50.0  # Configurable threshold
        
        passing_grades = []
        failing_grades = []
        
        for gs in grading_scale:
            min_mark_float = float(gs.min_mark)
            if min_mark_float >= passing_threshold_percentage:
                passing_grades.append(gs.grade)
            else:
                failing_grades.append(gs.grade)
        
        # If no passing grades were identified, fall back to reasonable defaults
        if not passing_grades:
            if session.class_level.educational_level.level_type in ['O_LEVEL', 'A_LEVEL']:
                passing_grades = ['A', 'B', 'C', 'D']
                failing_grades = ['E', 'F', 'S']
            else:
                passing_grades = ['A', 'B', 'C']
                failing_grades = ['D', 'E', 'F']
        
        # ============================================
        # Helper function to check if a grade is passing
        # ============================================
        def is_passing_grade(grade):
            return grade in passing_grades
        
        # Build score data
        score_data = []
        marks_list = []
        
        for score in scores:
            # Convert to float safely
            marks = float(score.marks)
            marks_list.append(marks)
            percentage = (marks / float(paper.max_marks)) * 100 if float(paper.max_marks) > 0 else 0
            
            # Calculate grade based on percentage
            grade = None
            grade_description = None
            for gs in grading_scale:
                min_mark_float = float(gs.min_mark)
                max_mark_float = float(gs.max_mark)
                if min_mark_float <= percentage <= max_mark_float:
                    grade = gs.grade
                    grade_description = gs.get_grade_display()
                    break
            
            score_data.append({
                'student': score.student,
                'marks': marks,
                'percentage': round(percentage, 2),
                'grade': grade,
                'grade_description': grade_description,
                'is_passing': is_passing_grade(grade) if grade else False,
                'has_score': True,
            })
        
        # Identify students without scores
        students_with_scores = {s['student'].pk for s in score_data}
        students_without_scores = [s for s in students if s.pk not in students_with_scores]
        
        # Sort score data for rankings
        sorted_by_marks = sorted(score_data, key=lambda x: x['marks'], reverse=True)
        
        # ============================================
        # STATISTICAL ANALYSIS
        # ============================================
        from statistics import mean, median, mode, stdev
        
        stats = {
            'total_students': len(students),
            'students_with_scores': len(score_data),
            'students_without_scores': len(students_without_scores),
            'completion_rate': (len(score_data) / len(students) * 100) if students else 0,
            'max_marks': float(paper.max_marks),
            'mean': mean(marks_list) if marks_list else 0,
            'median': median(marks_list) if marks_list else 0,
            'std_dev': stdev(marks_list) if len(marks_list) > 1 else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'mean_percentage': (mean(marks_list) / float(paper.max_marks) * 100) if marks_list and float(paper.max_marks) > 0 else 0,
        }
        
        # Calculate mode if possible
        try:
            stats['mode'] = mode(marks_list) if marks_list else 0
        except:
            stats['mode'] = 0
        
        # ============================================
        # TOP AND BOTTOM PERFORMERS
        # ============================================
        top_n = min(10, len(sorted_by_marks))
        bottom_n = min(10, len(sorted_by_marks))
        
        top_performers = sorted_by_marks[:top_n]
        bottom_performers = sorted_by_marks[-bottom_n:] if bottom_n > 0 else []
        
        # ============================================
        # SCORE DISTRIBUTION (Histogram bins)
        # ============================================
        max_marks_float = float(paper.max_marks)
        num_bins = 10
        bin_size = max_marks_float / num_bins if max_marks_float > 0 else 1
        
        score_bins = []
        for i in range(num_bins):
            bin_min = i * bin_size
            bin_max = (i + 1) * bin_size
            bin_count = sum(1 for m in marks_list if bin_min <= m < bin_max or (i == num_bins - 1 and m == bin_max))
            score_bins.append({
                'min': round(bin_min, 1),
                'max': round(bin_max, 1),
                'count': bin_count,
                'percentage': (bin_count / len(marks_list) * 100) if marks_list else 0,
            })
        
        # ============================================
        # GRADE DISTRIBUTION
        # ============================================
        grade_distribution = []
        total_with_grades = len(score_data)
        
        for gs in grading_scale:
            count = sum(1 for s in score_data if s['grade'] == gs.grade)
            percentage = (count / total_with_grades * 100) if total_with_grades > 0 else 0
            grade_distribution.append({
                'grade': gs.grade,
                'description': gs.get_grade_display(),
                'min_mark': float(gs.min_mark),
                'max_mark': float(gs.max_mark),
                'count': count,
                'percentage': round(percentage, 1),
                'is_passing': is_passing_grade(gs.grade),
            })
        
        # ============================================
        # SCORE RANGE ANALYSIS
        # ============================================
        score_ranges = [
            {'label': '90-100%', 'min': 90, 'max': 100, 'students': []},
            {'label': '80-89%', 'min': 80, 'max': 89, 'students': []},
            {'label': '70-79%', 'min': 70, 'max': 79, 'students': []},
            {'label': '60-69%', 'min': 60, 'max': 69, 'students': []},
            {'label': '50-59%', 'min': 50, 'max': 59, 'students': []},
            {'label': '40-49%', 'min': 40, 'max': 49, 'students': []},
            {'label': '30-39%', 'min': 30, 'max': 39, 'students': []},
            {'label': '0-29%', 'min': 0, 'max': 29, 'students': []},
        ]
        
        for result in score_data:
            for r in score_ranges:
                if r['min'] <= result['percentage'] <= r['max']:
                    r['students'].append(result)
        
        for r in score_ranges:
            r['count'] = len(r['students'])
            r['percentage'] = (r['count'] / len(score_data) * 100) if score_data else 0
        
        # ============================================
        # GENDER-BASED PERFORMANCE
        # ============================================
        gender_performance = {
            'Male': {'count': 0, 'total_marks': 0, 'marks_list': [], 'pass_count': 0, 'fail_count': 0},
            'Female': {'count': 0, 'total_marks': 0, 'marks_list': [], 'pass_count': 0, 'fail_count': 0},
            'Other': {'count': 0, 'total_marks': 0, 'marks_list': [], 'pass_count': 0, 'fail_count': 0},
        }
        
        # Get all possible grades from grading scale for distribution
        all_grades = [gs.grade for gs in grading_scale]
        gender_distribution = {
            'Male': {g: 0 for g in all_grades},
            'Female': {g: 0 for g in all_grades},
            'Other': {g: 0 for g in all_grades},
        }
        
        for result in score_data:
            gender_key = 'Male' if result['student'].gender == 'male' else 'Female' if result['student'].gender == 'female' else 'Other'
            gender_performance[gender_key]['count'] += 1
            gender_performance[gender_key]['total_marks'] += result['marks']
            gender_performance[gender_key]['marks_list'].append(result['marks'])
            
            if result['is_passing']:
                gender_performance[gender_key]['pass_count'] += 1
            else:
                gender_performance[gender_key]['fail_count'] += 1
            
            if result['grade'] in gender_distribution[gender_key]:
                gender_distribution[gender_key][result['grade']] += 1
        
        # Calculate averages and statistics for each gender
        for gender in gender_performance:
            count = gender_performance[gender]['count']
            if count > 0:
                marks_list_gender = gender_performance[gender]['marks_list']
                gender_performance[gender]['average'] = gender_performance[gender]['total_marks'] / count
                gender_performance[gender]['mean_percentage'] = (gender_performance[gender]['average'] / float(paper.max_marks)) * 100 if float(paper.max_marks) > 0 else 0
                
                # Calculate median
                if len(marks_list_gender) > 1:
                    from statistics import median
                    gender_performance[gender]['median'] = median(marks_list_gender)
                elif marks_list_gender:
                    gender_performance[gender]['median'] = marks_list_gender[0]
                else:
                    gender_performance[gender]['median'] = 0
                
                gender_performance[gender]['highest'] = max(marks_list_gender) if marks_list_gender else 0
                gender_performance[gender]['lowest'] = min(marks_list_gender) if marks_list_gender else 0
                gender_performance[gender]['pass_rate'] = (gender_performance[gender]['pass_count'] / count * 100) if count > 0 else 0
            else:
                gender_performance[gender]['average'] = 0
                gender_performance[gender]['mean_percentage'] = 0
                gender_performance[gender]['median'] = 0
                gender_performance[gender]['highest'] = 0
                gender_performance[gender]['lowest'] = 0
                gender_performance[gender]['pass_rate'] = 0
        
        # ============================================
        # PASS/FAIL ANALYSIS (Using dynamic passing grades)
        # ============================================
        passed = [s for s in score_data if s['is_passing']]
        failed = [s for s in score_data if not s['is_passing'] and s['grade'] is not None]
        
        pass_fail_stats = {
            'passed_count': len(passed),
            'failed_count': len(failed),
            'pass_rate': (len(passed) / len(score_data) * 100) if score_data else 0,
            'fail_rate': (len(failed) / len(score_data) * 100) if score_data else 0,
            'passing_grades': passing_grades,
            'failing_grades': failing_grades,
        }
        
        # Pass/Fail by Gender
        pass_fail_gender = {
            'Male': {'passed': 0, 'failed': 0},
            'Female': {'passed': 0, 'failed': 0},
            'Other': {'passed': 0, 'failed': 0},
        }
        
        for result in score_data:
            gender_key = 'Male' if result['student'].gender == 'male' else 'Female' if result['student'].gender == 'female' else 'Other'
            if result['is_passing']:
                pass_fail_gender[gender_key]['passed'] += 1
            elif result['grade'] is not None:
                pass_fail_gender[gender_key]['failed'] += 1
        
        # ============================================
        # COMPARISON WITH SUBJECT OVERALL PERFORMANCE
        # ============================================
        # Get subject results for comparison
        subject_results = StudentSubjectResult.objects.filter(
            exam_session=session,
            subject=subject
        ).select_related('student')
        
        subject_scores_map = {r.student_id: float(r.total_marks) for r in subject_results}
        
        # Calculate subject max (sum of all paper max marks for this subject)
        subject_papers = SubjectExamPaper.objects.filter(
            exam_session=session, 
            subject=subject
        )
        subject_max = sum(float(p.max_marks) for p in subject_papers) if subject_papers.exists() else 0
        
        comparison_data = []
        for result in score_data:
            student_id = result['student'].pk
            subject_total = subject_scores_map.get(student_id)
            if subject_total is not None and subject_max > 0:
                subject_percentage = (subject_total / subject_max * 100)
                comparison_data.append({
                    'student': result['student'],
                    'paper_marks': result['marks'],
                    'paper_percentage': result['percentage'],
                    'subject_total': subject_total,
                    'subject_percentage': round(subject_percentage, 2),
                    'difference': round(result['percentage'] - subject_percentage, 2),
                })
        
        # Find students who performed better/worse on paper vs overall
        better_on_paper = sorted(comparison_data, key=lambda x: x['difference'], reverse=True)[:10] if comparison_data else []
        worse_on_paper = sorted(comparison_data, key=lambda x: x['difference'])[:10] if comparison_data else []
        
        # ============================================
        # PERCENTILE RANKINGS
        # ============================================
        marks_list_sorted = sorted(marks_list)
        
        def calculate_percentile(marks):
            if not marks_list_sorted:
                return 0
            count_less = sum(1 for m in marks_list_sorted if m < marks)
            return (count_less / len(marks_list_sorted)) * 100
        
        percentile_data = []
        for result in score_data:
            percentile = calculate_percentile(result['marks'])
            percentile_data.append({
                'student': result['student'],
                'marks': result['marks'],
                'percentile': round(percentile, 1),
                'rank': sum(1 for m in marks_list_sorted if m > result['marks']) + 1,
            })
        
        percentile_data.sort(key=lambda x: x['percentile'], reverse=True)
        
        # Add rank to each student in score_data
        for i, result in enumerate(sorted_by_marks):
            result['rank'] = i + 1
        
        # Calculate subject average for display
        subject_avg = None
        subject_avg_percentage = None
        if subject_results:
            subject_totals = [float(r.total_marks) for r in subject_results]
            subject_avg = sum(subject_totals) / len(subject_totals) if subject_totals else 0
            subject_avg_percentage = (subject_avg / subject_max * 100) if subject_max > 0 else 0
        
        return render(request, self.template_name, {
            'paper': paper,
            'session': session,
            'subject': subject,
            'score_data': score_data,
            'students_without_scores': students_without_scores,
            'stats': stats,
            'top_performers': top_performers,
            'bottom_performers': bottom_performers,
            'score_bins': score_bins,
            'grade_distribution': grade_distribution,
            'score_ranges': score_ranges,
            'gender_performance': gender_performance,
            'gender_distribution': gender_distribution,
            'pass_fail_stats': pass_fail_stats,
            'pass_fail_gender': pass_fail_gender,
            'comparison_data': comparison_data[:50],  # Limit for display
            'better_on_paper': better_on_paper,
            'worse_on_paper': worse_on_paper,
            'percentile_data': percentile_data[:50],
            'passing_grades': passing_grades,
            'failing_grades': failing_grades,
            'subject_avg': subject_avg,
            'subject_avg_percentage': subject_avg_percentage,
            'subject_max': subject_max,
            'can_edit': session.status != 'published',
        })

# ════════════════════════════════════════════════════════════════════════════
# SUBJECT ANALYTICS EXPORT PDF VIEW
# ════════════════════════════════════════════════════════════════════════════

class ExportSubjectAnalyticsPDFView(ManagementRequiredMixin, View):
    """Export subject analytics to PDF format."""
    
    def get(self, request, session_pk, subject_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        subject = get_object_or_404(Subject, pk=subject_pk)
        
        # Get all papers for this subject
        papers = SubjectExamPaper.objects.filter(
            exam_session=session,
            subject=subject
        ).order_by('paper_number')
        
        if not papers.exists():
            messages.warning(request, f'No papers found for {subject.name}.')
            return redirect('management:subject_analytics', session_pk=session_pk, subject_pk=subject_pk)
        
        # Get enrolled students
        students = _enrolled_students(session)
        
        # Get all scores
        scores = StudentPaperScore.objects.filter(
            exam_paper__in=papers,
            student__in=students
        ).select_related('student', 'exam_paper')
        
        # Get grading scale
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # Build student results (simplified for PDF)
        total_max_possible = sum(float(p.max_marks) for p in papers)
        
        def calculate_grade(percentage):
            for gs in grading_scale:
                if float(gs.min_mark) <= percentage <= float(gs.max_mark):
                    return {'grade': gs.grade, 'description': gs.get_grade_display()}
            return {'grade': 'F', 'description': 'Fail'}
        
        student_results = []
        student_scores_map = {}
        
        for score in scores:
            student_id = score.student_id
            if student_id not in student_scores_map:
                student_scores_map[student_id] = {
                    'student': score.student,
                    'total': 0,
                }
            student_scores_map[student_id]['total'] += float(score.marks)
        
        for student_id, data in student_scores_map.items():
            percentage = (data['total'] / total_max_possible * 100) if total_max_possible > 0 else 0
            grade_info = calculate_grade(percentage)
            student_results.append({
                'student': data['student'],
                'total_marks': data['total'],
                'percentage': round(percentage, 2),
                'grade': grade_info['grade'],
                'grade_description': grade_info['description'],
            })
        
        student_results.sort(key=lambda x: x['total_marks'], reverse=True)
        
        # Calculate grade distribution
        grade_distribution = []
        for gs in grading_scale:
            count = sum(1 for r in student_results if r['grade'] == gs.grade)
            grade_distribution.append({
                'grade': gs.grade,
                'description': gs.get_grade_display(),
                'count': count,
                'percentage': (count / len(student_results) * 100) if student_results else 0,
            })
        
        # Generate HTML content for PDF
        context = {
            'session': session,
            'subject': subject,
            'papers': papers,
            'student_results': student_results[:100],  # Limit for PDF
            'total_max_possible': total_max_possible,
            'grade_distribution': grade_distribution,
            'total_students': len(students),
            'students_with_scores': len(student_results),
            'generated_date': timezone.now(),
            'request': request,
        }
        
        # Render HTML template
        html_string = render_to_string('portal_management/exams/subject_analytics_pdf.html', context)
        
        # Configure fonts
        font_config = FontConfiguration()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f'{subject.name}_{session.name}_analytics_{date.today()}.pdf'.replace(' ', '_').replace('/', '_')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        try:
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config
            )
            return response
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect('management:subject_analytics', session_pk=session_pk, subject_pk=subject_pk)


# ════════════════════════════════════════════════════════════════════════════
# PAPER ANALYTICS EXPORT PDF VIEW
# ════════════════════════════════════════════════════════════════════════════

class ExportPaperAnalyticsPDFView(ManagementRequiredMixin, View):
    """Export paper analytics to PDF format."""
    
    def get(self, request, paper_pk):
        paper = get_object_or_404(SubjectExamPaper, pk=paper_pk)
        session = paper.exam_session
        subject = paper.subject
        
        # Get enrolled students
        students = _enrolled_students(session)
        
        # Get scores
        scores = StudentPaperScore.objects.filter(
            exam_paper=paper,
            student__in=students
        ).select_related('student')
        
        # Build score data
        score_data = []
        marks_list = []
        
        for score in scores:
            marks = float(score.marks)
            marks_list.append(marks)
            score_data.append({
                'student': score.student,
                'marks': marks,
                'percentage': (marks / float(paper.max_marks)) * 100,
            })
        
        score_data.sort(key=lambda x: x['marks'], reverse=True)
        
        # Calculate statistics
        from statistics import mean, median
        stats = {
            'total_students': len(students),
            'students_with_scores': len(score_data),
            'students_without_scores': len(students) - len(score_data),
            'max_marks': float(paper.max_marks),
            'mean': mean(marks_list) if marks_list else 0,
            'median': median(marks_list) if marks_list else 0,
            'highest': max(marks_list) if marks_list else 0,
            'lowest': min(marks_list) if marks_list else 0,
            'mean_percentage': (mean(marks_list) / float(paper.max_marks) * 100) if marks_list else 0,
        }
        
        # Generate HTML content for PDF
        context = {
            'paper': paper,
            'session': session,
            'subject': subject,
            'score_data': score_data[:100],  # Limit for PDF
            'stats': stats,
            'generated_date': timezone.now(),
            'request': request,
        }
        
        # Render HTML template
        html_string = render_to_string('portal_management/exams/paper_analytics_pdf.html', context)
        
        # Configure fonts
        font_config = FontConfiguration()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f'{subject.name}_Paper{paper.paper_number}_{session.name}_analytics_{date.today()}.pdf'.replace(' ', '_').replace('/', '_')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        try:
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config
            )
            return response
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect('management:paper_analytics', paper_pk=paper_pk)





# ──────────────────────────────────────────────────────────────
# Helper functions
# ──────────────────────────────────────────────────────────────



class PaperResultsFilterView(ManagementRequiredMixin, View):
    """
    View paper results with filtering capabilities.
    Allows filtering students by:
        - All students
        - Students with scores (have marks)
        - Students without scores (no marks entered yet)
    """
    template_name = 'portal_management/exams/paper_results_filter.html'

    def get(self, request, paper_pk):
        paper = get_object_or_404(
            SubjectExamPaper.objects.select_related(
                'exam_session__class_level__educational_level',
                'exam_session__academic_year',
                'subject'
            ), pk=paper_pk
        )
        
        session = paper.exam_session
        
        # Get filter parameter
        filter_type = request.GET.get('filter', 'all')
        search_query = request.GET.get('search', '').strip()
        
        # Check if PDF export is requested
        export_pdf = request.GET.get('export_pdf', 'false').lower() == 'true'
        
        # Get all enrolled students
        students = _enrolled_students(session)
        
        # Get existing scores
        scores = {
            sc.student_id: sc
            for sc in StudentPaperScore.objects.filter(
                exam_paper=paper,
                student__in=students
            ).select_related('student')
        }
        
        # Build student list with score status
        student_results = []
        for student in students:
            score = scores.get(student.pk)
            has_score = score is not None
            marks = float(score.marks) if score else None
            percentage = (marks / float(paper.max_marks) * 100) if marks is not None else None
            
            # Calculate grade if percentage exists
            grade = None
            if percentage is not None:
                grading_scale = GradingScale.objects.filter(
                    education_level=session.class_level.educational_level
                ).order_by('-min_mark')
                for gs in grading_scale:
                    if float(gs.min_mark) <= percentage <= float(gs.max_mark):
                        grade = gs.grade
                        break
            
            student_results.append({
                'student': student,
                'has_score': has_score,
                'marks': marks,
                'percentage': round(percentage, 2) if percentage is not None else None,
                'grade': grade,
                'score': score,
            })
        
        # Apply filter based on selection
        if filter_type == 'with_scores':
            filtered_results = [r for r in student_results if r['has_score']]
        elif filter_type == 'without_scores':
            filtered_results = [r for r in student_results if not r['has_score']]
        else:
            filtered_results = student_results
        
        # Apply search filter
        if search_query:
            filtered_results = [
                r for r in filtered_results
                if search_query.lower() in r['student'].full_name.lower()
                or search_query.lower() in (r['student'].registration_number or '').lower()
            ]
        
        # Sort results
        filtered_results.sort(key=lambda x: (x['has_score'], x['student'].full_name))
        
        # Calculate statistics
        stats = {
            'total_students': len(students),
            'with_scores': sum(1 for r in student_results if r['has_score']),
            'without_scores': sum(1 for r in student_results if not r['has_score']),
            'completion_percentage': (sum(1 for r in student_results if r['has_score']) / len(students) * 100) if students else 0,
        }
        
        # Calculate score statistics if there are scores
        marks_list = [r['marks'] for r in student_results if r['has_score']]
        if marks_list:
            stats['average_score'] = round(mean(marks_list), 2)
            stats['highest_score'] = max(marks_list)
            stats['lowest_score'] = min(marks_list)
        else:
            stats['average_score'] = 0
            stats['highest_score'] = 0
            stats['lowest_score'] = 0
        
        # Calculate grade distribution
        grade_distribution = {}
        for result in filtered_results:
            if result['grade']:
                grade_distribution[result['grade']] = grade_distribution.get(result['grade'], 0) + 1
        
        # Get grading scale for reference
        grading_scale = GradingScale.objects.filter(
            education_level=session.class_level.educational_level
        ).order_by('-min_mark')
        
        # If PDF export requested, return PDF
        if export_pdf:
            return self.export_pdf(request, paper_pk, filter_type, search_query, 
                                   filtered_results, stats, paper, session, 
                                   grading_scale, grade_distribution)
        
        # Otherwise render HTML
        return render(request, self.template_name, {
            'paper': paper,
            'session': session,
            'filtered_results': filtered_results,
            'stats': stats,
            'filter_type': filter_type,
            'search_query': search_query,
            'grade_distribution': grade_distribution,
            'grading_scale': grading_scale,
            'can_edit': session.status != 'published',
            'total_students': len(students),
        })
    
    def export_pdf(self, request, paper_pk, filter_type, search_query, 
                   filtered_results, stats, paper, session, 
                   grading_scale, grade_distribution):
        """Export filtered paper results to PDF using WeasyPrint."""
        
        # Get school profile for this educational level
        school_profile = SchoolProfile.objects.get_active_profile(
            session.class_level.educational_level
        )
        
        # Prepare school info
        school_info = self._get_school_info(school_profile, session.class_level.educational_level)
        
        # Calculate additional statistics
        marks_list = [r['marks'] for r in filtered_results if r['has_score']]
        percentages = [r['percentage'] for r in filtered_results if r['has_score']]
        
        pdf_stats = {
            'total_students': len(filtered_results),
            'with_scores': len(marks_list),
            'without_scores': len(filtered_results) - len(marks_list),
            'average_score': mean(marks_list) if marks_list else 0,
            'median_score': median(marks_list) if marks_list else 0,
            'highest_score': max(marks_list) if marks_list else 0,
            'lowest_score': min(marks_list) if marks_list else 0,
            'average_percentage': mean(percentages) if percentages else 0,
            'completion_rate': (len(marks_list) / len(filtered_results) * 100) if filtered_results else 0,
        }
        
        # Calculate gender distribution
        gender_counts = {
            'Male': sum(1 for r in filtered_results if r['student'].gender == 'male'),
            'Female': sum(1 for r in filtered_results if r['student'].gender == 'female'),
            'Other': sum(1 for r in filtered_results if r['student'].gender not in ['male', 'female']),
        }
        
        # Prepare context for PDF template
        context = {
            'paper': paper,
            'session': session,
            'filter_type': filter_type,
            'search_query': search_query,
            'filtered_results': filtered_results,
            'stats': stats,
            'pdf_stats': pdf_stats,
            'school_info': school_info,
            'gender_counts': gender_counts,
            'filter_label': self._get_filter_label(filter_type),
            'generated_date': timezone.now(),
            'total_pages': (len(filtered_results) + 24) // 25 if filtered_results else 1,
            'request': request,
        }
        
        # Render HTML template
        from django.template.loader import render_to_string
        html_string = render_to_string(
            'portal_management/exams/paper_results_filter_pdf.html', 
            context
        )
        
        # Configure fonts
        from weasyprint import HTML
        from weasyprint.text.fonts import FontConfiguration
        
        font_config = FontConfiguration()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        safe_subject = paper.subject.code.replace(' ', '_')
        safe_filter = filter_type.replace(' ', '_')
        filename = f"{safe_subject}_Paper{paper.paper_number}_{safe_filter}_{date.today()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        try:
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config,
                presentational_hints=True,
            )
            return response
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect('management:paper_results_filter', paper_pk=paper.pk)
    
    def _get_filter_label(self, filter_type):
        """Get human-readable filter label."""
        labels = {
            'all': 'All Students',
            'with_scores': 'Students With Scores',
            'without_scores': 'Students Without Scores',
        }
        return labels.get(filter_type, 'All Students')
    
    def _get_school_info(self, school_profile, educational_level):
        """Get school information from profile or fallback to settings."""
        if school_profile:
            return {
                'name': school_profile.name,
                'address': school_profile.address,
                'phone': school_profile.get_contact_phone(),
                'email': school_profile.email,
                'motto': school_profile.motto,
                'registration_number': school_profile.registration_number,
                'logo': school_profile.logo.url if school_profile.logo else None,
                'contact_person': school_profile.get_contact_name(),
            }
        else:
            return {
                'name': getattr(settings, 'SCHOOL_NAME', 'School Management System'),
                'address': getattr(settings, 'SCHOOL_ADDRESS', ''),
                'phone': getattr(settings, 'SCHOOL_PHONE', ''),
                'email': getattr(settings, 'SCHOOL_EMAIL', ''),
                'motto': getattr(settings, 'SCHOOL_MOTTO', ''),
                'registration_number': getattr(settings, 'SCHOOL_REGISTRATION_NUMBER', ''),
                'logo': None,
                'contact_person': 'Headmaster',
            }