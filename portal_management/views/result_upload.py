"""
portal_management/views/result_upload.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
UploadResultsView — session-wide Excel upload that parses a file
produced by DownloadResultTemplateView and saves StudentPaperScore records.

Parse algorithm (sentinel-based — guaranteed row-4 scan):
  1. Read row 4 exactly (the hidden sentinel row)
  2. Any column whose value starts with 'PID:' identifies a SubjectExamPaper
  3. Column A (index 0) of each data row (row 5+) = registration number
  4. Columns mapped by PID: = marks for the corresponding paper
  5. Empty marks cell → skip (do not delete existing score)
  6. Bad marks value  → row-level error (no partial save; rest of row still processed)

All saves are wrapped in a single atomic transaction so a mid-upload
crash leaves the database unchanged.
"""

import io
import logging
from decimal import Decimal, InvalidOperation

import openpyxl

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import View

from core.mixins import ManagementRequiredMixin
from core.models import (
    ExamSession, StudentPaperScore, SubjectExamPaper,
)
from portal_management.views.exam_views import _enrolled_students


logger = logging.getLogger(__name__)


class UploadResultsView(ManagementRequiredMixin, View):
    """
    Parse a filled-in DownloadResultTemplateView Excel file and
    bulk-create / bulk-update StudentPaperScore records.

    GET  — render the upload form (with subject-wise download shortcuts)
    POST — handle file upload, return JSON on AJAX or redirect on normal POST
    """
    template_name = 'portal_management/exams/upload_results.html'

    # ── GET ───────────────────────────────────────────────────────────────

    def get(self, request, session_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        if session.status == 'published':
            messages.error(
                request,
                'This session is published. '
                'Unpublish it first if corrections are needed.'
            )
            return redirect('management:exam_session_detail', pk=session_pk)

        subjects = (
            SubjectExamPaper.objects
            .filter(exam_session=session)
            .values('subject_id', 'subject__name', 'subject__short_name')
            .distinct()
            .order_by('subject__name')
        )
        return render(request, self.template_name, {
            'session':  session,
            'subjects': subjects,
        })

    # ── POST ──────────────────────────────────────────────────────────────

    def post(self, request, session_pk):
        session  = get_object_or_404(ExamSession, pk=session_pk)
        is_ajax  = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        # Guard: published
        if session.status == 'published':
            return self._error(
                is_ajax, request, session_pk,
                'Cannot upload results to a published session.'
            )

        # Guard: file present
        xl_file = request.FILES.get('excel_file')
        if not xl_file:
            return self._error(
                is_ajax, request, session_pk,
                'No file uploaded. Please select an Excel file.'
            )

        # Guard: extension
        ext = xl_file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls'):
            return self._error(
                is_ajax, request, session_pk,
                'Invalid format — please upload an .xlsx or .xls file.'
            )

        # Load workbook
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(xl_file.read()), data_only=True
            )
            ws = wb.active
        except Exception as exc:
            return self._error(
                is_ajax, request, session_pk,
                f'Cannot read Excel file: {exc}'
            )

        # ── Scan row 4 for PID: sentinel markers ──────────────────────────
        paper_col_map: dict[int, SubjectExamPaper] = {}  # 0-based col → paper

        try:
            sentinel_row = list(
                ws.iter_rows(min_row=4, max_row=4, values_only=True)
            )[0]
        except IndexError:
            return self._error(
                is_ajax, request, session_pk,
                'Template too short — sentinel row 4 not found. '
                'Download a fresh template.'
            )

        for col_idx, cell_val in enumerate(sentinel_row):
            raw = str(cell_val).strip() if cell_val is not None else ''
            if raw.startswith('PID:'):
                try:
                    paper_id = int(raw[4:])
                    paper    = SubjectExamPaper.objects.get(
                        pk=paper_id, exam_session=session
                    )
                    paper_col_map[col_idx] = paper
                except (ValueError, SubjectExamPaper.DoesNotExist):
                    logger.warning(
                        'UploadResults: unknown paper ID %s in session %s',
                        raw, session_pk
                    )

        if not paper_col_map:
            return self._error(
                is_ajax, request, session_pk,
                'Paper ID row (row 4) not found or contains no valid PID: '
                'markers. Please download a fresh template — do not modify '
                'or delete row 4.'
            )

        # ── Build student lookup (reg_no → Student) ────────────────────────
        student_map = {
            s.registration_number: s
            for s in _enrolled_students(session)
            if s.registration_number
        }

        # ── Process data rows (row 5 onward) ──────────────────────────────
        saved_rows  = 0
        skipped     = 0
        error_rows  = []

        with transaction.atomic():
            for row_offset, row in enumerate(
                ws.iter_rows(min_row=5, values_only=True)
            ):
                excel_row = 5 + row_offset

                # Skip fully empty rows
                if not row or all(
                    cell is None or str(cell).strip() == ''
                    for cell in row
                ):
                    skipped += 1
                    continue

                # Column A = registration number
                reg_no = str(row[0]).strip() if row[0] else ''
                if not reg_no or reg_no.lower() == 'none':
                    skipped += 1
                    continue

                student = student_map.get(reg_no)
                if not student:
                    error_rows.append(
                        f'Row {excel_row}: '
                        f'"{reg_no}" — student not found or not enrolled '
                        f'in {session.class_level} / {session.academic_year}.'
                    )
                    continue

                # Process each paper column for this student
                row_saved = 0
                for col_idx, paper in paper_col_map.items():
                    if col_idx >= len(row):
                        continue
                    cell_val = row[col_idx]
                    raw_str  = str(cell_val).strip() if cell_val is not None else ''

                    if raw_str == '':
                        # Empty → leave existing score untouched
                        continue

                    try:
                        marks = Decimal(raw_str.replace(',', '.'))
                    except InvalidOperation:
                        error_rows.append(
                            f'Row {excel_row} ({reg_no}): '
                            f'invalid marks value "{cell_val}" '
                            f'for {paper.subject.name} P{paper.paper_number}.'
                        )
                        continue

                    if marks < 0:
                        error_rows.append(
                            f'Row {excel_row} ({reg_no}): '
                            f'marks cannot be negative ({marks}).'
                        )
                        continue

                    if marks > paper.max_marks:
                        error_rows.append(
                            f'Row {excel_row} ({reg_no}): '
                            f'marks {marks} exceed maximum '
                            f'{paper.max_marks} for '
                            f'{paper.subject.name} P{paper.paper_number}.'
                        )
                        continue

                    try:
                        StudentPaperScore.objects.update_or_create(
                            student=student,
                            exam_paper=paper,
                            defaults={'marks': marks},
                        )
                        row_saved += 1
                    except ValidationError as exc:
                        error_rows.append(
                            f'Row {excel_row} ({reg_no}): {exc}'
                        )
                    except Exception as exc:
                        logger.error(
                            'UploadResults save error row=%s student=%s: %s',
                            excel_row, reg_no, exc, exc_info=True,
                        )
                        error_rows.append(
                            f'Row {excel_row} ({reg_no}): '
                            f'unexpected error — {exc}'
                        )

                if row_saved > 0:
                    saved_rows += 1
                elif not any(
                    col_idx < len(row) and row[col_idx] not in (None, '')
                    for col_idx in paper_col_map
                ):
                    skipped += 1

        # ── Summary message ───────────────────────────────────────────────
        parts = [f'{saved_rows} student row(s) saved.']
        if skipped:
            parts.append(f'{skipped} empty row(s) skipped.')
        if error_rows:
            parts.append(f'{len(error_rows)} error(s).')
        msg = '  '.join(parts)

        summary = {
            'saved_rows':  saved_rows,
            'skipped_rows': skipped,
            'errors':      error_rows,
            'total':       saved_rows + skipped + len(error_rows),
        }

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': msg,
                'summary': summary,
            })

        if error_rows:
            messages.warning(request, msg)
        else:
            messages.success(request, msg)

        return redirect('management:exam_session_detail', pk=session_pk)

    # ── Helper ────────────────────────────────────────────────────────────

    @staticmethod
    def _error(is_ajax, request, session_pk, msg, status=400):
        if is_ajax:
            return JsonResponse({'success': False, 'message': msg}, status=status)
        messages.error(request, msg)
        return redirect('management:upload_results', session_pk=session_pk)
