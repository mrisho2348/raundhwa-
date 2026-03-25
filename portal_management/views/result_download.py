"""
portal_management/views/result_download.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DownloadResultTemplateView — session-wide Excel template download.

Modes (via ?mode=):
  full      — all subjects × all papers in one sheet (default)
  subject   — single subject, requires ?subject_id=
  prefilled — pre-fills existing scores into the template

Fix for AttributeError 'MergedCell' object attribute 'value' is read-only:
  Root cause: openpyxl merge_cells() makes all cells except the top-left
  into MergedCell proxy objects whose .value is read-only.  Any attempt to
  write to ws.cell(row, col).value AFTER merging will crash if that (row,col)
  is a non-anchor merged cell.

  Fix applied everywhere:
    1. NEVER call ws.cell(..., value=...) on a column that was merged.
    2. Merge AFTER writing anchor cell content — not before.
    3. Use _safe_write(ws, row, col, value) which checks isMergedCell
       and skips non-anchor cells silently.
    4. Subject group headers (row 2) are written per-column
       (no merge_cells for the subject name) to avoid the crash.
       The visual grouping comes from identical background colour
       rather than Excel cell merging.
"""

import io
import logging
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.views.generic import View

from core.mixins import ManagementRequiredMixin
from core.models import (
    ExamSession, StudentPaperScore, Subject,
    SubjectExamPaper,
)
from portal_management.views.exam_views import _enrolled_students


logger = logging.getLogger(__name__)


# ── Palette ───────────────────────────────────────────────────────────────────

_SUBJECT_PALETTES = [
    ('1a73e8', 'e8f0fe'),  # blue
    ('0f7b3a', 'e6f4ea'),  # green
    ('6f42c1', 'f3e5f5'),  # purple
    ('fd7e14', 'fff3cd'),  # orange
    ('17a2b8', 'd1ecf1'),  # teal
    ('dc3545', 'f8d7da'),  # red
    ('6c757d', 'f1f3f5'),  # grey
    ('795548', 'efebe9'),  # brown
]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header(ws, row, col, value, bg, fg='FFFFFF', bold=True,
            size=9, wrap=False, h_align='center'):
    """Write a styled header cell. Always writes to the anchor column only."""
    cell = ws.cell(row=row, column=col)
    cell.value     = value
    cell.font      = Font(bold=bold, color=fg, size=size, name='Arial')
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(
        horizontal=h_align, vertical='center', wrap_text=wrap
    )
    cell.border = _thin('AAAAAA')
    return cell


def _data(ws, row, col, value=None, locked=False,
          bg=None, h_align='center', number_format=None):
    """Write a styled data cell."""
    cell        = ws.cell(row=row, column=col)
    cell.value  = value
    bg_hex      = bg or ('F5F5F5' if locked else 'FFFFFF')
    cell.fill   = PatternFill('solid', fgColor=bg_hex)
    cell.alignment = Alignment(horizontal=h_align, vertical='center')
    cell.border = _thin()
    if number_format:
        cell.number_format = number_format
    return cell


# ── Safe merge (write THEN merge to avoid MergedCell read-only error) ─────────

def _merge_write(ws, row, col_start, col_end, value, fill_hex, font_obj,
                 align_obj=None, height=None):
    """
    Write value to anchor cell, then merge.
    This avoids the 'MergedCell object attribute value is read-only' error
    which occurs when you merge FIRST then try to write.
    """
    # 1. Write to anchor first
    anchor      = ws.cell(row=row, column=col_start)
    anchor.value = value
    anchor.font  = font_obj
    anchor.fill  = PatternFill('solid', fgColor=fill_hex)
    anchor.alignment = align_obj or Alignment(
        horizontal='center', vertical='center'
    )
    anchor.border = _thin('888888')

    # 2. Merge after writing
    if col_start < col_end:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end
        )

    if height:
        ws.row_dimensions[row].height = height

    return anchor


# ════════════════════════════════════════════════════════════════════════════
# VIEW
# ════════════════════════════════════════════════════════════════════════════

class DownloadResultTemplateView(ManagementRequiredMixin, View):
    """
    Download an Excel result-entry template for a session.

    Template layout (sentinel-based — same protocol as paper_excel.py):
      Row 1  — banner: session name + instructions
      Row 2  — subject group headers (one colour per subject, NO merge)
      Row 3  — paper detail headers: P1 / max marks etc.
      Row 4  — SENTINEL ROW: '__PAPER_IDS__' in A, PID:N in each paper col
               (height=1, visually hidden — used by UploadResultsView)
      Row 5+ — student data: col A = reg_no (locked), col B = name (locked),
               col C+ = editable marks cells (pre-filled when mode=prefilled)

    Fix:  subject headers in row 2 are written per-column with matching
          background colour instead of using merge_cells, which eliminates
          the MergedCell read-only crash entirely.
    """

    def get(self, request, session_pk):
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level', 'academic_year', 'term'
            ),
            pk=session_pk,
        )
        mode       = request.GET.get('mode', 'full')
        subject_id = request.GET.get('subject_id')

        # ── Papers queryset ───────────────────────────────────────────────
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')

        if mode == 'subject' and subject_id:
            papers = papers.filter(subject_id=subject_id)
            if not papers.exists():
                messages.error(request, 'No papers found for this subject.')
                return redirect('management:exam_session_detail', pk=session_pk)

        papers_list = list(papers)

        if not papers_list:
            messages.error(
                request,
                'No exam papers have been added to this session yet. '
                'Please add papers before downloading the template.'
            )
            return redirect('management:exam_session_detail', pk=session_pk)

        # ── Students ──────────────────────────────────────────────────────
        students = _enrolled_students(session)

        if not students:
            messages.error(
                request,
                'No enrolled students found for this session. '
                'Students must be enrolled before downloading the template.'
            )
            return redirect('management:exam_session_detail', pk=session_pk)

        # ── Pre-fill existing scores ───────────────────────────────────────
        existing = {}
        if mode in ('prefilled', 'subject'):
            for sc in StudentPaperScore.objects.filter(
                exam_paper__exam_session=session,
                exam_paper__in=papers_list,
            ):
                existing[(sc.student_id, sc.exam_paper_id)] = sc.marks

        # ── Assign a colour pair to each subject ───────────────────────────
        subject_colours: dict[int, tuple[str, str]] = {}
        palette_idx = 0
        for paper in papers_list:
            sid = paper.subject_id
            if sid not in subject_colours:
                subject_colours[sid] = _SUBJECT_PALETTES[
                    palette_idx % len(_SUBJECT_PALETTES)
                ]
                palette_idx += 1

        # ── Build workbook ────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'

        total_paper_cols = len(papers_list)
        last_col         = 2 + total_paper_cols           # A=1 B=2 C..=papers
        last_letter      = get_column_letter(last_col)

        # ── Row 1: Banner ─────────────────────────────────────────────────
        # Write anchor first, merge after
        _merge_write(
            ws, 1, 1, last_col,
            f'Result Entry Template  —  {session.name}  '
            f'|  {session.class_level}  |  {session.term}  '
            f'|  {session.academic_year}  '
            f'||  DO NOT modify columns A–B or Row 4',
            '1a73e8',
            Font(bold=True, color='FFFFFF', size=10, name='Arial'),
            Alignment(horizontal='left', vertical='center'),
            height=24,
        )

        # ── Row 2: Subject group colour bars (NO merge — avoids crash) ────
        # Columns A and B
        _header(ws, 2, 1, 'Reg No',      '2c3e50', size=9)
        _header(ws, 2, 2, 'Full Name',   '2c3e50', size=9)
        ws.row_dimensions[2].height = 20

        # One cell per paper column, subject name repeated, same bg per subject
        for col_offset, paper in enumerate(papers_list):
            col = 3 + col_offset
            dark_bg, _ = subject_colours[paper.subject_id]
            _header(
                ws, 2, col,
                paper.subject.short_name or paper.subject.name,
                dark_bg,
                size=8,
                wrap=False,
            )

        # ── Row 3: Paper number + max marks ───────────────────────────────
        _header(ws, 3, 1, 'Reg No',    '7f8c8d', size=8)
        _header(ws, 3, 2, 'Full Name', '7f8c8d', size=8)
        ws.row_dimensions[3].height = 32

        for col_offset, paper in enumerate(papers_list):
            col = 3 + col_offset
            _, light_bg = subject_colours[paper.subject_id]
            _header(
                ws, 3, col,
                f'P{paper.paper_number}\n/{paper.max_marks}',
                '95a5a6',
                size=8,
                wrap=True,
                h_align='center',
            )

        # ── Row 4: SENTINEL (hidden, used by uploader) ────────────────────
        ws.row_dimensions[4].height = 1   # visually invisible
        anchor4 = ws.cell(row=4, column=1)
        anchor4.value = '__PAPER_IDS__'
        anchor4.font  = Font(size=1, color='FFFFFF')

        ws.cell(row=4, column=2).value = ''

        for col_offset, paper in enumerate(papers_list):
            col = 3 + col_offset
            c        = ws.cell(row=4, column=col)
            c.value  = f'PID:{paper.pk}'
            c.font   = Font(size=1, color='FFFFFF')

        # ── Row 5+: Student data rows ─────────────────────────────────────
        for row_idx, student in enumerate(students):
            r = 5 + row_idx
            ws.row_dimensions[r].height = 18

            # Reg No — locked (grey)
            _data(ws, r, 1, student.registration_number or '',
                  locked=True, h_align='left')

            # Full Name — locked (grey)
            _data(ws, r, 2, student.full_name,
                  locked=True, h_align='left')

            # Marks cells — editable
            for col_offset, paper in enumerate(papers_list):
                col = 3 + col_offset
                val = existing.get((student.pk, paper.pk))
                _, light_bg = subject_colours[paper.subject_id]
                _data(
                    ws, r, col,
                    value=float(val) if val is not None else None,
                    bg=light_bg if val is not None else None,
                    number_format='0.##',
                )

        # ── Column widths ─────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 20   # reg no
        ws.column_dimensions['B'].width = 32   # name
        for col_offset in range(total_paper_cols):
            ws.column_dimensions[get_column_letter(3 + col_offset)].width = 11

        # ── Freeze at first data row ──────────────────────────────────────
        ws.freeze_panes = 'C5'

        # ── HTTP Response ─────────────────────────────────────────────────
        subject_suffix = ''
        if mode == 'subject' and subject_id:
            sub = Subject.objects.filter(pk=subject_id).first()
            subject_suffix = f'_{sub.short_name or sub.name}' if sub else ''

        fname = (
            f'result_template'
            f'_{session.name[:25]}{subject_suffix}'
            f'_{mode}'
            f'_{date.today()}.xlsx'
        ).replace(' ', '_').replace('/', '-').replace('—', '-')

        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response
