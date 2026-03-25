"""
portal_management/views/paper_excel.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Per-paper Excel upload/download views.

The two views share a single sentinel protocol so upload always
knows exactly which row contains data regardless of how many
header rows the download template writes.

Sentinel protocol (same as DownloadResultTemplateView / UploadResultsView):
  Row 1 — title bar
  Row 2 — session info
  Row 3 — paper details
  Row 4 — filter info (omitted if no filters)
  Row N — column headers  (visible, styled)
  Row N+1 — SENTINEL ROW  value '__DATA_START__' in column A (row_dimensions height=1)
  Row N+2+ — student data

The uploader scans for '__DATA_START__' to find where data begins.
It never guesses based on header text — it always uses the sentinel.
"""

import io
import logging
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.views.generic import View

from core.mixins import ManagementRequiredMixin
from core.models import (
    GradingScale, StudentPaperScore, SubjectExamPaper,
)
from portal_management.views.exam_views import _enrolled_students


logger = logging.getLogger(__name__)

# ── Sentinel value written into / scanned for in the hidden row ───────────
_SENTINEL = '__DATA_START__'


# ════════════════════════════════════════════════════════════════════════════
# HELPER — cell styling
# ════════════════════════════════════════════════════════════════════════════

def _h(ws, row, col, value, bg='1a73e8', fg='FFFFFF', bold=True, size=10, wrap=False):
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    thin = Side(style='thin', color='CCCCCC')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return cell


def _d(ws, row, col, value='', locked=False, left=False):
    """Write a data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    thin = Side(style='thin', color='DDDDDD')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell.alignment = Alignment(
        horizontal='left' if left else 'center',
        vertical='center',
    )
    if locked:
        cell.fill = PatternFill('solid', fgColor='F0F0F0')
    return cell


# ════════════════════════════════════════════════════════════════════════════
# DOWNLOAD — per-paper template with optional filters
# ════════════════════════════════════════════════════════════════════════════

class DownloadPaperTemplateView(ManagementRequiredMixin, View):
    """
    Download an Excel template for a single SubjectExamPaper.

    Query params (all optional):
      status_filter  — all | with_scores | without_scores
      gender_filter  — all | male | female
      letter_filter  — all | A–Z  (first letter of student name)
      search         — free-text search on name / reg_no

    Template layout:
      Row 1  — title (bold, dark green)
      Row 2  — session info
      Row 3  — paper details
      Row 4  — filter summary (only if filters are active)
      Row ?  — column headers (#, Reg No, Name, Gender, Marks)
      Row ?+1 — SENTINEL  '__DATA_START__' in col A  (height=1, hidden)
      Row ?+2+ — one row per student (locked: #, reg_no, name, gender; editable: marks)
    """

    def get(self, request, paper_pk):
        paper   = get_object_or_404(SubjectExamPaper, pk=paper_pk)
        session = paper.exam_session

        # ── Filters ────────────────────────────────────────────────────────
        status_filter = request.GET.get('status_filter', 'all')
        gender_filter = request.GET.get('gender_filter', 'all')
        letter_filter = request.GET.get('letter_filter', 'all')
        search_term   = request.GET.get('search', '').strip().lower()

        students = _enrolled_students(session)

        # Existing scores
        scores = {
            sc.student_id: sc.marks
            for sc in StudentPaperScore.objects.filter(
                exam_paper=paper, student__in=students
            )
        }

        # Apply filters
        rows = []
        for student in students:
            marks     = scores.get(student.pk)
            has_score = marks is not None

            if status_filter == 'with_scores'    and not has_score: continue
            if status_filter == 'without_scores' and has_score:     continue
            if gender_filter != 'all'            and student.gender != gender_filter: continue
            if letter_filter != 'all':
                first = (student.full_name or ' ')[0].upper()
                if first != letter_filter:
                    continue
            if search_term:
                haystack = (
                    (student.full_name or '').lower() +
                    (student.registration_number or '').lower()
                )
                if search_term not in haystack:
                    continue

            rows.append({
                'student': student,
                'marks': marks,
                'has_score': has_score,
            })

        # ── Build workbook ─────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'P{paper.paper_number}'

        r = 1  # current row pointer

        # Row 1 — title
        ws.merge_cells(f'A{r}:E{r}')
        _h(ws, r, 1,
           f'RESULT ENTRY TEMPLATE — {paper.subject.name.upper()} PAPER {paper.paper_number}',
           bg='1a472a', size=13)
        ws.row_dimensions[r].height = 30
        r += 1

        # Row 2 — session info
        session_info = (
            f'Session: {session.name}  |  '
            f'Class: {session.class_level.name}'
            + (f' — {session.stream_class}' if session.stream_class else '')
            + f'  |  {session.term}  |  {session.academic_year}'
        )
        ws.merge_cells(f'A{r}:E{r}')
        cell = ws.cell(row=r, column=1, value=session_info)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill('solid', fgColor='E8F0FE')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22
        r += 1

        # Row 3 — paper details
        paper_info = (
            f'Subject: {paper.subject.name} ({paper.subject.code})  |  '
            f'Paper {paper.paper_number}  |  Max Marks: {paper.max_marks}'
        )
        if paper.duration_minutes:
            paper_info += f'  |  Duration: {paper.duration_minutes} min'
        if paper.exam_date:
            paper_info += f'  |  Date: {paper.exam_date.strftime("%d %b %Y")}'
        ws.merge_cells(f'A{r}:E{r}')
        cell = ws.cell(row=r, column=1, value=paper_info)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill('solid', fgColor='F0F0F0')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        # Row 4 — filter summary (only when filters active)
        active_filters = []
        if status_filter != 'all':
            active_filters.append(
                'With Scores' if status_filter == 'with_scores' else 'Without Scores'
            )
        if gender_filter != 'all': active_filters.append(f'Gender: {gender_filter}')
        if letter_filter != 'all': active_filters.append(f'Starts: {letter_filter}')
        if search_term:             active_filters.append(f'Search: "{search_term}"')

        if active_filters:
            ws.merge_cells(f'A{r}:E{r}')
            cell = ws.cell(
                row=r, column=1,
                value=f'Filters: {" | ".join(active_filters)}  —  {len(rows)} student(s)'
            )
            cell.font = Font(size=9, color='856404')
            cell.fill = PatternFill('solid', fgColor='FFF3CD')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[r].height = 18
            r += 1

        # Column headers (no grade column)
        headers = [
            ('#', 8),
            ('Registration Number', 20),
            ('Student Name', 30),
            ('Gender', 10),
            (f'Marks (0-{paper.max_marks})', 16),
        ]

        for col_idx, (label, width) in enumerate(headers, 1):
            _h(ws, r, col_idx, label, bg='2C3E50')
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[r].height = 25
        r += 1

        # ── SENTINEL ROW ─────────────────────────────────────────────────
        # Uploader scans for this value in column A to find data start.
        ws.cell(row=r, column=1, value=_SENTINEL)
        # Also store paper PK so uploader can verify correct template
        ws.cell(row=r, column=2, value=f'PAPER_PK:{paper.pk}')
        ws.row_dimensions[r].height = 1   # visually hidden
        r += 1

        # ── Student rows ──────────────────────────────────────────────────
        for idx, item in enumerate(rows, 1):
            student = item['student']

            _d(ws, r, 1, idx, locked=True)
            _d(ws, r, 2, student.registration_number or '', locked=True)
            _d(ws, r, 3, student.full_name, locked=True, left=True)
            _d(ws, r, 4,
               student.get_gender_display() if student.gender else '', locked=True)

            # Marks cell — prefilled if score exists, editable background
            marks_cell = ws.cell(
                row=r, column=5,
                value=float(item['marks']) if item['has_score'] else None
            )
            marks_cell.number_format = '0.##'
            thin = Side(style='thin', color='DDDDDD')
            marks_cell.border = Border(
                top=thin, left=thin, right=thin, bottom=thin
            )
            marks_cell.alignment = Alignment(horizontal='center', vertical='center')
            # Green tint for prefilled scores
            if item['has_score']:
                marks_cell.fill = PatternFill('solid', fgColor='E8F5E9')

            r += 1

        # Freeze at the first data row
        ws.freeze_panes = f'A{r - len(rows)}'

        # ── Response ──────────────────────────────────────────────────────
        filename_parts = [
            f'paper{paper.paper_number}',
            (paper.subject.short_name or paper.subject.name).replace(' ', '_')[:15],
            session.name[:20].replace(' ', '_'),
        ]
        if active_filters:
            filename_parts.append('filtered')
        filename = '_'.join(filename_parts) + f'_{date.today()}.xlsx'
        filename = filename.replace('/', '_').replace('—', '-')

        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ════════════════════════════════════════════════════════════════════════════
# UPLOAD — parse the per-paper template and save scores
# ════════════════════════════════════════════════════════════════════════════

class BulkExcelUploadView(ManagementRequiredMixin, View):
    """
    Upload a filled DownloadPaperTemplateView Excel file to save scores
    for a single SubjectExamPaper.

    Parse algorithm (sentinel-based — no header guessing):
      1. Scan all rows for a cell whose value == '__DATA_START__'
      2. The row immediately after the sentinel is the first student row
      3. Column layout: A=#, B=reg_no, C=name, D=gender, E=marks
      4. Marks in column E are saved / updated / cleared
      5. If PAPER_PK:N is found in the sentinel row column B, verify it
         matches paper_pk from the URL to catch wrong-template uploads

    Validation:
      - Marks must be between 0 and paper.max_marks (inclusive)
      - Empty marks = clear score
      - Invalid marks = error
      - Student not enrolled = error
    """

    def post(self, request, paper_pk):
        paper   = get_object_or_404(SubjectExamPaper, pk=paper_pk)
        session = paper.exam_session

        # ── Guard: published session ───────────────────────────────────────
        if session.status == 'published':
            return JsonResponse({
                'success': False,
                'message': 'Cannot upload results to a published session.',
            }, status=400)

        # ── Guard: file present and valid extension ────────────────────────
        xl_file = request.FILES.get('excel_file')
        if not xl_file:
            return JsonResponse({
                'success': False,
                'message': 'No file uploaded. Please select an Excel file.',
            }, status=400)

        ext = xl_file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls'):
            return JsonResponse({
                'success': False,
                'message': 'Invalid format. Upload an .xlsx or .xls file.',
            }, status=400)

        # ── Load workbook ──────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(xl_file.read()), data_only=True
            )
            ws = wb.active
        except Exception as exc:
            return JsonResponse({
                'success': False,
                'message': f'Cannot read Excel file: {exc}',
            }, status=400)

        # ── Find sentinel row ──────────────────────────────────────────────
        sentinel_row = None
        all_rows = list(ws.iter_rows(values_only=True))

        for row_idx, row in enumerate(all_rows):
            if row and str(row[0]).strip() == _SENTINEL:
                sentinel_row = row_idx  # 0-based index into all_rows
                break

        if sentinel_row is None:
            return JsonResponse({
                'success': False,
                'message': (
                    'Template format not recognised — sentinel row missing. '
                    'Please download a fresh template from the "Download Template" button '
                    'and do not delete or modify row A that contains internal markers.'
                ),
            }, status=400)

        # ── Optional: verify paper PK in sentinel row ──────────────────────
        sentinel_row_data = all_rows[sentinel_row]
        if len(sentinel_row_data) >= 2 and sentinel_row_data[1]:
            pk_cell = str(sentinel_row_data[1]).strip()
            if pk_cell.startswith('PAPER_PK:'):
                template_paper_pk = int(pk_cell.replace('PAPER_PK:', ''))
                if template_paper_pk != paper.pk:
                    template_paper = SubjectExamPaper.objects.filter(
                        pk=template_paper_pk
                    ).first()
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Wrong template uploaded. '
                            f'This template belongs to '
                            f'"{template_paper or f"Paper PK={template_paper_pk}"}", '
                            f'but you are uploading for '
                            f'"{paper.subject.name} Paper {paper.paper_number}". '
                            f'Please download the correct template.'
                        ),
                    }, status=400)

        # ── Build student lookup ───────────────────────────────────────────
        enrolled  = _enrolled_students(session)
        stud_map  = {
            s.registration_number: s
            for s in enrolled
            if s.registration_number
        }

        # ── Grading for the response payload ──────────────────────────────
        grading_scale = list(
            GradingScale.objects.filter(
                education_level=session.class_level.educational_level
            ).order_by('-min_mark')
        )

        def calc_grade(marks):
            pct = float(marks) / float(paper.max_marks) * 100
            for gs in grading_scale:
                if float(gs.min_mark) <= pct <= float(gs.max_mark):
                    return gs.grade
            return 'F'

        # ── Process data rows ──────────────────────────────────────────────
        data_rows       = all_rows[sentinel_row + 1:]   # rows after sentinel
        saved           = 0
        errors          = []
        updated_scores  = []
        skipped         = 0
        processed       = 0

        with transaction.atomic():
            for offset, row in enumerate(data_rows):
                row_num = sentinel_row + offset + 2  # 1-based Excel row number

                # Skip fully empty rows
                if not row or all(
                    cell is None or str(cell).strip() == ''
                    for cell in row
                ):
                    skipped += 1
                    continue

                processed += 1

                # Column B (index 1) = registration number
                reg_no    = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                # Column E (index 4) = marks
                marks_raw = row[4] if len(row) > 4 else None

                if not reg_no:
                    errors.append(
                        f'Row {row_num}: Missing registration number.'
                    )
                    continue

                student = stud_map.get(reg_no)
                if not student:
                    errors.append(
                        f'Row {row_num}: Student "{reg_no}" not found or not enrolled '
                        f'in {session.class_level} for {session.academic_year}.'
                    )
                    continue

                # ── Empty marks → clear existing score ────────────────────
                if marks_raw is None or str(marks_raw).strip() == '':
                    deleted = StudentPaperScore.objects.filter(
                        student=student, exam_paper=paper
                    ).delete()
                    if deleted[0]:
                        saved += 1
                        updated_scores.append({
                            'student_id':   student.pk,
                            'student_name': student.full_name,
                            'marks':        None,
                            'cleared':      True,
                        })
                    continue

                # ── Parse and validate marks ───────────────────────────────
                try:
                    marks_str = str(marks_raw).strip().replace(',', '.')
                    marks     = Decimal(marks_str)
                except (InvalidOperation, ValueError):
                    errors.append(
                        f'Row {row_num} ({reg_no}): '
                        f'Invalid marks value "{marks_raw}". '
                        f'Expected a number between 0 and {paper.max_marks}.'
                    )
                    continue

                # Validate marks range (0 to max_marks)
                if marks < 0:
                    errors.append(
                        f'Row {row_num} ({reg_no}): Marks cannot be negative ({marks}).'
                    )
                    continue

                if marks > paper.max_marks:
                    errors.append(
                        f'Row {row_num} ({reg_no}): '
                        f'Marks ({marks}) exceed maximum ({paper.max_marks}).'
                    )
                    continue

                # ── Save score ─────────────────────────────────────────────
                try:
                    score, created = StudentPaperScore.objects.update_or_create(
                        student=student,
                        exam_paper=paper,
                        defaults={'marks': marks},
                    )
                    grade = calc_grade(marks)
                    pct   = float(marks) / float(paper.max_marks) * 100

                    saved += 1
                    updated_scores.append({
                        'student_id':   student.pk,
                        'student_name': student.full_name,
                        'marks':        float(marks),
                        'percentage':   round(pct, 1),
                        'grade':        grade,
                        'created':      created,
                    })

                except ValidationError as exc:
                    errors.append(f'Row {row_num} ({reg_no}): {exc}')
                except Exception as exc:
                    logger.error(
                        'BulkExcelUpload save error row=%s student=%s: %s',
                        row_num, reg_no, exc, exc_info=True
                    )
                    errors.append(f'Row {row_num} ({reg_no}): Unexpected error — {exc}')

        # ── Post-upload stats ──────────────────────────────────────────────
        all_scores  = StudentPaperScore.objects.filter(
            exam_paper=paper, student__in=enrolled
        )
        marks_list  = [float(s.marks) for s in all_scores]
        total_stud  = len(enrolled)
        scored      = len(marks_list)

        stats = {
            'total_students':        total_stud,
            'students_with_scores':  scored,
            'students_without_scores': total_stud - scored,
            'average':    round(sum(marks_list) / scored, 2) if scored else 0,
            'highest':    max(marks_list)   if marks_list else 0,
            'lowest':     min(marks_list)   if marks_list else 0,
            'completion_percentage': round(scored / total_stud * 100, 1) if total_stud else 0,
        }

        # ── Build response message ─────────────────────────────────────────
        if saved:
            msg = f'Successfully processed {saved} score(s).'
        else:
            msg = 'No scores were processed.'

        if errors:
            msg += f' {len(errors)} error(s) — see details below.'
        if skipped:
            msg += f' {skipped} empty row(s) skipped.'

        response_data = {
            'success':         True,
            'message':         msg,
            'saved':           saved,
            'total_processed': processed,
            'skipped_rows':    skipped,
            'errors':          errors,
            'updated_scores':  updated_scores[:50],
            'stats':           stats,
        }

        if not saved and not errors:
            response_data['warning'] = (
                'No valid data found. '
                'Make sure you downloaded the template for this specific paper '
                'and entered marks in column E (values between 0 and {paper.max_marks}).'
            )
            response_data['message'] = response_data['warning']

        return JsonResponse(response_data)