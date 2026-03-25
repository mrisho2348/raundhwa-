# views/export_result_analytics.py

from django.conf import settings
from django.views import View
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from decimal import Decimal
import re
import logging
from core.models import (
    ExamSession, StudentEnrollment, Student, StudentExamMetrics,
    StudentExamPosition, StudentSubjectResult, StudentPaperScore,
    SubjectExamPaper, EducationalLevel, StudentStreamAssignment, Subject,
    GradingScale, DivisionScale, Combination, StudentCombinationAssignment
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Shared style helpers
# ---------------------------------------------------------------------------

BRAND_BLUE    = "1A73E8"
BRAND_DARK    = "1A237E"
ACCENT_GOLD   = "F9AB00"
HEADER_BG     = "1A73E8"
ALT_ROW_BG    = "EAF1FB"
TOTAL_ROW_BG  = "D2E3FC"
WHITE         = "FFFFFF"
LIGHT_GREY    = "F8F9FA"
BORDER_COLOUR = "BDC1C6"

def _thin_border(colour=BORDER_COLOUR):
    s = Side(style="thin", color=colour)
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FONT      = Font(name="Arial", bold=True,  color=WHITE,      size=11)
SUB_HEADER_FONT  = Font(name="Arial", bold=True,  color=BRAND_DARK, size=12)
SCHOOL_FONT      = Font(name="Arial", bold=True,  color=BRAND_BLUE, size=16)
INFO_FONT        = Font(name="Arial",              color="5F6368",   size=9)
SESSION_FONT     = Font(name="Arial", italic=True, color="5F6368",   size=9)
BODY_FONT        = Font(name="Arial",                                size=10)
BODY_BOLD_FONT   = Font(name="Arial", bold=True,                     size=10)
FOOTER_FONT      = Font(name="Arial", italic=True, color="9AA0A6",   size=8)

HEADER_FILL      = PatternFill("solid", fgColor=HEADER_BG)
ALT_FILL         = PatternFill("solid", fgColor=ALT_ROW_BG)
TOTAL_FILL       = PatternFill("solid", fgColor=TOTAL_ROW_BG)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT   = Alignment(horizontal="right",  vertical="center")


class ExportResultAnalyticsView(LoginRequiredMixin, View):
    """
    Export result analytics data to Excel.
    Modes: students | top_performers | bottom_performers |
           subject_performance | cross_matrix | grade_distribution
    """

    SCHOOL_NAME    =  getattr(settings, 'SCHOOL_NAME', 'SCHOOL NAME NOT SET'),
    SCHOOL_ADDRESS = getattr(settings, 'SCHOOL_ADDRESS', ''),
    SCHOOL_PHONE   = getattr(settings, 'SCHOOL_PHONE', ''),
    SCHOOL_EMAIL   = getattr(settings, 'SCHOOL_EMAIL', ''),

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------

    def _sanitize_worksheet_name(self, name):
        name = re.sub(r'[\\/*?:\[\]]', '', str(name))
        name = re.sub(r'\s+', '_', name)
        name = re.sub(r'[^\w\-_.]', '', name)
        return (name[:31] if len(name) > 31 else name) or "Sheet"

    # ------------------------------------------------------------------
    # HTTP entry point
    # ------------------------------------------------------------------

    def get(self, request, pk):
        session = get_object_or_404(ExamSession, pk=pk)
        mode = request.GET.get('mode', 'students')
        top_n = int(request.GET.get('top_n', 10))
        bottom_n = int(request.GET.get('bottom_n', 10))
        div_filter = request.GET.get('division', '')
        grade_filter = request.GET.get('grade', '')
        gen_filter = request.GET.get('gender', '')
        combination_filter = request.GET.get('combination', '')
        search = request.GET.get('search', '')

        educational_level = session.class_level.educational_level
        level_type = educational_level.level_type

        data = self._load_data(session, educational_level, level_type,
                               div_filter, grade_filter, gen_filter, 
                               combination_filter, search, top_n, bottom_n)

        # Dispatch based on mode and level type
        if mode == 'grade_distribution':
            wb = self._export_grade_distribution(data, session)
        elif mode == 'students':
            wb = self._export_students(data, session)
        elif mode == 'top_performers':
            wb = self._export_top_performers(data, session, top_n)
        elif mode == 'bottom_performers':
            wb = self._export_bottom_performers(data, session, bottom_n)
        elif mode == 'subject_performance':
            wb = self._export_subject_performance(data, session)
        elif mode == 'cross_matrix':
            wb = self._export_cross_matrix(data, session)
        else:
            wb = self._export_students(data, session)

        mode_display = {
            'students': 'Complete_Results',
            'top_performers': f'Top_{top_n}_Performers',
            'bottom_performers': f'Bottom_{bottom_n}_Performers',
            'subject_performance': 'Subject_Performance',
            'cross_matrix': 'Cross_Matrix',
            'grade_distribution': 'Grade_Distribution'
        }.get(mode, 'Results')

        date_str = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{mode_display}_{date_str}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # ------------------------------------------------------------------
    # Header / footer builders
    # ------------------------------------------------------------------

    def _add_school_header(self, ws, title, session, num_cols, start_row=1):
        end_col = max(num_cols, 1)

        def _merge_write(row, value, font, fill=None, height=None):
            cell = ws.cell(row=row, column=1, value=value)
            cell.font = font
            cell.alignment = CENTER
            if fill:
                cell.fill = fill
            if end_col > 1:
                ws.merge_cells(
                    start_row=row, start_column=1,
                    end_row=row, end_column=end_col
                )
            if height:
                ws.row_dimensions[row].height = height
            return cell

        r = start_row

        _merge_write(r, self.SCHOOL_NAME, SCHOOL_FONT,
                     PatternFill("solid", fgColor="1A237E"), height=28)
        ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, color=WHITE, size=16)
        r += 1

        _merge_write(r, self.SCHOOL_ADDRESS, INFO_FONT, height=16)
        r += 1

        _merge_write(r, f"Tel: {self.SCHOOL_PHONE}   |   Email: {self.SCHOOL_EMAIL}",
                     INFO_FONT, height=14)
        r += 1

        accent_cell = ws.cell(row=r, column=1, value=" ")
        accent_cell.fill = PatternFill("solid", fgColor=ACCENT_GOLD)
        if end_col > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
        ws.row_dimensions[r].height = 4
        r += 1

        _merge_write(r, title, SUB_HEADER_FONT,
                     PatternFill("solid", fgColor="E8F0FE"), height=22)
        r += 1

        class_name = session.class_level.name
        year_name = session.academic_year.name
        meta = f"Exam: {session.name}   |   Class: {class_name}   |   Year: {year_name}"
        if session.stream_class:
            meta += f"   |   Stream: {session.stream_class.name}"
        meta += f"   |   Date: {session.exam_date.strftime('%d %B %Y')}"
        _merge_write(r, meta, SESSION_FONT, height=14)
        r += 1

        ws.row_dimensions[r].height = 6
        r += 1

        return r

    def _add_footer(self, ws, row):
        cell = ws.cell(row=row + 1, column=1,
                       value=f"Generated: {timezone.now().strftime('%d %B %Y at %H:%M:%S')}   |   {self.SCHOOL_NAME}")
        cell.font = FOOTER_FONT
        cell.alignment = LEFT
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"1:{row - 1}"

    def _apply_header_row(self, ws, headers, col_widths, title, session, start_row=1):
        data_row = self._add_school_header(ws, title, session, len(headers), start_row)

        ws.row_dimensions[data_row].height = 30
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = _thin_border(WHITE)

        for col_idx, width in (col_widths or {}).items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = ws.cell(row=data_row + 1, column=1)
        return data_row + 1

    @staticmethod
    def _style_data_row(ws, row_num, num_cols, is_alt=False):
        fill = ALT_FILL if is_alt else None
        bdr = _thin_border()
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = BODY_FONT
            cell.border = bdr
            if fill:
                cell.fill = fill
            if col == 1:
                cell.alignment = CENTER
            elif col == 2:
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER

    @staticmethod
    def _style_total_row(ws, row_num, num_cols):
        bdr = _thin_border(BRAND_DARK)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = BODY_BOLD_FONT
            cell.fill = TOTAL_FILL
            cell.border = bdr
            cell.alignment = CENTER

    # ------------------------------------------------------------------
    # Data loader
    # ------------------------------------------------------------------

    def _load_data(self, session, educational_level, level_type,division_filter, grade_filter, gender_filter,  combination_filter, search_query, top_n, bottom_n):

        grading_scale = list(
            GradingScale.objects.filter(education_level=educational_level).order_by('-min_mark')
        )
        division_scale = {
            ds.division: ds for ds in
            DivisionScale.objects.filter(education_level=educational_level)
        }
        all_divisions = self._all_division_labels(division_scale)

        # Get all combinations for A-Level
        combinations = []
        if level_type == 'A_LEVEL':
            combinations = list(Combination.objects.filter(
                educational_level=educational_level
            ).order_by('code'))

        all_level_subjects = list(
            Subject.objects.filter(educational_level=educational_level).order_by('name')
        )
        all_level_subject_objects = [
            {'id': s.id, 'name': s.name,
            'short_name': s.short_name or s.name[:15], 'code': s.code}
            for s in all_level_subjects
        ]

        enrolled_qs = StudentEnrollment.objects.filter(
            academic_year=session.academic_year,
            class_level=session.class_level,
            status='active'
        ).select_related('student', 'student__user')
        
        if session.stream_class:
            enrolled_qs = enrolled_qs.filter(
                stream_assignment__stream_class=session.stream_class
            )
        enrolled = enrolled_qs.order_by('student__first_name', 'student__last_name')

        positions = {
            p.student_id: {'class_position': p.class_position,
                        'stream_position': p.stream_position}
            for p in StudentExamPosition.objects.filter(exam_session=session)
        }
        
        metrics_data = {
            m.student_id: {
                'total_marks': float(m.total_marks),
                'average_marks': float(m.average_marks),
                'total_points': float(m.total_points) if m.total_points else 0,
                'division': m.division,
            }
            for m in StudentExamMetrics.objects.filter(exam_session=session)
        }

        session_papers = SubjectExamPaper.objects.filter(
            exam_session=session).select_related('subject')
        papers_by_subject = {}
        for paper in session_papers:
            sid = paper.subject_id
            papers_by_subject.setdefault(sid, {'subject': paper.subject, 'papers': []})
            papers_by_subject[sid]['papers'].append({
                'paper_number': paper.paper_number,
                'max_marks': float(paper.max_marks),
            })

        paper_scores_qs = StudentPaperScore.objects.filter(
            exam_paper__exam_session=session
        ).select_related('exam_paper', 'exam_paper__subject')
        student_paper_scores = {}
        for score in paper_scores_qs:
            student_paper_scores.setdefault(score.student_id, {})
            student_paper_scores[score.student_id].setdefault(score.exam_paper.subject_id, [])
            student_paper_scores[score.student_id][score.exam_paper.subject_id].append({
                'paper_number': score.exam_paper.paper_number,
                'marks': float(score.marks),
                'max_marks': float(score.exam_paper.max_marks),
            })

        subject_results_qs = StudentSubjectResult.objects.filter(
            exam_session=session).select_related('student', 'subject')
        subject_results_map = {
            (r.student_id, r.subject_id): r for r in subject_results_qs
        }

        students_data = []
        for enrollment in enrolled:
            student = enrollment.student
            student_id = student.pk

            if gender_filter and student.gender != gender_filter:
                continue
            if search_query:
                if not (search_query.lower() in student.full_name.lower() or
                        search_query.lower() in (student.registration_number or '').lower()):
                    continue

            pos_data = positions.get(student_id, {})
            metrics = metrics_data.get(student_id, {})
            papers_scores = student_paper_scores.get(student_id, {})

            has_metrics = student_id in metrics_data
            has_paper_scores = bool(papers_scores)
            has_subject_results = any(
                (student_id, s['id']) in subject_results_map
                for s in all_level_subject_objects
            )
            has_results = has_metrics or has_paper_scores or has_subject_results

            # Get combination for A-Level
            combination_code = None
            combination_obj = None
            if level_type == 'A_LEVEL':
                combination_obj = self._get_student_combination(student_id, session.academic_year)
                combination_code = combination_obj.code if combination_obj else None
                if combination_filter:
                    if combination_obj:
                        if str(combination_obj.id) != combination_filter:
                            continue
                    else:
                        continue

            subject_performance = []
            for subject in all_level_subject_objects:
                subj_id = subject['id']
                subj_papers = papers_scores.get(subj_id, [])
                db_result = subject_results_map.get((student_id, subj_id))
                has_ep = subj_id in papers_by_subject
                has_ps = bool(subj_papers)

                if has_ps:
                    avg = self._compute_subject_average(subj_papers)
                    grade = self._grade_from_average(avg, grading_scale)
                elif db_result:
                    avg = float(db_result.total_marks)
                    grade = db_result.grade
                else:
                    avg = 0.0
                    grade = None

                if not has_results or not has_ep or (not has_ps and not db_result):
                    display_grade = ''
                    display_avg = ''
                else:
                    display_grade = grade or ''
                    display_avg = round(avg, 2) if avg > 0 else ''

                subject_performance.append({
                    'subject_name': subject['name'],
                    'grade': display_grade,
                    'average': display_avg,
                })

            # Overall figures - different for Primary/Nursery vs O/A-Level
            if has_metrics:
                total_marks = metrics.get('total_marks', 0)
                average_marks = metrics.get('average_marks', 0)
                division = metrics.get('division', '')
                points = metrics.get('total_points', 0)
            elif has_paper_scores or has_subject_results:
                scored = [s['average'] for s in subject_performance
                        if s['average'] not in ('', 'Absent', 'N/A')]
                if scored:
                    total_marks = sum(scored)
                    average_marks = round(total_marks / len(scored), 2)
                    division = ''
                    points = ''
                else:
                    total_marks = average_marks = 0
                    division = points = ''
            else:
                total_marks = average_marks = 0
                division = points = ''

            # Derive overall grade for Primary/Nursery
            overall_grade = None
            if level_type in ['PRIMARY', 'NURSERY'] and has_results:
                subject_grades = [s['grade'] for s in subject_performance
                                if s['grade'] not in ('', 'Absent', 'N/A')]
                if subject_grades:
                    grade_map = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'E': 1, 'F': 0}
                    avg_grade = sum(grade_map.get(g, 0) for g in subject_grades) / len(subject_grades)
                    for g, val in sorted(grade_map.items(), key=lambda x: x[1], reverse=True):
                        if avg_grade >= val:
                            overall_grade = g
                            break

            # Apply filters
            if grade_filter and level_type in ['PRIMARY', 'NURSERY']:
                if overall_grade != grade_filter:
                    continue
            if division_filter and level_type in ['O_LEVEL', 'A_LEVEL']:
                if division != division_filter:
                    continue

            students_data.append({
                'student_id': student_id,
                'full_name': student.full_name,
                'registration_number': student.registration_number or 'N/A',
                'gender': student.get_gender_display() or 'Not specified',
                'has_results': has_results,
                'class_position': pos_data.get('class_position'),
                'total_marks': total_marks if has_results else '',
                'average_marks': average_marks if has_results else '',
                'division': division if division else '',
                'points': points if points else '',
                'grade': overall_grade if level_type in ['PRIMARY', 'NURSERY'] else (division if division else ''),
                'combination': combination_code,
                'subjects': subject_performance,
            })

        # Sort by position
        with_pos = sorted([s for s in students_data if s['class_position'] is not None],
                        key=lambda x: x['class_position'])
        without_pos = [s for s in students_data if s['class_position'] is None]
        students_sorted = with_pos + without_pos

        # Calculate subject performance
        subject_performance_stats = []
        for subject in all_level_subject_objects:
            avgs = []
            absent_count = no_paper_count = 0
            for student in students_sorted:
                if not student['has_results']:
                    continue
                for subj in student['subjects']:
                    if subj['subject_name'] == subject['name']:
                        if subj['average'] == '':
                            absent_count += 1
                        elif subj['average'] == 'N/A':
                            no_paper_count += 1
                        else:
                            try:
                                avgs.append(float(subj['average']))
                            except (ValueError, TypeError):
                                pass
            avg = sum(avgs) / len(avgs) if avgs else 0
            subject_performance_stats.append({
                'subject_name': subject['name'],
                'average_marks': round(avg, 2),
                'total_students': len(avgs),
                'absent_count': absent_count,
                'no_paper_count': no_paper_count,
            })

        # Get grades list from grading scale for Primary/Nursery
        if level_type in ['PRIMARY', 'NURSERY']:
            grades = [gs.grade for gs in grading_scale] if grading_scale else ['A', 'B', 'C', 'D', 'E', 'F']
        else:
            grades = []

        # Calculate cross matrix with proper initialization
        if level_type in ['PRIMARY', 'NURSERY']:
            # Initialize cross matrix with all grades set to 0
            cross_matrix = {
                'male': {grade: 0 for grade in grades},
                'female': {grade: 0 for grade in grades},
                'total': {grade: 0 for grade in grades}
            }
            
            # Populate cross matrix with actual data
            for student in with_pos:
                if not student['has_results']:
                    continue
                gender = 'male' if student['gender'] == 'Male' else 'female'
                grade_val = student.get('grade', '')
                if grade_val and grade_val in cross_matrix[gender]:
                    cross_matrix[gender][grade_val] += 1
                    cross_matrix['total'][grade_val] += 1
        else:
            # For O/A-Level, use division cross matrix
            cross_matrix = {
                'male': {d: 0 for d in all_divisions},
                'female': {d: 0 for d in all_divisions},
                'total': {d: 0 for d in all_divisions}
            }
            
            for student in with_pos:
                if not student['has_results']:
                    continue
                gender = 'male' if student['gender'] == 'Male' else 'female'
                division_val = student.get('division', '')
                if division_val and division_val in cross_matrix[gender]:
                    cross_matrix[gender][division_val] += 1
                    cross_matrix['total'][division_val] += 1

        # Calculate grade distribution for stats
        grade_dist = {}
        if level_type in ['PRIMARY', 'NURSERY']:
            for student in with_pos:
                g = student.get('grade', '')
                if g and g != 'N/A':
                    grade_dist[g] = grade_dist.get(g, 0) + 1

        # Stats dictionary with grade distribution
        stats = {
            'total_students': len(students_data),
            'students_with_results': len(with_pos),
            'students_absent': len(without_pos),
            'average_total_marks': 0,
            'highest_total_marks': 0,
            'lowest_total_marks': 0,
            'average_points': 0,
            'division_distribution': {},
            'grade_distribution': grade_dist,
            'gender_distribution': {},
            'subject_performance': subject_performance_stats,
        }

        # Calculate averages if there are students with results
        if with_pos:
            total_marks_list = [float(s['total_marks']) for s in with_pos if s['total_marks']]
            if total_marks_list:
                stats['average_total_marks'] = round(sum(total_marks_list) / len(total_marks_list), 2)
                stats['highest_total_marks'] = round(max(total_marks_list), 2)
                stats['lowest_total_marks'] = round(min(total_marks_list), 2)
            
            # Points for O/A-Level
            if level_type in ['O_LEVEL', 'A_LEVEL']:
                points_list = [float(s['points']) for s in with_pos if s['points']]
                if points_list:
                    stats['average_points'] = round(sum(points_list) / len(points_list), 2)
            
            # Division distribution for O/A-Level
            if level_type in ['O_LEVEL', 'A_LEVEL']:
                for d in all_divisions:
                    stats['division_distribution'][d] = 0
                for student in with_pos:
                    div = student.get('division', '')
                    if div and div in stats['division_distribution']:
                        stats['division_distribution'][div] += 1

        return {
            'students': students_sorted,
            'students_with_results': with_pos,
            'subject_performance': subject_performance_stats,
            'cross_matrix': cross_matrix,
            'divisions': all_divisions,
            'grades': grades,
            'subjects': all_level_subject_objects,
            'level_type': level_type,
            'session': session,
            'stats': stats,
        }
    # ------------------------------------------------------------------
    # Computation helpers
    # ------------------------------------------------------------------

    def _compute_subject_average(self, paper_scores):
        if not paper_scores:
            return 0.0
        total = sum(p['marks'] for p in paper_scores)
        max_tot = sum(p['max_marks'] for p in paper_scores)
        return (total / max_tot * 100) if max_tot else 0.0

    def _grade_from_average(self, average, grading_scale):
        for gs in grading_scale:
            if float(gs.min_mark) <= average <= float(gs.max_mark):
                return gs.grade
        return ''

    def _all_division_labels(self, division_scale):
        order = ['I', 'II', 'III', 'IV', '0']
        defined = set(division_scale.keys())
        ordered = [d for d in order if d in defined]
        ordered += sorted(defined - set(order))
        return ordered

    def _get_student_combination(self, student_id, academic_year):
        try:
            assignment = StudentCombinationAssignment.objects.filter(
                student_id=student_id,
                enrollment__academic_year=academic_year,
                is_active=True
            ).select_related('combination').first()
            return assignment.combination if assignment else None
        except Exception:
            return None

    def _calculate_division_cross_matrix(self, students_data, level_type, all_divisions):
        """Calculate gender × division cross matrix for O/A-Level."""
        matrix = {'male': {}, 'female': {}, 'total': {}}
        for d in all_divisions:
            matrix['male'][d] = matrix['female'][d] = matrix['total'][d] = 0
        for student in students_data:
            if not student['has_results']:
                continue
            gender = 'male' if student['gender'] == 'Male' else 'female'
            value = student['division']
            if not value or value == 'N/A':
                continue
            matrix[gender][value] = matrix[gender].get(value, 0) + 1
            matrix['total'][value] = matrix['total'].get(value, 0) + 1
        return matrix

    def _calculate_grade_cross_matrix(self, students_data, level_type, grades):
        """Calculate gender × grade cross matrix for Primary/Nursery."""
        matrix = {'male': {}, 'female': {}, 'total': {}}
        for g in grades:
            matrix['male'][g] = matrix['female'][g] = matrix['total'][g] = 0
        for student in students_data:
            if not student['has_results']:
                continue
            gender = 'male' if student['gender'] == 'Male' else 'female'
            value = student.get('grade', '')
            if not value or value == 'N/A':
                continue
            matrix[gender][value] = matrix[gender].get(value, 0) + 1
            matrix['total'][value] = matrix['total'].get(value, 0) + 1
        return matrix

    # ------------------------------------------------------------------
    # Student row writer (handles both O/A-Level and Primary/Nursery)
    # ------------------------------------------------------------------

    def _write_student_rows(self, ws, students, subjects, data_start_row, level_type):
        is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
        is_primary = level_type in ['PRIMARY', 'NURSERY']
        
        if is_secondary:
            fixed_cols = 5 if level_type == 'A_LEVEL' else 4
            num_cols = fixed_cols + len(subjects) + 5
        else:
            fixed_cols = 4
            num_cols = fixed_cols + len(subjects) + 3  # No division/points for primary

        for idx, student in enumerate(students, 1):
            row_num = data_start_row + idx - 1
            is_alt = (idx % 2 == 0)
            self._style_data_row(ws, row_num, num_cols, is_alt)

            col = 1
            ws.cell(row=row_num, column=col, value=idx)
            col += 1

            name_cell = ws.cell(row=row_num, column=col, value=student['full_name'])
            name_cell.alignment = LEFT
            name_cell.font = BODY_FONT
            col += 1

            ws.cell(row=row_num, column=col, value=student['registration_number'])
            col += 1

            ws.cell(row=row_num, column=col, value=student['gender'])
            col += 1

            # A-Level combination column
            if level_type == 'A_LEVEL':
                combination = student.get('combination', '')
                comb_cell = ws.cell(row=row_num, column=col, value=combination or '—')
                if combination:
                    comb_cell.font = Font(name="Arial", bold=True, color="6F42C1", size=10)
                col += 1

            # Subject scores
            subj_map = {s['subject_name']: s for s in student['subjects']}
            for subject in subjects:
                subj = subj_map.get(subject['name'], {})
                grade = subj.get('grade', '')
                avg = subj.get('average', '')
                cell = ws.cell(row=row_num, column=col,
                               value=f"{grade}-{avg}" if grade else '')
                if grade in ('A', 'A+'):
                    cell.font = Font(name="Arial", bold=True, color="0F9D58", size=10)
                elif grade in ('F', 'E'):
                    cell.font = Font(name="Arial", color="D93025", size=10)
                col += 1

            # Metrics
            ws.cell(row=row_num, column=col, value=student['total_marks'])
            ws.cell(row=row_num, column=col + 1, value=student['average_marks'])
            
            if is_secondary:
                ws.cell(row=row_num, column=col + 2, value=student['division'])
                ws.cell(row=row_num, column=col + 3, value=student['points'])
                pos_col = col + 4
            else:
                # For Primary/Nursery, show grade instead of division/points
                grade_val = student.get('grade', '')
                grade_cell = ws.cell(row=row_num, column=col + 2, value=grade_val)
                if grade_val in ('A', 'B', 'C', 'D', 'E', 'F'):
                    grade_colors = {'A': "0F9D58", 'B': "0F9D58", 'C': "FFB300", 
                                    'D': "FF6B00", 'E': "D93025", 'F': "D93025"}
                    grade_cell.font = Font(name="Arial", bold=True, color=grade_colors.get(grade_val, "5F6368"), size=10)
                pos_col = col + 3

            pos_val = f"#{student['class_position']}" if student['class_position'] else ''
            pos_cell = ws.cell(row=row_num, column=pos_col, value=pos_val)
            if student['class_position'] and student['class_position'] <= 3:
                pos_cell.font = Font(name="Arial", bold=True, color=ACCENT_GOLD, size=10)

        return data_start_row + len(students)

    # ------------------------------------------------------------------
    # Build headers (level-aware)
    # ------------------------------------------------------------------

    def _build_headers_and_widths(self, subjects, level_type):
        is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
        is_primary = level_type in ['PRIMARY', 'NURSERY']
        
        headers = ['#', 'Student Name', 'Reg. Number', 'Gender']
        if level_type == 'A_LEVEL':
            headers.append('Combination')
        
        for s in subjects:
            headers.append(s['short_name'])
        
        if is_secondary:
            headers += ['Total', 'Average', 'Division', 'Points', 'Position']
        else:
            headers += ['Total', 'Average', 'Grade', 'Position']

        # Set column widths
        widths = {1: 5, 2: 28, 3: 18, 4: 10}
        start_col = 6 if level_type == 'A_LEVEL' else 5
        
        for i in range(len(subjects)):
            widths[start_col + i] = 9
        
        base = start_col + len(subjects)
        if is_secondary:
            widths.update({base: 9, base+1: 9, base+2: 10, base+3: 8, base+4: 10})
        else:
            widths.update({base: 9, base+1: 9, base+2: 8, base+3: 10})
        
        return headers, widths

    # ------------------------------------------------------------------
    # Export methods
    # ------------------------------------------------------------------

    def _export_grade_distribution(self, data, session):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Grade Distribution")
        
        headers = ['Grade', 'Number of Students', 'Percentage']
        widths = {1: 12, 2: 18, 3: 15}
        
        data_row = self._apply_header_row(ws, headers, widths,
                                          "GRADE DISTRIBUTION", session)
        
        grade_dist = data['stats'].get('grade_distribution', {}) if 'stats' in data else {}
        total_students = len(data['students_with_results'])
        
        grade_order = ['A', 'B', 'C', 'D', 'E', 'F']
        
        for idx, grade in enumerate(grade_order, 1):
            count = grade_dist.get(grade, 0)
            percentage = (count / total_students * 100) if total_students > 0 else 0
            row_num = data_row + idx - 1
            is_alt = (idx % 2 == 0)
            self._style_data_row(ws, row_num, 3, is_alt)
            
            ws.cell(row=row_num, column=1, value=grade)
            ws.cell(row=row_num, column=2, value=count)
            percent_cell = ws.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
            if percentage >= 50:
                percent_cell.font = Font(name="Arial", color="0F9D58", bold=True)
            elif percentage <= 10:
                percent_cell.font = Font(name="Arial", color="D93025")
        
        self._add_footer(ws, data_row + len(grade_order))
        return wb

    def _export_students(self, data, session):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Student Results")
        
        headers, widths = self._build_headers_and_widths(data['subjects'], data['level_type'])
        data_row = self._apply_header_row(ws, headers, widths,
                                          "STUDENT RESULTS REPORT", session)
        last_row = self._write_student_rows(ws, data['students'],
                                            data['subjects'], data_row, data['level_type'])
        self._add_footer(ws, last_row)
        return wb

    def _export_top_performers(self, data, session, top_n):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Top Performers")
        
        top = data['students_with_results'][:top_n]
        if not top:
            ws.cell(row=1, column=1, value="No top performers found.")
            return wb
        
        headers, widths = self._build_headers_and_widths(data['subjects'], data['level_type'])
        data_row = self._apply_header_row(ws, headers, widths,
                                          f"TOP {len(top)} PERFORMERS", session)
        last_row = self._write_student_rows(ws, top, data['subjects'], data_row, data['level_type'])
        self._add_footer(ws, last_row)
        return wb

    def _export_bottom_performers(self, data, session, bottom_n):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Bottom Performers")
        
        students = data['students_with_results']
        bottom = students[-bottom_n:] if len(students) >= bottom_n else students
        if not bottom:
            ws.cell(row=1, column=1, value="No bottom performers found.")
            return wb
        
        headers, widths = self._build_headers_and_widths(data['subjects'], data['level_type'])
        data_row = self._apply_header_row(ws, headers, widths,
                                          f"BOTTOM {len(bottom)} PERFORMERS", session)
        last_row = self._write_student_rows(ws, bottom, data['subjects'], data_row, data['level_type'])
        self._add_footer(ws, last_row)
        return wb

    def _export_subject_performance(self, data, session):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Subject Performance")
        
        headers = ['Subject', 'Average Marks (%)', 'Students Scored', 'Absent', 'No Paper']
        widths = {1: 32, 2: 15, 3: 16, 4: 10, 5: 10}
        data_row = self._apply_header_row(ws, headers, widths,
                                          "SUBJECT PERFORMANCE COMPARISON", session)
        
        for idx, subject in enumerate(data['subject_performance'], 1):
            row_num = data_row + idx - 1
            is_alt = (idx % 2 == 0)
            self._style_data_row(ws, row_num, 5, is_alt)
            
            avg = subject['average_marks']
            ws.cell(row=row_num, column=1, value=subject['subject_name']).alignment = LEFT
            avg_cell = ws.cell(row=row_num, column=2,
                               value=round(avg, 2) if avg > 0 else '')
            avg_cell.number_format = '0.00"%"'
            if avg >= 60:
                avg_cell.font = Font(name="Arial", color="0F9D58", bold=True, size=10)
            elif avg < 40:
                avg_cell.font = Font(name="Arial", color="D93025", size=10)
            ws.cell(row=row_num, column=3, value=subject['total_students'])
            ws.cell(row=row_num, column=4, value=subject['absent_count'])
            ws.cell(row=row_num, column=5, value=subject['no_paper_count'])
        
        self._add_footer(ws, data_row + len(data['subject_performance']))
        return wb

    def _export_cross_matrix(self, data, session):
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_worksheet_name("Cross Matrix")
        
        if data['level_type'] in ['PRIMARY', 'NURSERY']:
            headers = ['Gender / Grade'] + [f'Grade {g}' for g in data['grades']] + ['Total']
            widths = {1: 20}
            for i in range(len(data['grades'])):
                widths[i+2] = 12
            widths[len(data['grades'])+2] = 12
        else:
            headers = ['Gender / Division'] + [f'Division {d}' for d in data['divisions']] + ['Total']
            widths = {1: 20}
            for i in range(len(data['divisions'])):
                widths[i+2] = 12
            widths[len(data['divisions'])+2] = 12
        
        data_row = self._apply_header_row(ws, headers, widths,
                                        "GENDER × CROSS MATRIX", session)
        
        row_num = data_row
        cross_matrix = data['cross_matrix']
        
        # Male row
        self._style_data_row(ws, row_num, len(headers))
        male_row = ['Male']
        male_total = 0
        if data['level_type'] in ['PRIMARY', 'NURSERY']:
            for grade in data['grades']:
                cnt = cross_matrix['male'].get(grade, 0)
                male_row.append(cnt)
                male_total += cnt
        else:
            for division in data['divisions']:
                cnt = cross_matrix['male'].get(division, 0)
                male_row.append(cnt)
                male_total += cnt
        male_row.append(male_total)
        
        for col, val in enumerate(male_row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = CENTER if col > 1 else LEFT
        row_num += 1
        
        # Female row
        self._style_data_row(ws, row_num, len(headers))
        female_row = ['Female']
        female_total = 0
        if data['level_type'] in ['PRIMARY', 'NURSERY']:
            for grade in data['grades']:
                cnt = cross_matrix['female'].get(grade, 0)
                female_row.append(cnt)
                female_total += cnt
        else:
            for division in data['divisions']:
                cnt = cross_matrix['female'].get(division, 0)
                female_row.append(cnt)
                female_total += cnt
        female_row.append(female_total)
        
        for col, val in enumerate(female_row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = CENTER if col > 1 else LEFT
        row_num += 1
        
        # Total row
        self._style_total_row(ws, row_num, len(headers))
        total_row = ['Total']
        grand_total = 0
        if data['level_type'] in ['PRIMARY', 'NURSERY']:
            for grade in data['grades']:
                cnt = cross_matrix['total'].get(grade, 0)
                total_row.append(cnt)
                grand_total += cnt
        else:
            for division in data['divisions']:
                cnt = cross_matrix['total'].get(division, 0)
                total_row.append(cnt)
                grand_total += cnt
        total_row.append(grand_total)
        
        for col, val in enumerate(total_row, 1):
            ws.cell(row=row_num, column=col, value=val).alignment = CENTER
        
        self._add_footer(ws, row_num + 1)
        return wb