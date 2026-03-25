"""
portal_management/views/students.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
All student-related views for the Management portal.

32 views organised into 7 sections:
  ── STUDENT CRUD ──────────────────── StudentListView, StudentCreateView,
                                        StudentDetailView, StudentUpdateView
  ── ENROLLMENT ────────────────────── StudentEnrollView, StudentStreamAssignView
  ── DRAFTS ────────────────────────── StudentDraftListView, StudentDraftCreateView,
                                        StudentDraftEditView, StudentDraftPublishView,
                                        StudentDraftDeleteView
  ── LIFECYCLE ─────────────────────── StudentSuspendView, StudentLiftSuspensionView,
                                        StudentTransferView, StudentWithdrawView
  ── ACCOUNT ───────────────────────── StudentResetPasswordView
  ── PARENT MANAGEMENT ─────────────── StudentParentManagementView,
                                        StudentAddParentView, StudentParentUpdateView,
                                        StudentParentRemoveView, StudentParentSetPrimaryView,
                                        StudentParentBulkAddView, ParentEditView,
                                        ParentDeleteView, GetParentDetailsView,
                                        ParentListView, ParentCreateView, ParentUpdateView
  ── AJAX HELPERS ──────────────────── SearchParentsView, GetStreamsView,
                                        GetCombinationsView
"""

from datetime import date, datetime
import json
import logging

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Q, Avg, Max, Min, Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import DetailView, ListView, TemplateView, View
from weasyprint import HTML

from core.mixins import ManagementRequiredMixin
from core.models import (
    AcademicYear, AuditLog, ClassLevel, Combination, DivisionScale, ExamSession, GradingScale,
    Parent, SchoolProfile, StreamClass, Student, StudentCombinationAssignment, StudentEnrollment, StudentExamMetrics, StudentExamPosition, StudentPaperScore,
    StudentParent, StudentStreamAssignment, StudentSubjectAssignment, StudentSubjectResult, StudentSuspension, StudentTransferOut, SubjectExamPaper, Term,
)
from portal_management.forms.parent_form import ParentForm
from portal_management.forms.student_form import StudentEnrollmentForm, StudentForm
from portal_management.forms.student_parent_form import StudentParentForm
# ===== OPENPYXL =====
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== DJANGO =====
from django.template.loader import render_to_string

# ===== WEASYPRINT =====
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration


logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# STUDENT CRUD
# ════════════════════════════════════════════════════════════════════════════

class StudentListView(ManagementRequiredMixin, TemplateView):
    """Display all students in a DataTable with status summary pills."""
    template_name = 'portal_management/students/list.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        status_filter = self.request.GET.get('status')
        qs = Student.objects.select_related('user').order_by('registration_number')
        if status_filter:
            qs = qs.filter(status=status_filter)
        ctx['students'] = qs
        ctx['status_counts'] = {
            s: Student.objects.filter(status=s).count()
            for s in ['active', 'suspended', 'withdrawn', 'completed', 'transferred']
        }
        ctx['selected_status'] = status_filter
        return ctx


class StudentCreateView(ManagementRequiredMixin, View):
    """
    Enroll a new student.
    Supports multiple submit actions via POST param 'action':
      save_view    -> redirect to student detail  (default)
      save_list    -> redirect to student list
      save_new     -> redirect back to create form
      save_enroll  -> redirect to enrollment page
      save_parent  -> redirect to parent management page
      save_draft   -> save as draft (session-based)
    Supports AJAX: returns JSON when X-Requested-With header is present.
    """
    template_name = 'portal_management/students/form.html'

    def get(self, request):
        form = StudentForm()
        form.fields['admission_date'].initial = timezone.now().date()

        # Load session draft if requested
        draft_id = request.GET.get('draft')
        if draft_id:
            draft_data = request.session.get(f'draft_{draft_id}', {})
            if draft_data:
                form = StudentForm(initial=draft_data)
                messages.info(request, 'Loading draft data.')

        return render(request, self.template_name, {
            'form': form,
            'title': 'Enroll New Student',
            'action': 'Create',
        })

    def post(self, request):
        form = StudentForm(request.POST, request.FILES)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        next_action = request.POST.get('action', 'save_view')

        # ── Draft save (no DB record created) ────────────────────────────
        if next_action == 'save_draft' and form.is_valid():
            draft_data = {
                k: v for k, v in form.cleaned_data.items()
                if k != 'profile_picture'
            }
            draft_id = int(timezone.now().timestamp())
            request.session[f'draft_{draft_id}'] = {
                'data': draft_data,
                'timestamp': timezone.now().isoformat(),
            }
            # Trim old drafts (keep last 10)
            keys = [k for k in request.session.keys() if k.startswith('draft_')]
            for old in sorted(keys)[:-10]:
                del request.session[old]
            messages.success(request, 'Draft saved. You can continue later.')
            if is_ajax:
                return JsonResponse({'success': True, 'draft_id': draft_id,
                                     'message': 'Draft saved successfully.'})
            return redirect('management:student_draft_list')

        # ── Regular save ──────────────────────────────────────────────────
        if form.is_valid():
            try:
                with transaction.atomic():
                    student = form.save()
                    if not Student.objects.filter(pk=student.pk).exists():
                        raise ValidationError('Student was not properly saved.')

                    # Clear session drafts
                    for k in [k for k in request.session.keys()
                               if k.startswith('draft_')]:
                        del request.session[k]

                    messages.success(
                        request,
                        f'Student {student.full_name} enrolled successfully. '
                        f'Registration Number: {student.registration_number}'
                    )

                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'student_id': student.pk,
                            'registration_number': student.registration_number,
                            'redirect_url': reverse('management:student_detail',
                                                    kwargs={'pk': student.pk}),
                        })

                    if next_action == 'save_list':
                        return redirect('management:student_list')
                    elif next_action == 'save_new':
                        messages.info(request, 'Add another student.')
                        return redirect('management:student_create')
                    elif next_action == 'save_enroll':
                        messages.info(request, f'Complete enrollment for {student.full_name}.')
                        return redirect('management:student_enroll', pk=student.pk)
                    elif next_action == 'save_parent':
                        messages.info(request, f'Add parent/guardian for {student.full_name}.')
                        return redirect('management:student_parent_management', pk=student.pk)
                    else:
                        return redirect('management:student_detail', pk=student.pk)

            except ValidationError as e:
                error_msg = ', '.join(e.messages) if hasattr(e, 'messages') else str(e)
                messages.error(request, f'Validation error: {error_msg}')
                logger.error('StudentCreate validation error: %s', e, exc_info=True)
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg})

            except Exception as e:
                messages.error(request, f'Error saving student: {e}')
                logger.error('StudentCreate error: %s', e, exc_info=True)
                if is_ajax:
                    return JsonResponse({'success': False, 'error': str(e)})
        else:
            error_count = len(form.errors)
            messages.error(
                request,
                f'Please correct the {error_count} '
                f'error{"s" if error_count > 1 else ""} below.'
            )
            logger.error('StudentCreate form errors: %s', form.errors)
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'error_messages': [
                        f'{form.fields[f].label if f in form.fields else f}: '
                        f'{", ".join(errs)}'
                        for f, errs in form.errors.items()
                    ],
                })

        return render(request, self.template_name, {
            'form': form,
            'title': 'Enroll New Student',
            'action': 'Create',
        })


class StudentDetailView(ManagementRequiredMixin, DetailView):
    """Full student profile page."""
    model = Student
    template_name = 'portal_management/students/detail.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        student = self.object
        
        # Get enrollments with proper select_related (no combination field)
        ctx['enrollments'] = student.enrollments.select_related(
            'academic_year', 
            'class_level'
        ).order_by('-academic_year__start_date')
        
        # For each enrollment, get the combination if applicable (A-Level)
        for enrollment in ctx['enrollments']:
            # Get the current combination for this enrollment
            if enrollment.class_level.educational_level.level_type == 'A_LEVEL':
                combination_assignment = enrollment.combination_assignments.filter(is_active=True).first()
                enrollment.combination = combination_assignment.combination if combination_assignment else None
            else:
                enrollment.combination = None
        
        ctx['parents'] = student.parents.all()
        
        # Get active enrollment
        active_enrollment = student.enrollments.filter(status='active').first()
        ctx['active_enrollment'] = active_enrollment
        
        # Get stream and combination for active enrollment
        ctx['stream'] = None
        ctx['combination'] = None
        
        if active_enrollment:
            if hasattr(active_enrollment, 'stream_assignment'):
                ctx['stream'] = active_enrollment.stream_assignment.stream_class
            
            # Get combination for A-Level
            if active_enrollment.class_level.educational_level.level_type == 'A_LEVEL':
                combination_assignment = active_enrollment.combination_assignments.filter(is_active=True).first()
                ctx['combination'] = combination_assignment.combination if combination_assignment else None
        
        ctx['suspensions'] = student.suspensions.order_by('-suspension_date')
        ctx['active_suspension'] = student.suspensions.filter(is_lifted=False).first()
        ctx['has_transfer'] = hasattr(student, 'transfer_out')
        ctx['has_withdrawal'] = hasattr(student, 'withdrawal')
        ctx['audit_logs'] = AuditLog.objects.filter(
            object_id=student.pk,
        ).select_related('user').order_by('-timestamp')[:20]
        
        return ctx


class StudentUpdateView(ManagementRequiredMixin, View):
    """
    Edit an existing student.
    Same save pattern as StudentCreateView — supports the same action values
    and AJAX responses with proper error handling.
    """
    template_name = 'portal_management/students/form.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        if student.status in ['transferred', 'withdrawn']:
            messages.warning(
                request,
                f'This student is {student.get_status_display()}. '
                f'Some fields may be read-only.'
            )
        return render(request, self.template_name, {
            'form': StudentForm(instance=student),
            'student': student,
            'title': f'Edit — {student.full_name}',
            'action': 'Update',
            'is_update': True,
        })

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        form = StudentForm(request.POST, request.FILES, instance=student)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        next_action = request.POST.get('action', 'save_view')

        if form.is_valid():
            try:
                with transaction.atomic():
                    student = form.save()
                    
                    # Verify save was successful
                    if not Student.objects.filter(pk=student.pk).exists():
                        raise ValidationError('Student was not properly saved.')

                    success_message = f'Student {student.full_name} updated successfully.'
                    messages.success(request, success_message)
                    
                    logger.info(
                        f'Student updated: {student.full_name} (ID: {student.pk}) '
                        f'by user {request.user.username}'
                    )

                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'message': success_message,
                            'redirect_url': reverse('management:student_detail', kwargs={'pk': student.pk}),
                        })

                    # Handle redirect based on action
                    if next_action == 'save_list':
                        return redirect('management:student_list')
                    elif next_action == 'save_new':
                        return redirect('management:student_create')
                    elif next_action == 'save_enroll':
                        return redirect('management:student_enroll', pk=student.pk)
                    elif next_action == 'save_parent':
                        return redirect('management:student_parent_management', pk=student.pk)
                    else:
                        return redirect('management:student_detail', pk=student.pk)

            except ValidationError as e:
                error_msg = ', '.join(e.messages) if hasattr(e, 'messages') else str(e)
                messages.error(request, f'Validation error: {error_msg}')
                logger.error('StudentUpdate validation error: %s', e, exc_info=True)
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': error_msg,
                        'field_errors': {}
                    }, status=400)
                    
            except IntegrityError as e:
                error_msg = f'Database integrity error: {str(e)}'
                messages.error(request, error_msg)
                logger.error('StudentUpdate integrity error: %s', e, exc_info=True)
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': error_msg,
                        'field_errors': {}
                    }, status=400)
                    
            except Exception as e:
                error_msg = f'Error updating student: {str(e)}'
                messages.error(request, error_msg)
                logger.error('StudentUpdate error: %s', e, exc_info=True)
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': error_msg,
                        'field_errors': {}
                    }, status=500)
        else:
            # Form is invalid - collect detailed error information
            error_count = len(form.errors)
            error_messages = []
            field_errors = {}
            
            # Log all form errors for debugging
            logger.error(f'StudentUpdate form errors for student ID {pk}:')
            
            for field, errors in form.errors.items():
                if field == '__all__':
                    # Non-field errors
                    for error in errors:
                        error_messages.append(str(error))
                        logger.error(f'  Non-field error: {error}')
                else:
                    # Field-specific errors
                    field_errors[field] = [str(error) for error in errors]
                    for error in errors:
                        error_messages.append(f'{field}: {error}')
                        logger.error(f'  {field}: {error}')
            
            # Create a user-friendly error message
            if error_messages:
                error_summary = ', '.join(error_messages[:3])  # Show first 3 errors
                if len(error_messages) > 3:
                    error_summary += f' and {len(error_messages) - 3} more error(s)'
            else:
                error_summary = f'Please correct the {error_count} error(s) below.'
            
            messages.error(request, error_summary)
            
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'error': error_summary,
                    'field_errors': field_errors,
                    'error_count': error_count,
                    'errors': {
                        field: [str(error) for error in errors]
                        for field, errors in form.errors.items()
                    }
                }, status=400)
            
            # For non-AJAX, keep the form with errors to display in template
            return render(request, self.template_name, {
                'form': form,
                'student': student,
                'title': f'Edit — {student.full_name}',
                'action': 'Update',
                'is_update': True,
                'form_errors': form.errors,  # Pass errors explicitly to template
            })

        # Fallback for non-AJAX errors (should not reach here normally)
        return render(request, self.template_name, {
            'form': form,
            'student': student,
            'title': f'Edit — {student.full_name}',
            'action': 'Update',
            'is_update': True,
        })


# ════════════════════════════════════════════════════════════════════════════
# ENROLLMENT
# ════════════════════════════════════════════════════════════════════════════

class StudentEnrollView(ManagementRequiredMixin, View):
    """
    Enroll an existing student into a class level for an academic year.
    Validates: no duplicate enrollment, A-Level combination required,
    stream capacity, and validates Term belongs to academic year.
    """
    template_name = 'portal_management/students/enroll.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        if student.status not in ['active', 'completed']:
            messages.warning(
                request,
                f'Student is {student.get_status_display()}. '
                f'Only active or completed students can be enrolled.'
            )
            return redirect('management:student_detail', pk=pk)

        active_year = AcademicYear.objects.filter(is_active=True).first()
        current_enrollment = student.enrollments.filter(status='active').first()
        
        # Initialize form with default values
        initial_data = {
            'student': student,
            'academic_year': active_year,
            'enrollment_date': timezone.now().date(),
        }
        form = StudentEnrollmentForm(initial=initial_data)
        
        return render(request, self.template_name, {
            'form': form,
            'student': student,
            'current_enrollment': current_enrollment,
            'active_year': active_year,
            'title': f'Enroll {student.full_name}',
        })

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Create form with POST data
        form = StudentEnrollmentForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    enrollment = form.save(commit=False)
                    enrollment.student = student
                    
                    # Validate enrollment before saving
                    self._validate_enrollment(enrollment)
                    enrollment.save()
                    
                    # Handle combination assignment for A-Level
                    combination_id = request.POST.get('combination')
                    level_type = enrollment.class_level.educational_level.level_type
                    
                    if level_type == 'A_LEVEL' and combination_id:
                        # Create combination assignment
                        combination = get_object_or_404(Combination, pk=combination_id)
                        StudentCombinationAssignment.objects.create(
                            student=student,
                            enrollment=enrollment,
                            combination=combination,
                            assigned_date=timezone.now().date(),
                            is_active=True
                        )
                        logger.info(f'A-Level combination {combination.code} assigned to {student.full_name}')
                    
                    # Handle stream assignment if provided
                    stream_id = request.POST.get('stream_class')
                    if stream_id:
                        stream = get_object_or_404(StreamClass, pk=stream_id)
                        # Validate stream belongs to class level
                        if stream.class_level != enrollment.class_level:
                            raise ValidationError(
                                f'Stream {stream.name} does not belong to {enrollment.class_level.name}'
                            )
                        
                        # Check stream capacity
                        current_count = StudentStreamAssignment.objects.filter(
                            stream_class=stream,
                            enrollment__academic_year=enrollment.academic_year,
                        ).count()
                        
                        if current_count >= stream.capacity:
                            raise ValidationError(
                                f'Stream {stream.name} has reached its maximum capacity of {stream.capacity}'
                            )
                        
                        StudentStreamAssignment.objects.create(
                            enrollment=enrollment,
                            stream_class=stream,
                            assigned_date=timezone.now().date()
                        )
                        logger.info(f'Student {student.full_name} assigned to stream {stream.name}')
                    
                    # Handle action type
                    action = request.POST.get('action', 'save')
                    
                    # Prepare response data
                    response_data = {
                        'success': True,
                        'message': f'{student.full_name} enrolled in {enrollment.class_level.name} for {enrollment.academic_year.name}.',
                        'enrollment_id': enrollment.id,
                    }
                    
                    # Handle redirect based on action
                    if action == 'save':
                        response_data['redirect_url'] = reverse('management:student_detail', args=[pk])
                    elif action == 'save_and_assign_stream':
                        # Redirect to stream assignment page
                        response_data['redirect_url'] = reverse('management:enrollment_assign_stream', args=[enrollment.id])
                    elif action == 'save_add_another':
                        response_data['redirect_url'] = reverse('management:student_enroll', args=[pk])
                    
                    # Log the enrollment
                    logger.info(f'Student {student.full_name} (ID: {student.pk}) enrolled in {enrollment.class_level.name} for {enrollment.academic_year.name}')
                    
                    if is_ajax:
                        return JsonResponse(response_data)
                    else:
                        messages.success(request, response_data['message'])
                        return redirect(response_data['redirect_url'])
                        
            except ValidationError as e:
                error_message = str(e)
                logger.error('Enrollment validation error: %s', error_message, exc_info=True)
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': error_message,
                        'errors': {'__all__': [error_message]}
                    }, status=400)
                else:
                    messages.error(request, error_message)
                    
            except Exception as e:
                error_message = f'Error enrolling student: {str(e)}'
                logger.error('Enrollment error: %s', error_message, exc_info=True)
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': error_message
                    }, status=500)
                else:
                    messages.error(request, error_message)
        else:
            # Form is invalid
            error_count = len(form.errors)
            error_message = f'Please correct the {error_count} error{"s" if error_count > 1 else ""} below.'
            logger.error('Enrollment form errors: %s', form.errors)
            
            if is_ajax:
                # Format errors for AJAX response
                formatted_errors = {}
                for field, errors in form.errors.items():
                    formatted_errors[field] = [str(error) for error in errors]
                
                return JsonResponse({
                    'success': False,
                    'error': error_message,
                    'errors': formatted_errors
                }, status=400)
            else:
                messages.error(request, error_message)
        
        # For non-AJAX requests with errors, return the form
        if not is_ajax:
            current_enrollment = student.enrollments.filter(status='active').first()
            active_year = AcademicYear.objects.filter(is_active=True).first()
            
            return render(request, self.template_name, {
                'form': form,
                'student': student,
                'current_enrollment': current_enrollment,
                'active_year': active_year,
                'title': f'Enroll {student.full_name}',
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    @staticmethod
    def _validate_enrollment(enrollment):
        """
        Validate enrollment business rules
        NOTE: Combination validation is now handled separately since StudentEnrollment
        does NOT have a combination field. Combinations are managed through
        StudentCombinationAssignment model.
        """
        # Check for duplicate enrollment in the same academic year
        if StudentEnrollment.objects.filter(
            student=enrollment.student,
            academic_year=enrollment.academic_year
        ).exists():
            raise ValidationError(
                f'{enrollment.student.full_name} is already enrolled '
                f'in {enrollment.academic_year.name}.'
            )
        
        # Validate that enrollment date is within academic year
        if enrollment.enrollment_date:
            if enrollment.enrollment_date < enrollment.academic_year.start_date:
                raise ValidationError(
                    f'Enrollment date cannot be before academic year start date '
                    f'({enrollment.academic_year.start_date}).'
                )
            if enrollment.enrollment_date > enrollment.academic_year.end_date:
                raise ValidationError(
                    f'Enrollment date cannot be after academic year end date '
                    f'({enrollment.academic_year.end_date}).'
                )
        
        # Check if the student is already in an active enrollment for the current academic year
        current_active = enrollment.student.enrollments.filter(
            academic_year=enrollment.academic_year,
            status='active'
        ).exclude(pk=enrollment.pk)
        
        if current_active.exists():
            raise ValidationError(
                f'{enrollment.student.full_name} already has an active enrollment '
                f'in {enrollment.academic_year.name}.'
            )
        
        # Optional: Validate class level progression
        # This checks if the student is moving to the next appropriate level
        # Get the most recent completed enrollment
        last_enrollment = enrollment.student.enrollments.filter(
            status__in=['promoted', 'completed']
        ).order_by('-academic_year__start_date').first()
        
        if last_enrollment and last_enrollment.class_level:
            # If there was a previous enrollment, ensure progression makes sense
            if enrollment.class_level.order > last_enrollment.class_level.order + 1:
                raise ValidationError(
                    f'Student cannot skip from {last_enrollment.class_level.name} to '
                    f'{enrollment.class_level.name}. They must complete intermediate levels first.'
                )
            
            # Check if moving between educational levels
            last_level_type = last_enrollment.class_level.educational_level.level_type
            new_level_type = enrollment.class_level.educational_level.level_type
            
            if last_level_type != new_level_type:
                # Moving to a new educational level (e.g., Primary → O-Level)
                # This is allowed, but log it for tracking
                import logging
                logger = logging.getLogger(__name__)
                logger.info(
                    f'Student {enrollment.student.full_name} moving from '
                    f'{last_level_type} to {new_level_type}'
                )


# In your views.py file (add this after your StudentEnrollView)

class EnrollmentAssignStreamView(ManagementRequiredMixin, View):
    """
    Assign a stream to an enrolled student.
    This view handles both displaying the stream assignment form and processing it.
    """
    template_name = 'portal_management/students/assign_stream.html'

    def get(self, request, enrollment_id):
        """Display stream assignment form"""
        enrollment = get_object_or_404(StudentEnrollment, pk=enrollment_id)
        
        # Check if enrollment is active
        if enrollment.status != 'active':
            messages.warning(
                request,
                f'Cannot assign stream to enrollment that is {enrollment.get_status_display()}.'
            )
            return redirect('management:student_detail', pk=enrollment.student.pk)
        
        # Check if already has a stream assignment
        existing_assignment = StudentStreamAssignment.objects.filter(enrollment=enrollment).first()
        
        # Get available streams for this class level
        available_streams = StreamClass.objects.filter(
            class_level=enrollment.class_level
        ).order_by('stream_letter')
        
        # Check capacities for each stream
        streams_with_capacity = []
        for stream in available_streams:
            current_count = StudentStreamAssignment.objects.filter(
                stream_class=stream,
                enrollment__academic_year=enrollment.academic_year
            ).count()
            
            streams_with_capacity.append({
                'stream': stream,
                'current_count': current_count,
                'available_capacity': stream.capacity - current_count,
                'has_capacity': current_count < stream.capacity
            })
        
        return render(request, self.template_name, {
            'enrollment': enrollment,
            'existing_assignment': existing_assignment,
            'streams': streams_with_capacity,
            'title': f'Assign Stream - {enrollment.student.full_name}',
        })

    def post(self, request, enrollment_id):
        """Process stream assignment"""
        enrollment = get_object_or_404(StudentEnrollment, pk=enrollment_id)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        stream_id = request.POST.get('stream_class')
        action = request.POST.get('action', 'assign')
        
        try:
            with transaction.atomic():
                # Handle removal of existing assignment
                if action == 'remove':
                    existing = StudentStreamAssignment.objects.filter(enrollment=enrollment).first()
                    if existing:
                        existing.delete()
                        message = f'Stream assignment removed for {enrollment.student.full_name}'
                        logger.info(message)
                        
                        if is_ajax:
                            return JsonResponse({
                                'success': True,
                                'message': message,
                                'redirect_url': reverse('management:student_detail', args=[enrollment.student.pk])
                            })
                        else:
                            messages.success(request, message)
                            return redirect('management:student_detail', pk=enrollment.student.pk)
                    else:
                        raise ValidationError('No stream assignment found to remove.')
                
                # Assign or update stream
                if not stream_id:
                    raise ValidationError('Please select a stream to assign.')
                
                stream = get_object_or_404(StreamClass, pk=stream_id)
                
                # Validate stream belongs to class level
                if stream.class_level != enrollment.class_level:
                    raise ValidationError(
                        f'Stream {stream.name} does not belong to {enrollment.class_level.name}.'
                    )
                
                # Check stream capacity
                current_count = StudentStreamAssignment.objects.filter(
                    stream_class=stream,
                    enrollment__academic_year=enrollment.academic_year,
                ).count()
                
                # Check if we're updating an existing assignment
                existing = StudentStreamAssignment.objects.filter(enrollment=enrollment).first()
                if existing and existing.stream_class == stream:
                    # Same stream - no capacity check needed
                    pass
                elif current_count >= stream.capacity:
                    raise ValidationError(
                        f'Stream {stream.name} has reached its maximum capacity of {stream.capacity} students. '
                        f'Please select a different stream or increase the capacity.'
                    )
                
                # Create or update assignment
                if existing:
                    existing.stream_class = stream
                    existing.assigned_date = timezone.now().date()
                    existing.remarks = request.POST.get('remarks', '')
                    existing.save()
                    message = f'Stream updated to {stream.name} for {enrollment.student.full_name}'
                else:
                    StudentStreamAssignment.objects.create(
                        enrollment=enrollment,
                        stream_class=stream,
                        assigned_date=timezone.now().date(),
                        remarks=request.POST.get('remarks', '')
                    )
                    message = f'{enrollment.student.full_name} assigned to stream {stream.name}'
                
                logger.info(message)
                
                response_data = {
                    'success': True,
                    'message': message,
                    'redirect_url': reverse('management:student_detail', args=[enrollment.student.pk])
                }
                
                if is_ajax:
                    return JsonResponse(response_data)
                else:
                    messages.success(request, message)
                    return redirect(response_data['redirect_url'])
                    
        except ValidationError as e:
            error_message = str(e)
            logger.error('Stream assignment validation error: %s', error_message, exc_info=True)
            
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'error': error_message
                }, status=400)
            else:
                messages.error(request, error_message)
                return redirect('management:enrollment_assign_stream', enrollment_id=enrollment_id)
                
        except Exception as e:
            error_message = f'Error assigning stream: {str(e)}'
            logger.error('Stream assignment error: %s', error_message, exc_info=True)
            
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'error': error_message
                }, status=500)
            else:
                messages.error(request, error_message)
                return redirect('management:enrollment_assign_stream', enrollment_id=enrollment_id)
                            

class StudentStreamAssignView(ManagementRequiredMixin, View):
    """Assign or change stream for an enrolled student."""
    template_name = 'portal_management/students/stream_assign.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        active_enrollment = student.enrollments.filter(status='active').first()
        if not active_enrollment:
            messages.warning(request, f'{student.full_name} is not currently enrolled.')
            return redirect('management:student_detail', pk=pk)
        current_stream = None
        if hasattr(active_enrollment, 'stream_assignment'):
            current_stream = active_enrollment.stream_assignment.stream_class
        return render(request, self.template_name, {
            'student': student,
            'enrollment': active_enrollment,
            'streams': StreamClass.objects.filter(
                class_level=active_enrollment.class_level
            ).order_by('stream_letter'),
            'current_stream': current_stream,
            'title': f'Assign Stream — {student.full_name}',
        })

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        active_enrollment = student.enrollments.filter(status='active').first()
        if not active_enrollment:
            messages.error(request, 'No active enrollment found.')
            return redirect('management:student_detail', pk=pk)
        try:
            with transaction.atomic():
                from core.models import StudentStreamAssignment
                stream_id = request.POST.get('stream')
                if not stream_id:
                    if hasattr(active_enrollment, 'stream_assignment'):
                        active_enrollment.stream_assignment.delete()
                        messages.success(
                            request, f'Stream assignment removed for {student.full_name}.'
                        )
                else:
                    stream = get_object_or_404(StreamClass, pk=stream_id)
                    count = StudentStreamAssignment.objects.filter(
                        stream_class=stream,
                        enrollment__academic_year=active_enrollment.academic_year
                    ).exclude(enrollment=active_enrollment).count()
                    if count >= stream.capacity:
                        raise ValidationError(
                            f'Stream {stream.name} has reached maximum capacity.'
                        )
                    if hasattr(active_enrollment, 'stream_assignment'):
                        sa = active_enrollment.stream_assignment
                        sa.stream_class = stream
                        sa.assigned_date = timezone.now().date()
                        sa.save()
                    else:
                        StudentStreamAssignment.objects.create(
                            enrollment=active_enrollment,
                            stream_class=stream,
                            assigned_date=timezone.now().date()
                        )
                    messages.success(
                        request, f'{student.full_name} assigned to stream {stream.name}.'
                    )
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error assigning stream: {e}')
            logger.error('Stream assignment error: %s', e, exc_info=True)
        return redirect('management:student_detail', pk=pk)


# ════════════════════════════════════════════════════════════════════════════
# DRAFTS
# ════════════════════════════════════════════════════════════════════════════

class StudentDraftListView(ManagementRequiredMixin, ListView):
    """List all students with draft status."""
    template_name = 'portal_management/students/drafts/list.html'
    context_object_name = 'drafts'
    paginate_by = 20

    def get_queryset(self):
        return Student.objects.filter(status='draft').order_by('-updated_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total_drafts'] = self.get_queryset().count()
        return ctx


class StudentDraftCreateView(ManagementRequiredMixin, View):
    """Save partial student data as a draft via AJAX (JSON body)."""

    def post(self, request):
        try:
            data = json.loads(request.body)
            with transaction.atomic():
                student = Student.objects.create(
                    first_name=data.get('first_name', ''),
                    middle_name=data.get('middle_name', ''),
                    last_name=data.get('last_name', ''),
                    gender=data.get('gender', ''),
                    date_of_birth=data.get('date_of_birth') or None,
                    address=data.get('address', ''),
                    national_id=data.get('national_id', ''),
                    physical_disability=data.get('physical_disability', ''),
                    admission_date=data.get('admission_date') or timezone.now().date(),
                    status='draft',
                )
                request.session[f'draft_{student.pk}'] = data
                return JsonResponse({
                    'success': True,
                    'draft_id': student.pk,
                    'message': 'Draft saved successfully.'
                })
        except Exception as e:
            logger.error('Draft create error: %s', e, exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class StudentDraftEditView(ManagementRequiredMixin, View):
    """Edit an existing draft student."""
    template_name = 'portal_management/students/drafts/edit.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk, status='draft')
        draft_data = request.session.get(f'draft_{student.pk}', {})
        return render(request, self.template_name, {
            'student': student,
            'draft_data': draft_data,
            'title': f'Edit Draft — {student.full_name or "New Student"}',
        })


class StudentDraftPublishView(ManagementRequiredMixin, View):
    """Promote a draft student to active status."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk, status='draft')
        try:
            with transaction.atomic():
                student.status = 'active'
                student.save()
                request.session.pop(f'draft_{student.pk}', None)
                messages.success(
                    request,
                    f'Draft published. Registration: {student.registration_number}'
                )
                return redirect('management:student_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error publishing draft: {e}')
            logger.error('Draft publish error: %s', e, exc_info=True)
            return redirect('management:student_draft_edit', pk=pk)


class StudentDraftDeleteView(ManagementRequiredMixin, View):
    """Permanently delete a draft student record."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk, status='draft')
        try:
            student.delete()
            request.session.pop(f'draft_{pk}', None)
            messages.success(request, 'Draft deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting draft: {e}')
            logger.error('Draft delete error: %s', e, exc_info=True)
        return redirect('management:student_draft_list')


# ════════════════════════════════════════════════════════════════════════════
# LIFECYCLE
# ════════════════════════════════════════════════════════════════════════════

class StudentSuspendView(ManagementRequiredMixin, View):
    """Suspend a student. Model.save() auto-sets status='suspended'."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        try:
            with transaction.atomic():
                StudentSuspension.objects.create(
                    student=student,
                    suspension_date=request.POST.get('suspension_date'),
                    expected_return_date=request.POST.get(
                        'expected_return_date'
                    ) or None,
                    reason=request.POST.get('reason'),
                    remarks=request.POST.get('remarks', ''),
                    authorised_by=getattr(request.user, 'staff_profile', None),
                )
                messages.success(request, f'{student.full_name} has been suspended.')
        except Exception as e:
            messages.error(request, f'Error suspending student: {e}')
            logger.error('Suspend error: %s', e, exc_info=True)
        return redirect('management:student_detail', pk=pk)


class StudentLiftSuspensionView(ManagementRequiredMixin, View):
    """Lift an active suspension. Model.save() restores status='active'."""

    def post(self, request, pk, suspension_pk):
        student = get_object_or_404(Student, pk=pk)
        suspension = get_object_or_404(
            StudentSuspension, pk=suspension_pk, student=student
        )
        try:
            suspension.is_lifted = True
            suspension.lifted_date = timezone.now().date()
            suspension.lifted_by = getattr(request.user, 'staff_profile', None)
            suspension.save()
            messages.success(
                request,
                f'Suspension lifted. {student.full_name} is now active.'
            )
        except Exception as e:
            messages.error(request, f'Error lifting suspension: {e}')
            logger.error('Lift suspension error: %s', e, exc_info=True)
        return redirect('management:student_detail', pk=pk)


class StudentTransferView(ManagementRequiredMixin, View):
    """Transfer student out. Model.save() auto-sets status='transferred'."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        if hasattr(student, 'transfer_out'):
            messages.warning(
                request, f'{student.full_name} already has a transfer record.'
            )
            return redirect('management:student_detail', pk=pk)
        try:
            with transaction.atomic():
                ae = student.enrollments.filter(
                    status='active'
                ).select_related('class_level', 'academic_year').first()
                StudentTransferOut.objects.create(
                    student=student,
                    transfer_date=request.POST.get('transfer_date'),
                    destination_school_name=request.POST.get('destination_school', ''),
                    reason=request.POST.get('reason', 'voluntary'),
                    last_class_level=ae.class_level if ae else None,
                    last_academic_year=ae.academic_year if ae else None,
                    transfer_letter_issued=(
                        request.POST.get('transfer_letter_issued') == 'on'
                    ),
                    transcript_issued=(
                        request.POST.get('transcript_issued') == 'on'
                    ),
                    remarks=request.POST.get('remarks', ''),
                    authorised_by=getattr(request.user, 'staff_profile', None),
                )
                messages.success(
                    request,
                    f'{student.full_name} transferred out. Status updated.'
                )
        except Exception as e:
            messages.error(request, f'Error transferring student: {e}')
            logger.error('Transfer error: %s', e, exc_info=True)
        return redirect('management:student_detail', pk=pk)


class StudentWithdrawView(ManagementRequiredMixin, View):
    """Record withdrawal. Model.save() auto-sets status='withdrawn'."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        if hasattr(student, 'withdrawal'):
            messages.warning(
                request, f'{student.full_name} already has a withdrawal record.'
            )
            return redirect('management:student_detail', pk=pk)
        try:
            with transaction.atomic():
                from core.models import StudentWithdrawal
                ae = student.enrollments.filter(
                    status='active'
                ).select_related('class_level', 'academic_year').first()
                StudentWithdrawal.objects.create(
                    student=student,
                    withdrawal_date=request.POST.get('withdrawal_date'),
                    reason=request.POST.get('reason', 'other'),
                    last_class_level=ae.class_level if ae else None,
                    last_academic_year=ae.academic_year if ae else None,
                    transcript_issued=(
                        request.POST.get('transcript_issued') == 'on'
                    ),
                    remarks=request.POST.get('remarks', ''),
                    authorised_by=getattr(request.user, 'staff_profile', None),
                )
                messages.success(
                    request,
                    f'{student.full_name} withdrawal recorded. Status updated.'
                )
        except Exception as e:
            messages.error(request, f'Error recording withdrawal: {e}')
            logger.error('Withdrawal error: %s', e, exc_info=True)
        return redirect('management:student_detail', pk=pk)


# ════════════════════════════════════════════════════════════════════════════
# ACCOUNT
# ════════════════════════════════════════════════════════════════════════════

class StudentResetPasswordView(ManagementRequiredMixin, View):
    """Reset student portal password back to their registration number (AJAX)."""

    def post(self, request, user_id):
        try:
            from core.models import CustomUser, UserType
            user = get_object_or_404(
                CustomUser, pk=user_id, user_type=UserType.STUDENT
            )
            user.set_password(user.username)
            user.save(update_fields=['password'])
            logger.info(
                'Password reset: student=%s by=%s', user.username, request.user
            )
            return JsonResponse({'success': True})
        except Exception as e:
            logger.error('Password reset error: %s', e, exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ════════════════════════════════════════════════════════════════════════════
# PARENT MANAGEMENT
# ════════════════════════════════════════════════════════════════════════════

class StudentParentManagementView(ManagementRequiredMixin, View):
    """Comprehensive parent management page for a student."""
    template_name = 'portal_management/students/parent_management.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        relationships = StudentParent.objects.filter(
            student=student
        ).select_related('parent').order_by(
            '-is_primary_contact', 'parent__full_name'
        )
        
        # Calculate counts in the view
        total_parents = relationships.count()
        primary_count = relationships.filter(is_primary_contact=True).count()
        fee_responsible_count = relationships.filter(is_fee_responsible=True).count()
        
        recent_parents = Parent.objects.order_by('-id')[:5]
        
        return render(request, self.template_name, {
            'student': student,
            'relationships': relationships,
            'recent_parents': recent_parents,
            'parent_form': StudentParentForm(),
            'title': f'Manage Parents — {student.full_name}',
            'relationship_choices': Parent.RELATIONSHIP_CHOICES,
            # Add these counts
            'total_parents': total_parents,
            'primary_count': primary_count,
            'fee_responsible_count': fee_responsible_count,
        })
    

class StudentAddParentView(ManagementRequiredMixin, View):
    """Add a new parent or link an existing parent to a student."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        action = request.POST.get('parent_action', 'create_new')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        try:
            with transaction.atomic():
                if action == 'link_existing':
                    result = self._link_existing(request, student)
                else:
                    result = self._create_new(request, student)

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': result['message'],
                        'redirect_url': reverse(
                            'management:student_parent_management',
                            kwargs={'pk': student.pk}
                        ),
                    })
                messages.success(request, result['message'])
                return redirect('management:student_parent_management', pk=pk)

        except ValidationError as e:
            error_msg = ', '.join(e.messages) if hasattr(e, 'messages') else str(e)
            logger.error('AddParent validation error student=%s: %s', pk, e,
                         exc_info=True)
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg,
                                     'errors': getattr(e, 'message_dict',
                                                       {'__all__': [error_msg]})})
            messages.error(request, f'Error: {error_msg}')
            return redirect('management:student_parent_management', pk=pk)

        except Exception as e:
            logger.error('AddParent error student=%s: %s', pk, e, exc_info=True)
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)},
                                    status=500)
            messages.error(request, f'Error adding parent: {e}')
            return redirect('management:student_parent_management', pk=pk)

    # ── Private helpers ───────────────────────────────────────────────────

    def _link_existing(self, request, student):
        parent_id = request.POST.get('parent_id')
        if not parent_id:
            raise ValidationError('Please select a parent to link.')
        parent = get_object_or_404(Parent, pk=parent_id)
        relationship, created = StudentParent.objects.get_or_create(
            student=student, parent=parent,
            defaults={
                'is_primary_contact': (
                    request.POST.get('is_primary_contact_link') == 'on'
                ),
                'is_fee_responsible': (
                    request.POST.get('is_fee_responsible_link') == 'on'
                ),
                'fee_responsible_from': (
                    request.POST.get('fee_responsible_from_link') or None
                ),
            }
        )
        if not created:
            relationship.is_primary_contact = (
                request.POST.get('is_primary_contact_link') == 'on'
            )
            relationship.is_fee_responsible = (
                request.POST.get('is_fee_responsible_link') == 'on'
            )
            if request.POST.get('fee_responsible_from_link'):
                relationship.fee_responsible_from = (
                    request.POST.get('fee_responsible_from_link')
                )
            relationship.save()
        if relationship.is_primary_contact:
            StudentParent.objects.filter(
                student=student
            ).exclude(pk=relationship.pk).update(is_primary_contact=False)
        msg = (
            f'{parent.full_name} linked to student successfully.'
            if created else
            f'{parent.full_name}\'s relationship updated successfully.'
        )
        return {'message': msg}

    def _create_new(self, request, student):
        form = StudentParentForm(request.POST)
        if not form.is_valid():
            raise ValidationError(form.errors)
        parent = form.save()
        relationship = StudentParent.objects.create(
            student=student, parent=parent,
            is_primary_contact=request.POST.get('is_primary_contact') == 'on',
            is_fee_responsible=request.POST.get('is_fee_responsible') == 'on',
            fee_responsible_from=request.POST.get('fee_responsible_from') or None,
        )
        if relationship.is_primary_contact:
            StudentParent.objects.filter(
                student=student
            ).exclude(pk=relationship.pk).update(is_primary_contact=False)
        return {'message': f'Parent {parent.full_name} added and linked successfully.'}


class StudentParentUpdateView(ManagementRequiredMixin, View):
    """Update the is_primary_contact / is_fee_responsible flags on a relationship."""

    def post(self, request, pk, relationship_pk):
        relationship = get_object_or_404(
            StudentParent, pk=relationship_pk, student_id=pk
        )
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            with transaction.atomic():
                # Handle is_primary_contact - accepts 'on', 'true', True, '1', 1
                primary_val = request.POST.get('is_primary_contact')
                if primary_val is not None:
                    relationship.is_primary_contact = self._to_boolean(primary_val)
                
                # Handle is_fee_responsible - accepts 'on', 'true', True, '1', 1
                fee_val = request.POST.get('is_fee_responsible')
                if fee_val is not None:
                    relationship.is_fee_responsible = self._to_boolean(fee_val)
                
                # Handle fee_responsible_from date
                fee_from = request.POST.get('fee_responsible_from')
                if fee_from:
                    relationship.fee_responsible_from = fee_from
                
                relationship.save()
                
                # Handle primary contact logic
                if relationship.is_primary_contact:
                    StudentParent.objects.filter(
                        student_id=pk
                    ).exclude(pk=relationship_pk).update(is_primary_contact=False)
                
                msg = f'{relationship.parent.full_name}\'s relationship updated.'
                
                if is_ajax:
                    return JsonResponse({
                        'success': True, 
                        'message': msg,
                        'data': {
                            'is_primary_contact': relationship.is_primary_contact,
                            'is_fee_responsible': relationship.is_fee_responsible,
                            'fee_responsible_from': relationship.fee_responsible_from,
                        }
                    })
                messages.success(request, msg)
                
        except Exception as e:
            logger.error('ParentUpdate error: %s', e, exc_info=True)
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f'Error: {e}')
            
        return redirect('management:student_parent_management', pk=pk)
    
    def _to_boolean(self, value):
        """Convert various truthy values to boolean."""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, str)):
            if str(value).lower() in ('true', '1', 'on', 'yes'):
                return True
        return False
    

class StudentParentRemoveView(ManagementRequiredMixin, View):
    """Unlink a parent from a student (StudentParent junction record deleted)."""

    def post(self, request, pk, relationship_pk):
        relationship = get_object_or_404(
            StudentParent, pk=relationship_pk, student_id=pk
        )
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            parent_name = relationship.parent.full_name
            was_primary = relationship.is_primary_contact
            with transaction.atomic():
                relationship.delete()
                new_primary = None
                if was_primary:
                    next_rel = StudentParent.objects.filter(
                        student_id=pk
                    ).first()
                    if next_rel:
                        next_rel.is_primary_contact = True
                        next_rel.save()
                        new_primary = {
                            'id': next_rel.pk,
                            'name': next_rel.parent.full_name,
                        }
                msg = f'{parent_name} removed from student\'s record.'
                if is_ajax:
                    return JsonResponse({
                        'success': True, 'message': msg,
                        'removed_id': relationship_pk,
                        'was_primary': was_primary,
                        'new_primary': new_primary,
                    })
                messages.success(request, msg)
        except Exception as e:
            logger.error('ParentRemove error: %s', e, exc_info=True)
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f'Error removing parent: {e}')
        return redirect('management:student_parent_management', pk=pk)


class StudentParentSetPrimaryView(ManagementRequiredMixin, View):
    """Set one parent as the primary contact, clearing all others."""

    def post(self, request, pk, relationship_pk):
        relationship = get_object_or_404(
            StudentParent, pk=relationship_pk, student_id=pk
        )
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        try:
            with transaction.atomic():
                StudentParent.objects.filter(
                    student_id=pk
                ).update(is_primary_contact=False)
                relationship.is_primary_contact = True
                relationship.save()
                msg = (
                    f'{relationship.parent.full_name} is now the primary contact.'
                )
                if is_ajax:
                    return JsonResponse({
                        'success': True, 'message': msg,
                        'relationship_id': relationship.pk,
                        'parent_name': relationship.parent.full_name,
                    })
                messages.success(request, msg)
        except Exception as e:
            logger.error('SetPrimary error: %s', e, exc_info=True)
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f'Error: {e}')
        return redirect('management:student_parent_management', pk=pk)


class StudentParentBulkAddView(ManagementRequiredMixin, View):
    """Link multiple existing parents to a student at once."""

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        parent_ids = request.POST.getlist('parent_ids')
        if not parent_ids:
            messages.warning(request, 'No parents selected.')
            return redirect('management:student_parent_management', pk=pk)
        try:
            with transaction.atomic():
                added = 0
                for pid in parent_ids:
                    parent = get_object_or_404(Parent, pk=pid)
                    _, created = StudentParent.objects.get_or_create(
                        student=student, parent=parent,
                        defaults={
                            'is_primary_contact': False,
                            'is_fee_responsible': False,
                        }
                    )
                    if created:
                        added += 1
                if added:
                    messages.success(
                        request,
                        f'Successfully added {added} parent(s) to {student.full_name}.'
                    )
                else:
                    messages.info(request, 'Selected parents were already linked.')
        except Exception as e:
            messages.error(request, f'Error adding parents: {e}')
            logger.error('BulkAdd error: %s', e, exc_info=True)
        return redirect('management:student_parent_management', pk=pk)


class ParentEditView(ManagementRequiredMixin, View):
    """Edit a Parent record. Returns JSON on AJAX, redirects on normal POST."""
    template_name = 'portal_management/students/edit_parent_modal.html'

    def get(self, request, parent_pk):
        parent = get_object_or_404(Parent, pk=parent_pk)
        return render(request, self.template_name, {
            'parent': parent,
            'form': ParentForm(instance=parent),
        })

    def post(self, request, parent_pk):
        parent = get_object_or_404(Parent, pk=parent_pk)
        form = ParentForm(request.POST, instance=parent)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if form.is_valid():
            try:
                parent = form.save()
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': f'{parent.full_name} updated successfully.',
                        'parent': {
                            'id': parent.pk,
                            'full_name': parent.full_name,
                            'phone_number': parent.phone_number,
                            'email': parent.email or '',
                            'relationship': parent.get_relationship_display(),
                        }
                    })
                messages.success(request, f'{parent.full_name} updated successfully.')
                return redirect(request.META.get('HTTP_REFERER',
                                                  'management:student_list'))
            except Exception as e:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': str(e)},
                                        status=500)
                messages.error(request, f'Error updating parent: {e}')
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': form.errors},
                                    status=400)
        return render(request, self.template_name, {
            'parent': parent, 'form': form,
        })


class ParentDeleteView(ManagementRequiredMixin, View):
    """Delete a parent only if not linked to any students. AJAX only."""

    def delete(self, request, parent_pk):
        parent = get_object_or_404(Parent, pk=parent_pk)
        linked = parent.students.count()
        if linked:
            return JsonResponse({
                'success': False,
                'error': (
                    f'Cannot delete {parent.full_name} — '
                    f'linked to {linked} student(s).'
                )
            }, status=400)
        try:
            name = parent.full_name
            parent.delete()
            return JsonResponse({'success': True,
                                  'message': f'{name} deleted successfully.'})
        except Exception as e:
            logger.error('ParentDelete error: %s', e, exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class GetParentDetailsView(ManagementRequiredMixin, View):
    """AJAX — return parent details as JSON for the edit modal."""

    def get(self, request, parent_pk):
        parent = get_object_or_404(Parent, pk=parent_pk)
        return JsonResponse({
            'id': parent.pk,
            'full_name': parent.full_name,
            'relationship': parent.relationship,
            'phone_number': parent.phone_number,
            'alternate_phone': parent.alternate_phone or '',
            'email': parent.email or '',
            'address': parent.address or '',
        })


class ParentListView(ManagementRequiredMixin, ListView):
    """List all parents/guardians with search support."""
    model = Parent
    template_name = 'portal_management/students/parent_list.html'
    context_object_name = 'parents'
    paginate_by = 20

    def get_queryset(self):
        qs = Parent.objects.prefetch_related('students').order_by('full_name')
        search = self.request.GET.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(phone_number__icontains=search) |
                Q(email__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total'] = Parent.objects.count()
        ctx['search'] = self.request.GET.get('search', '')
        return ctx


class ParentCreateView(ManagementRequiredMixin, View):
    """Create a standalone parent record (not yet linked to any student)."""
    template_name = 'portal_management/students/parent_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': ParentForm(),
            'title': 'Add Parent / Guardian',
        })

    def post(self, request):
        form = ParentForm(request.POST)
        if form.is_valid():
            parent = form.save()
            messages.success(request, f'Parent {parent.full_name} added.')
            return redirect('management:parent_list')
        return render(request, self.template_name, {
            'form': form,
            'title': 'Add Parent / Guardian',
        })


class ParentUpdateView(ManagementRequiredMixin, View):
    """Update a parent/guardian record."""
    template_name = 'portal_management/parents/form.html'

    def get(self, request, pk):
        parent = get_object_or_404(Parent, pk=pk)
        return render(request, self.template_name, {
            'form': ParentForm(instance=parent),
            'parent': parent,
            'title': f'Edit — {parent.full_name}',
            'action': 'Update',
        })

    def post(self, request, pk):
        parent = get_object_or_404(Parent, pk=pk)
        form = ParentForm(request.POST, instance=parent)
        if form.is_valid():
            try:
                parent = form.save()
                messages.success(
                    request, f'Parent {parent.full_name} updated successfully.'
                )
                return redirect('management:parent_list')
            except Exception as e:
                messages.error(request, f'Error updating parent: {e}')
                logger.error('ParentUpdate error: %s', e, exc_info=True)
        else:
            messages.error(request, 'Please correct the errors below.')
        return render(request, self.template_name, {
            'form': form,
            'parent': parent,
            'title': f'Edit — {parent.full_name}',
            'action': 'Update',
        })


# ════════════════════════════════════════════════════════════════════════════
# AJAX HELPERS
# ════════════════════════════════════════════════════════════════════════════

class SearchParentsView(ManagementRequiredMixin, View):
    """AJAX — search parents by name or phone. Requires ?q= (min 2 chars)."""

    def get(self, request):
        query = request.GET.get('q', '').strip()
        if len(query) < 2:
            return JsonResponse([], safe=False)
        parents = Parent.objects.filter(
            Q(full_name__icontains=query) |
            Q(phone_number__icontains=query)
        ).order_by('full_name')[:10]
        return JsonResponse([{
            'id': p.pk,
            'full_name': p.full_name,
            'phone_number': p.phone_number,
            'email': p.email or '',
            'relationship': p.get_relationship_display(),
            'initials': p.full_name[:2].upper(),
        } for p in parents], safe=False)


class GetStreamsView(ManagementRequiredMixin, View):
    """AJAX — return streams for a class level. Requires ?class_level_id=."""

    def get(self, request):
        class_level_id = request.GET.get('class_level_id')
        if not class_level_id:
            return JsonResponse([], safe=False)
        streams = StreamClass.objects.filter(
            class_level_id=class_level_id
        ).order_by('stream_letter')
        return JsonResponse([{
            'id': s.pk,
            'name': s.name or str(s),
            'capacity': s.capacity,
            'student_count': s.student_count,
            'available': s.capacity - s.student_count,
        } for s in streams], safe=False)


class GetCombinationsView(ManagementRequiredMixin, View):
    """AJAX — return A-Level combinations for a class level. Requires ?class_level_id=."""

    def get(self, request):
        class_level_id = request.GET.get('class_level_id')
        if not class_level_id:
            return JsonResponse([], safe=False)
        cl = get_object_or_404(
            ClassLevel.objects.select_related('educational_level'),
            pk=class_level_id
        )
        if cl.educational_level.level_type != 'A_LEVEL':
            return JsonResponse([], safe=False)
        combos = Combination.objects.filter(
            educational_level=cl.educational_level
        ).order_by('code')
        return JsonResponse([{
            'id': c.pk,
            'name': c.name,
            'code': c.code,
            'display': f'{c.code} — {c.name}',
        } for c in combos], safe=False)



class StudentExamSessionsView(ManagementRequiredMixin, View):
    """
    Display all exam sessions conducted for a specific student across all enrollments.
    Shows a list of exam sessions with summary information and links to view detailed results.
    Handles different educational levels appropriately.
    """
    template_name = 'portal_management/students/student_exam_sessions.html'

    def get(self, request, student_pk):
        student = get_object_or_404(
            Student.objects.select_related('user'),
            pk=student_pk
        )
        
        # Get all enrollments for this student with educational level info
        enrollments = StudentEnrollment.objects.filter(
            student=student
        ).select_related(
            'academic_year',
            'class_level__educational_level'
        ).order_by('-academic_year__start_date')
        
        if not enrollments.exists():
            messages.info(request, f'No enrollment records found for {student.full_name}.')
            # Return early if no enrollments
            return render(request, self.template_name, {
                'student': student,
                'enrollments': [],
                'sessions_data': [],
                'sessions_by_year': {},
                'stats': self._empty_stats(),
                'educational_levels_summary': [],
                'total_sessions': 0,
                'generated_date': date.today(),
            })
        
        # Extract IDs to avoid complex Q objects
        class_level_ids = [e.class_level_id for e in enrollments]
        academic_year_ids = [e.academic_year_id for e in enrollments]
        
        # Get all exam sessions for the class levels and academic years this student was enrolled in
        exam_sessions = ExamSession.objects.filter(
            class_level_id__in=class_level_ids,
            academic_year_id__in=academic_year_ids
        ).select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level__educational_level',
            'stream_class'
        ).prefetch_related(
            Prefetch('student_metrics', 
                     queryset=StudentExamMetrics.objects.filter(student=student),
                     to_attr='student_metrics_list'),
            Prefetch('student_positions',
                     queryset=StudentExamPosition.objects.filter(student=student),
                     to_attr='student_positions_list')
        ).order_by('-exam_date')
        
        # Annotate each session with student's performance data based on educational level
        sessions_data = []
        
        for session in exam_sessions:
            educational_level = session.class_level.educational_level
            level_type = educational_level.level_type
            is_primary_nursery = level_type in ['PRIMARY', 'NURSERY']
            is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
            
            # Get student's metrics for this session
            metrics = None
            if hasattr(session, 'student_metrics_list') and session.student_metrics_list:
                metrics = session.student_metrics_list[0]
            
            # Get student's position for this session
            position = None
            if hasattr(session, 'student_positions_list') and session.student_positions_list:
                position = session.student_positions_list[0]
            
            # Check if student has results for this session
            has_results = metrics is not None
            
            # Get enrollment info for this session's academic year and class level
            enrollment = enrollments.filter(
                academic_year=session.academic_year,
                class_level=session.class_level
            ).first()
            
            # Calculate overall grade for Primary/Nursery if metrics available
            overall_grade = None
            if is_primary_nursery and metrics and metrics.average_marks:
                grading_scale = GradingScale.objects.filter(
                    education_level=educational_level
                ).order_by('-min_mark')
                for gs in grading_scale:
                    if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                        overall_grade = gs.grade
                        break
            
            # Get division scale info for secondary level
            division_info = None
            if is_secondary and metrics and metrics.division:
                try:
                    division_scale = DivisionScale.objects.get(
                        education_level=educational_level,
                        division=metrics.division
                    )
                    division_info = {
                        'min_points': division_scale.min_points,
                        'max_points': division_scale.max_points,
                    }
                except DivisionScale.DoesNotExist:
                    pass
            
            sessions_data.append({
                'session': session,
                'educational_level': educational_level,
                'level_type': level_type,
                'is_primary_nursery': is_primary_nursery,
                'is_secondary': is_secondary,
                'metrics': metrics,
                'position': position,
                'has_results': has_results,
                'enrollment': enrollment,
                # Common metrics
                'total_marks': metrics.total_marks if metrics else None,
                'average_marks': metrics.average_marks if metrics else None,
                'class_position': position.class_position if position else None,
                'stream_position': position.stream_position if position else None,
                # Primary/Nursery specific
                'overall_grade': overall_grade,
                # Secondary specific
                'division': metrics.division if metrics and is_secondary else None,
                'total_points': metrics.total_points if metrics and is_secondary else None,
                'division_info': division_info,
            })
        
        # Calculate statistics with educational level consideration
        stats = self._calculate_stats(sessions_data)
        
        # Group sessions by academic year
        sessions_by_year = {}
        for session_data in sessions_data:
            year_name = session_data['session'].academic_year.name
            if year_name not in sessions_by_year:
                sessions_by_year[year_name] = []
            sessions_by_year[year_name].append(session_data)
        
        # Get educational level summary
        educational_levels_summary = self._get_educational_levels_summary(enrollments)
        
        context = {
            'student': student,
            'enrollments': enrollments,
            'sessions_data': sessions_data,
            'sessions_by_year': sessions_by_year,
            'stats': stats,
            'educational_levels_summary': educational_levels_summary,
            'total_sessions': len(sessions_data),
            'generated_date': date.today(),
        }
        
        return render(request, self.template_name, context)
    
    def _empty_stats(self):
        """Return empty stats structure when no enrollments exist."""
        return {
            'total_sessions': 0,
            'sessions_with_results': 0,
            'sessions_without_results': 0,
            'by_level_type': {
                'PRIMARY': {'total': 0, 'with_results': 0},
                'NURSERY': {'total': 0, 'with_results': 0},
                'O_LEVEL': {'total': 0, 'with_results': 0},
                'A_LEVEL': {'total': 0, 'with_results': 0},
            },
            'best_performance': {
                'primary_nursery': {'grade': None, 'session': None, 'average': None},
                'secondary': {'division': None, 'points': None, 'session': None, 'total_marks': None},
            },
            'best_total_marks': None,
            'best_average': None,
            'best_division': None,
            'best_grade': None,
        }
    
    def _calculate_stats(self, sessions_data):
        """Calculate statistics considering different educational levels."""
        stats = {
            'total_sessions': len(sessions_data),
            'sessions_with_results': 0,
            'sessions_without_results': 0,
            'by_level_type': {
                'PRIMARY': {'total': 0, 'with_results': 0},
                'NURSERY': {'total': 0, 'with_results': 0},
                'O_LEVEL': {'total': 0, 'with_results': 0},
                'A_LEVEL': {'total': 0, 'with_results': 0},
            },
            'best_performance': {
                'primary_nursery': {'grade': None, 'session': None, 'average': None},
                'secondary': {'division': None, 'points': None, 'session': None, 'total_marks': None},
            },
            'best_total_marks': None,
            'best_average': None,
            'best_division': None,
            'best_grade': None,
        }
        
        # Count sessions with results
        sessions_with_results = [s for s in sessions_data if s['has_results']]
        stats['sessions_with_results'] = len(sessions_with_results)
        stats['sessions_without_results'] = stats['total_sessions'] - stats['sessions_with_results']
        
        # Count by level type
        for session in sessions_data:
            level_type = session['level_type']
            if level_type in stats['by_level_type']:
                stats['by_level_type'][level_type]['total'] += 1
                if session['has_results']:
                    stats['by_level_type'][level_type]['with_results'] += 1
        
        # Find best performance by level
        primary_nursery_sessions = [
            s for s in sessions_with_results 
            if s['is_primary_nursery'] and s['overall_grade']
        ]
        if primary_nursery_sessions:
            # For Primary/Nursery, find best grade (A > B > C...)
            grade_order = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'E': 1, 'F': 0}
            best_grade_session = max(
                primary_nursery_sessions,
                key=lambda x: grade_order.get(x['overall_grade'], -1)
            )
            stats['best_performance']['primary_nursery'] = {
                'grade': best_grade_session['overall_grade'],
                'session': best_grade_session['session'],
                'average': best_grade_session['average_marks'],
            }
            stats['best_grade'] = best_grade_session['overall_grade']
        
        # For Secondary, find best division (I > II > III > IV > 0)
        secondary_sessions = [
            s for s in sessions_with_results 
            if s['is_secondary'] and s['division']
        ]
        if secondary_sessions:
            division_order = {'I': 5, 'II': 4, 'III': 3, 'IV': 2, '0': 1}
            best_division_session = max(
                secondary_sessions,
                key=lambda x: division_order.get(x['division'], 0)
            )
            stats['best_performance']['secondary'] = {
                'division': best_division_session['division'],
                'points': best_division_session['total_points'],
                'session': best_division_session['session'],
                'total_marks': best_division_session['total_marks'],
            }
            stats['best_division'] = best_division_session['division']
        
        # Find best total marks across all levels
        if sessions_with_results:
            best_total = max((s['total_marks'] for s in sessions_with_results if s['total_marks']), default=0)
            if best_total:
                stats['best_total_marks'] = best_total
            
            best_avg = max((s['average_marks'] for s in sessions_with_results if s['average_marks']), default=0)
            if best_avg:
                stats['best_average'] = best_avg
        
        return stats
    
    def _get_educational_levels_summary(self, enrollments):
        """Get summary of educational levels the student has been enrolled in."""
        levels_summary = []
        seen_levels = set()
        
        for enrollment in enrollments:
            level = enrollment.class_level.educational_level
            if level.pk not in seen_levels:
                seen_levels.add(level.pk)
                levels_summary.append({
                    'name': level.name,
                    'level_type': level.get_level_type_display(),
                    'class_levels': [e.class_level.name for e in enrollments if e.class_level.educational_level == level],
                })
        
        return levels_summary
    



class StudentResultDetailView(ManagementRequiredMixin, View):
    """
    Display comprehensive analytics for a student in a specific exam session.
    Handles different educational levels appropriately:
        - Primary/Nursery: Show grades only, no points/divisions
        - O-Level/A-Level: Show points and divisions
    """
    template_name = 'portal_management/students/student_result_detail.html'

    def get(self, request, student_pk, session_pk):
        student = get_object_or_404(Student, pk=student_pk)
        session = get_object_or_404(
            ExamSession.objects.select_related(
                'class_level__educational_level',
                'academic_year',
                'term',
                'exam_type'
            ), pk=session_pk
        )
        
        # Get educational level info
        educational_level = session.class_level.educational_level
        level_type = educational_level.level_type
        is_primary_nursery = level_type in ['PRIMARY', 'NURSERY']
        is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
        
        # Check if student was enrolled in this session
        enrollment = student.enrollments.filter(
            academic_year=session.academic_year,
            class_level=session.class_level
        ).first()
        
        if not enrollment:
            messages.warning(request, f'{student.full_name} was not enrolled in this session.')
            return redirect('management:student_exam_sessions', student_pk=student_pk)
        
        # Get student's metrics for this session
        metrics = StudentExamMetrics.objects.filter(
            student=student,
            exam_session=session
        ).first()
        
        # Get student's position
        position = StudentExamPosition.objects.filter(
            student=student,
            exam_session=session
        ).first()
        
        # Get subject results
        subject_results = StudentSubjectResult.objects.filter(
            student=student,
            exam_session=session
        ).select_related('subject').order_by('subject__name')
        
        # Get all papers for this session
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        # Get paper scores
        paper_scores = {
            ps.exam_paper_id: ps
            for ps in StudentPaperScore.objects.filter(
                student=student,
                exam_paper__exam_session=session
            ).select_related('exam_paper')
        }
        
        # Get grading scale for this educational level
        grading_scale = GradingScale.objects.filter(
            education_level=educational_level
        ).order_by('-min_mark')
        
        # Get division scale if applicable (O-Level/A-Level only)
        division_scale = None
        if is_secondary:
            division_scale = DivisionScale.objects.filter(
                education_level=educational_level
            ).order_by('min_points')
        
        # Calculate class statistics for comparison
        class_metrics = StudentExamMetrics.objects.filter(
            exam_session=session
        ).aggregate(
            class_avg_total=Avg('total_marks'),
            class_avg_average=Avg('average_marks'),
            class_highest=Max('total_marks'),
            class_lowest=Min('total_marks'),
        )
        
        # Calculate subject-wise statistics including positions
        subject_class_stats = {}
        all_subject_results = StudentSubjectResult.objects.filter(
            exam_session=session
        ).select_related('subject')
        
        # First pass: collect all subject results and calculate totals
        subject_results_by_student = {}
        for result in all_subject_results:
            subj_id = result.subject_id
            student_id = result.student_id
            
            if subj_id not in subject_results_by_student:
                subject_results_by_student[subj_id] = []
            subject_results_by_student[subj_id].append({
                'student_id': student_id,
                'total_marks': float(result.total_marks),
                'grade': result.grade,
            })
        
        # Calculate class statistics for each subject
        for subj_id, results_list in subject_results_by_student.items():
            # Sort by total marks descending to determine positions
            sorted_results = sorted(results_list, key=lambda x: x['total_marks'], reverse=True)
            
            # Calculate statistics
            total_marks_sum = sum(r['total_marks'] for r in results_list)
            count = len(results_list)
            avg_marks = total_marks_sum / count if count > 0 else 0
            
            # Calculate grade distribution
            grade_distribution = {}
            for r in results_list:
                grade = r['grade']
                if grade:
                    grade_distribution[grade] = grade_distribution.get(grade, 0) + 1
            
            # Build position map
            position_map = {}
            for idx, r in enumerate(sorted_results, 1):
                position_map[r['student_id']] = idx
            
            subject_class_stats[subj_id] = {
                'avg_marks': avg_marks,
                'count': count,
                'total_marks': total_marks_sum,
                'grade_distribution': grade_distribution,
                'sorted_results': sorted_results,
                'position_map': position_map,
            }
        
        # Build subject-wise paper data with enhanced metrics
        subjects_data = []
        for paper in papers:
            subject = paper.subject
            existing_subject = next((s for s in subjects_data if s['subject'].pk == subject.pk), None)
            
            if not existing_subject:
                subject_result = subject_results.filter(subject=subject).first()
                class_stats = subject_class_stats.get(subject.pk, {
                    'avg_marks': 0, 
                    'grade_distribution': {},
                    'position_map': {},
                    'count': 0
                })
                
                # Calculate student's percentage and grade for this subject
                subject_max_marks = self._get_subject_max_marks(papers, subject.pk)
                
                if subject_result:
                    student_percentage = (float(subject_result.total_marks) / subject_max_marks * 100) if subject_max_marks > 0 else 0
                    student_grade = subject_result.grade
                    student_position = class_stats.get('position_map', {}).get(student.pk)
                else:
                    student_percentage = None
                    student_grade = None
                    student_position = None
                
                subjects_data.append({
                    'subject': subject,
                    'papers': [],
                    'result': subject_result,
                    'total_marks': subject_result.total_marks if subject_result else None,
                    'grade': subject_result.grade if subject_result else None,
                    'points': subject_result.points if subject_result else None,
                    'percentage': round(student_percentage, 2) if student_percentage else None,
                    'position': student_position,
                    'total_students': class_stats.get('count', 0),
                    'class_avg_marks': round(class_stats['avg_marks'], 2) if class_stats['avg_marks'] else 0,
                    'class_avg_percentage': round((class_stats['avg_marks'] / subject_max_marks * 100), 2) if class_stats['avg_marks'] and subject_max_marks > 0 else 0,
                    'class_grade_distribution': class_stats.get('grade_distribution', {}),
                    'performance_vs_class': None,
                })
                
                # Calculate performance vs class average
                if subject_result and class_stats['avg_marks'] > 0:
                    diff = float(subject_result.total_marks) - class_stats['avg_marks']
                    subjects_data[-1]['performance_vs_class'] = {
                        'difference': round(diff, 2),
                        'percentage_diff': round((diff / class_stats['avg_marks'] * 100), 1) if class_stats['avg_marks'] > 0 else 0,
                        'is_above': diff > 0,
                    }
            
            # Add paper data
            current_subject = next(s for s in subjects_data if s['subject'].pk == subject.pk)
            score = paper_scores.get(paper.pk)
            percentage = (score.marks / paper.max_marks * 100) if score and paper.max_marks > 0 else 0
            
            current_subject['papers'].append({
                'paper': paper,
                'score': score,
                'marks': score.marks if score else None,
                'percentage': round(percentage, 2) if score else None,
                'max_marks': paper.max_marks,
                'has_score': score is not None,
            })
        
        # Sort subjects by name
        subjects_data.sort(key=lambda x: x['subject'].name)
        
        # Calculate performance metrics based on educational level
        performance_metrics = {
            'total_marks': metrics.total_marks if metrics else None,
            'average_marks': metrics.average_marks if metrics else None,
            'class_position': position.class_position if position else None,
            'stream_position': position.stream_position if position else None,
            'class_avg': class_metrics['class_avg_total'],
            'class_avg_percentage': class_metrics['class_avg_average'],
            'class_highest': class_metrics['class_highest'],
            'class_lowest': class_metrics['class_lowest'],
        }
        
        # Add level-specific metrics
        if is_secondary:
            performance_metrics['total_points'] = metrics.total_points if metrics else None
            performance_metrics['division'] = metrics.division if metrics else None
        else:
            # For Primary/Nursery, calculate overall grade from subject averages
            overall_grade = None
            if metrics and metrics.average_marks:
                for gs in grading_scale:
                    if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                        overall_grade = gs.grade
                        break
            performance_metrics['overall_grade'] = overall_grade
        
        # Calculate performance vs class average
        if metrics and class_metrics['class_avg_total']:
            diff = metrics.total_marks - class_metrics['class_avg_total']
            performance_metrics['vs_class_avg'] = {
                'difference': diff,
                'percentage_diff': (diff / class_metrics['class_avg_total'] * 100) if class_metrics['class_avg_total'] > 0 else 0,
                'is_above': diff > 0,
            }
        
        # Grade distribution for this student
        grade_counts = {}
        for data in subjects_data:
            grade = data['grade']
            if grade:
                grade_counts[grade] = grade_counts.get(grade, 0) + 1
        
        # Calculate subject strengths and weaknesses
        subject_strengths = []
        subject_weaknesses = []
        
        for data in subjects_data:
            if data['result'] and data['total_marks'] and data['class_avg_marks'] > 0:
                diff = float(data['total_marks']) - data['class_avg_marks']
                subject_info = {
                    'subject': data['subject'],
                    'marks': data['total_marks'],
                    'percentage': data['percentage'],
                    'grade': data['grade'],
                    'position': data['position'],
                    'total_students': data['total_students'],
                    'class_avg': data['class_avg_marks'],
                    'class_avg_percentage': data['class_avg_percentage'],
                    'difference': diff,
                    'percentage_diff': (diff / data['class_avg_marks'] * 100) if data['class_avg_marks'] > 0 else 0,
                }
                if is_secondary and data['points']:
                    subject_info['points'] = data['points']
                
                if diff > 0:
                    subject_strengths.append(subject_info)
                else:
                    subject_weaknesses.append(subject_info)
        
        # Sort strengths and weaknesses
        subject_strengths.sort(key=lambda x: x['difference'], reverse=True)
        subject_weaknesses.sort(key=lambda x: x['difference'])
        
        # Get available sessions for navigation
        available_sessions = ExamSession.objects.filter(
            academic_year__in=student.enrollments.values_list('academic_year', flat=True),
            class_level__in=student.enrollments.values_list('class_level', flat=True)
        ).exclude(pk=session_pk).order_by('-exam_date')[:5]
        
        context = {
            'student': student,
            'session': session,
            'enrollment': enrollment,
            'educational_level': educational_level,
            'level_type': level_type,
            'is_primary_nursery': is_primary_nursery,
            'is_secondary': is_secondary,
            'metrics': metrics,
            'position': position,
            'subjects_data': subjects_data,
            'performance_metrics': performance_metrics,
            'grade_counts': grade_counts,
            'grading_scale': grading_scale,
            'division_scale': division_scale,
            'subject_strengths': subject_strengths[:5],
            'subject_weaknesses': subject_weaknesses[:5],
            'available_sessions': available_sessions,
            'total_subjects': len(subjects_data),
            'subjects_with_results': sum(1 for d in subjects_data if d['result']),
            'generated_date': timezone.now(),
        }
        
        return render(request, self.template_name, context)
    
    def _get_subject_max_marks(self, papers, subject_id):
        """Calculate total maximum marks for a subject across all papers."""
        subject_papers = [p for p in papers if p.subject_id == subject_id]
        return sum(float(p.max_marks) for p in subject_papers)


# views/export_student_result_excel_view.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import date

from core.mixins import ManagementRequiredMixin
from core.models import (
    Student, ExamSession, StudentSubjectResult, SubjectExamPaper,
    StudentPaperScore, StudentExamMetrics, StudentExamPosition, GradingScale,
)


# ══════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=10, color='000000', italic=False, name='Arial'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════════════════════════
#  PALETTE
# ══════════════════════════════════════════════════════════════════════════════

C_BRAND_DARK  = '0D3349'
C_BRAND_MID   = '1A5276'
C_ACCENT      = '148F77'
C_HEADER_BG   = '1B4F72'
C_SUBHDR_BG   = '2980B9'
C_STATS_BG    = '1A5276'
C_STATS_ROW   = 'D6EAF8'
C_WHITE       = 'FFFFFF'
C_GOLD        = 'F0B429'

# Grade badge colours — best → worst (cycling palette identical to analytics view)
_BADGE_PALETTE  = [
    '1E7E34', '1565C0', 'E65100', '6A1E6E',
    'B71C1C', '004D40', '37474F', '880E4F', '4A148C', '3E2723',
]
_ROW_BG_PALETTE = [
    'E8F5E9', 'E3F2FD', 'FFF3E0', 'F3E5F5',
    'FFEBEE', 'E0F2F1', 'ECEFF1', 'FCE4EC', 'EDE7F6', 'EFEBE9',
]

# Division badge map
_DIV_BADGE = {
    'I':   ('1E7E34', 'E8F5E9'),
    'II':  ('1565C0', 'E3F2FD'),
    'III': ('E65100', 'FFF3E0'),
    'IV':  ('880E4F', 'FCE4EC'),
    '0':   ('B71C1C', 'FFEBEE'),
}
_DIV_BADGE_DEFAULT = ('37474F', 'ECEFF1')

# Paper-score performance bands
_PERF_BANDS = [
    (75, '1B5E20', 'E8F5E9'),   # ≥75 %  → deep green / light green
    (50, '1A237E', 'E3F2FD'),   # ≥50 %  → deep blue  / light blue
    (30, 'E65100', 'FFF3E0'),   # ≥30 %  → deep orange/ light orange
    (0,  'B71C1C', 'FFEBEE'),   # < 30 % → deep red   / light red
]


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD GRADE-STYLE MAPS  (derived entirely from GradingScale queryset)
# ══════════════════════════════════════════════════════════════════════════════

def _build_grade_style_maps(grading_scale_list):
    """
    Returns (badge_fill, text_color, row_bg) dicts keyed by grade value.
    grading_scale_list must be ordered -min_mark (best grade first).
    """
    badge_fill = {}
    text_color = {}
    row_bg     = {}
    for idx, gs in enumerate(grading_scale_list):
        badge_fill[gs.grade] = _fill(_BADGE_PALETTE[idx % len(_BADGE_PALETTE)])
        text_color[gs.grade] = C_WHITE
        row_bg[gs.grade]     = _ROW_BG_PALETTE[idx % len(_ROW_BG_PALETTE)]
    return badge_fill, text_color, row_bg


# ══════════════════════════════════════════════════════════════════════════════
#  SHARED SHEET UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _school_block(ws, session, report_title, report_subtitle,
                  last_col, school_name, school_address,
                  school_phone, school_email, school_motto,
                  school_reg_no, generated_by):
    """
    Writes the standard 11-row premium school header used across every sheet.
    Returns the next free row number (12) where callers continue.
    """
    L = get_column_letter(last_col)

    def band(row, value, fill_hex, fnt, aln=None, h=None):
        ws.merge_cells(f'A{row}:{L}{row}')
        c = ws.cell(row=row, column=1, value=value)
        c.fill      = _fill(fill_hex)
        c.font      = fnt
        c.alignment = aln or _align('center', 'center')
        if h:
            ws.row_dimensions[row].height = h
        return c

    band(1, '', C_BRAND_DARK, _font(), h=8)

    band(2, school_name.upper(), C_BRAND_DARK,
         _font(bold=True, size=18, color=C_GOLD),
         _align('center', 'center'), h=38)

    contact = '   |   '.join(p for p in [school_address, school_phone, school_email] if p)
    band(3, contact, C_BRAND_MID,
         _font(size=9, color='D6EAF8', italic=True),
         _align('center', 'center'), h=18)

    motto_line = ''
    if school_motto:  motto_line += f'"{school_motto}"'
    if school_reg_no: motto_line += f'   Reg No: {school_reg_no}'
    band(4, motto_line, C_BRAND_MID,
         _font(size=9, color='AED6F1', italic=bool(school_motto)),
         _align('center', 'center'), h=18)

    band(5, '', C_GOLD, _font(), h=4)

    band(6, report_title, C_ACCENT,
         _font(bold=True, size=13, color=C_WHITE),
         _align('center', 'center'), h=28)

    band(7, report_subtitle, '0E6655',
         _font(bold=True, size=11, color='A9DFBF'),
         _align('center', 'center'), h=22)

    stream_part = f'   │   Stream: {session.stream_class.name}' if session.stream_class else ''
    band(8,
         f'Session: {session.name}   │   Class: {session.class_level.name}'
         f'   │   Term: {session.term}   │   Year: {session.academic_year}{stream_part}',
         'D6EAF8', _font(bold=True, size=9, color='1A5276'),
         _align('center', 'center'), h=20)

    band(9,
         f'Education Level: {session.class_level.educational_level}'
         f'   │   Exam Date: {session.exam_date.strftime("%d %B %Y")}',
         'EBF5FB', _font(size=9, color='1A5276'),
         _align('center', 'center'), h=18)

    band(10,
         f'Generated: {timezone.now().strftime("%A, %d %B %Y  %H:%M")}   │   By: {generated_by}',
         'F2F3F4', _font(size=8, color='555555', italic=True),
         _align('center', 'center'), h=16)

    band(11, '', 'D5D8DC', _font(), h=4)

    return 12   # next free row


def _col_headers(ws, row, headers, last_col):
    ws.row_dimensions[row].height = 28
    for col, h in enumerate(headers, 1):
        c           = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align('center', 'center')
        c.border    = _thin_border(C_BRAND_DARK)


def _footer(ws, row, last_col, school_name):
    L = get_column_letter(last_col)
    ws.merge_cells(f'A{row}:{L}{row}')
    c = ws.cell(row=row, column=1,
                value=(f'This report is computer-generated and is valid without a signature.  '
                       f'© {date.today().year} {school_name}  │  '
                       f'Printed: {timezone.now().strftime("%d %b %Y %H:%M")}'))
    c.font      = _font(size=8, color='888888', italic=True)
    c.fill      = _fill('F2F3F4')
    c.alignment = _align('center', 'center')
    ws.row_dimensions[row].height = 16

    ws.merge_cells(f'A{row+1}:{L}{row+1}')
    ws.cell(row=row + 1, column=1).fill = _fill(C_GOLD)
    ws.row_dimensions[row + 1].height  = 5


def _print_setup(ws, freeze_row, last_col):
    ws.freeze_panes                           = f'A{freeze_row}'
    ws.print_title_rows                       = f'1:{freeze_row - 1}'
    ws.page_setup.orientation                 = 'landscape'
    ws.page_setup.paperSize                   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage                   = True
    ws.page_setup.fitToWidth                  = 1
    ws.page_setup.fitToHeight                 = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _perf_colours(pct):
    """Return (font_hex, bg_hex) for a percentage value using band thresholds."""
    for threshold, fc, bg in _PERF_BANDS:
        if pct >= threshold:
            return fc, bg
    return _PERF_BANDS[-1][1], _PERF_BANDS[-1][2]


# ══════════════════════════════════════════════════════════════════════════════
#  GRADING-SCALE REFERENCE TABLE  (appended at the bottom of multiple sheets)
# ══════════════════════════════════════════════════════════════════════════════

def _grading_ref_table(ws, start_row, grading_scale_list,
                        badge_fill, text_color, last_col, ed_level):
    L = get_column_letter(last_col)
    ws.merge_cells(f'A{start_row}:{L}{start_row}')
    c = ws.cell(row=start_row, column=1,
                value=f'GRADING SCALE REFERENCE  –  {ed_level}')
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_BRAND_MID)
    c.alignment = _align('center', 'center')
    ws.row_dimensions[start_row].height = 22

    hdr = start_row + 1
    for ci, lbl in enumerate(['GRADE', 'MIN %', 'MAX %', 'POINTS', 'DESCRIPTION'], 1):
        c           = ws.cell(row=hdr, column=ci, value=lbl)
        c.font      = _font(bold=True, size=9, color=C_WHITE)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align('center', 'center')
        c.border    = _thin_border()
    ws.row_dimensions[hdr].height = 18

    for ri, gs in enumerate(grading_scale_list, hdr + 1):
        for ci, val in enumerate(
            [gs.grade, f'{gs.min_mark:.0f}', f'{gs.max_mark:.0f}',
             f'{gs.points:.1f}', gs.description or ''], 1
        ):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = badge_fill.get(gs.grade, _fill('EEEEEE')) if ci == 1 else _fill('F7F9FC')
            c.font      = _font(bold=(ci == 1), size=9,
                                color=text_color.get(gs.grade, '000000') if ci == 1 else '1A252F')
            c.alignment = _align('center', 'center')
            c.border    = _thin_border()
        ws.row_dimensions[ri].height = 16

    return hdr + len(grading_scale_list) + 1


# ══════════════════════════════════════════════════════════════════════════════
#  VIEW
# ══════════════════════════════════════════════════════════════════════════════

class ExportStudentResultExcelView(ManagementRequiredMixin, View):
    """
    Export a single student's exam results to a richly-formatted Excel workbook.

    Sheets
    ──────
    1. Result Card        — student bio + exam info + performance summary
    2. Subject Results    — per-subject marks, grade badge, class avg, position
    3. Paper Breakdown    — individual paper scores with performance colours
    4. Performance Analysis — grade distribution + above/below average analysis
    """

    def get(self, request, student_pk, session_pk):
        student = get_object_or_404(Student, pk=student_pk)
        session = get_object_or_404(ExamSession, pk=session_pk)

        educational_level = session.class_level.educational_level
        level_type        = educational_level.level_type
        is_secondary      = level_type in ['O_LEVEL', 'A_LEVEL']
        is_alevel         = level_type == 'A_LEVEL'

        # ── School metadata ───────────────────────────────────────────────
        school_meta = {
            'school_name': getattr(settings, 'SCHOOL_NAME', 'SCHOOL NAME NOT SET'),
            'school_address': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'school_phone': getattr(settings, 'SCHOOL_PHONE', ''),
            'school_email': getattr(settings, 'SCHOOL_EMAIL', ''),
            'school_motto': getattr(settings, 'SCHOOL_MOTTO', ''),
            'school_reg_no': getattr(settings, 'SCHOOL_REGISTRATION_NUMBER', ''),
            'generated_by': request.user.get_full_name() or request.user.username,
        }      
                # ── Get student's combination for A-Level ─────────────────────────
        student_combination = None
        if is_alevel:
            enrollment = student.enrollments.filter(
                academic_year=session.academic_year,
                class_level=session.class_level
            ).first()
            if enrollment:
                combination_assignment = enrollment.combination_assignments.filter(is_active=True).first()
                if combination_assignment:
                    student_combination = combination_assignment.combination

        # ── Grading scale ─────────────────────────────────────────────────
        grading_scale_list = list(
            GradingScale.objects.filter(education_level=educational_level)
            .order_by('-min_mark')
        )
        badge_fill, text_color, row_bg_map = _build_grade_style_maps(grading_scale_list)
        grade_points = {gs.grade: float(gs.points) for gs in grading_scale_list}

        def _grade_from_pct(pct):
            for gs in grading_scale_list:
                if float(gs.min_mark) <= pct <= float(gs.max_mark):
                    return gs.grade
            return grading_scale_list[-1].grade if grading_scale_list else '?'

        # ── DB queries ────────────────────────────────────────────────────
        subject_results = list(
            StudentSubjectResult.objects.filter(
                student=student, exam_session=session
            ).select_related('subject').order_by('subject__name')
        )

        papers_qs = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        papers_list = list(papers_qs)

        paper_scores = {
            ps.exam_paper_id: ps
            for ps in StudentPaperScore.objects.filter(
                student=student, exam_paper__exam_session=session
            ).select_related('exam_paper')
        }

        metrics  = StudentExamMetrics.objects.filter(
            student=student, exam_session=session
        ).first()

        position = StudentExamPosition.objects.filter(
            student=student, exam_session=session
        ).first()

        # ── Class averages per subject ─────────────────────────────────────
        all_subject_results = list(
            StudentSubjectResult.objects.filter(
                exam_session=session
            ).select_related('subject')
        )

        class_stats = {}   # subject_id → {total, count, scores[]}
        for r in all_subject_results:
            sid = r.subject_id
            class_stats.setdefault(sid, {'total': 0.0, 'count': 0, 'scores': []})
            class_stats[sid]['total']  += float(r.total_marks)
            class_stats[sid]['count']  += 1
            class_stats[sid]['scores'].append(float(r.total_marks))
        for sid in class_stats:
            cs = class_stats[sid]
            cs['avg'] = cs['total'] / cs['count'] if cs['count'] else 0.0

        # Max marks per subject (sum of paper max_marks)
        subject_max = {}    # subject_id → float
        for p in papers_list:
            subject_max[p.subject_id] = subject_max.get(p.subject_id, 0.0) + float(p.max_marks)

        # ── Workbook ──────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        wb.remove(wb.active)    # we name all sheets ourselves

        # ─────────────────────────────────────────────────────────────────
        #  SHEET 1 — RESULT CARD
        # ─────────────────────────────────────────────────────────────────
        self._sheet_result_card(
            wb, student, session, metrics, position,
            educational_level, level_type, is_secondary,
            grading_scale_list, badge_fill, text_color, row_bg_map,
            student_combination,  is_alevel,
            grade_points, school_meta,
        )

        # ─────────────────────────────────────────────────────────────────
        #  SHEET 2 — SUBJECT RESULTS
        # ─────────────────────────────────────────────────────────────────
        self._sheet_subject_results(
            wb, student, session, subject_results, papers_list,
            class_stats, subject_max, educational_level, level_type,
            is_secondary, grading_scale_list, badge_fill, text_color,
            row_bg_map, school_meta,
        )

        # ─────────────────────────────────────────────────────────────────
        #  SHEET 3 — PAPER BREAKDOWN
        # ─────────────────────────────────────────────────────────────────
        if papers_list:
            self._sheet_paper_breakdown(
                wb, student, session, papers_list, paper_scores,
                educational_level, grading_scale_list, badge_fill,
                text_color, school_meta,
            )

        # ─────────────────────────────────────────────────────────────────
        #  SHEET 4 — PERFORMANCE ANALYSIS
        # ─────────────────────────────────────────────────────────────────
        self._sheet_performance_analysis(
            wb, student, session, subject_results, papers_list,
            class_stats, subject_max, educational_level, level_type,
            grading_scale_list, badge_fill, text_color, row_bg_map,
            school_meta,
        )

        # ── HTTP response ─────────────────────────────────────────────────
        reg = (student.registration_number or 'STUDENT').replace('/', '-')
        safe_session = session.name.replace(' ', '_').replace('/', '-')
        filename = f'Result_{reg}_{safe_session}_{date.today()}.xlsx'
        filename = ''.join(c for c in filename if c.isalnum() or c in '._-')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 1 — RESULT CARD
    # ══════════════════════════════════════════════════════════════════════

    def _sheet_result_card(
        self, wb, student, session, metrics, position,
        educational_level, level_type, is_secondary,
        grading_scale_list, badge_fill, text_color, is_alevel, row_bg_map,student_combination,
        grade_points, school_meta,
    ):
        last_col = 4
        L = get_column_letter(last_col)
        ws = wb.create_sheet('Result Card')

        sub = f'{student.full_name.upper()}  —  {student.registration_number or "N/A"}'
        cur = _school_block(ws, session,
                            'INDIVIDUAL STUDENT RESULT CARD', sub,
                            last_col, **school_meta)

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        def _section_header(row, title, colour='1B4F72'):
            ws.merge_cells(f'A{row}:{L}{row}')
            c           = ws.cell(row=row, column=1, value=title)
            c.font      = _font(bold=True, size=10, color=C_WHITE)
            c.fill      = _fill(colour)
            c.alignment = _align('center', 'center')
            ws.row_dimensions[row].height = 22

        def _kv_row(row, label, value, val_bold=False,
                    val_fill=None, val_text=None):
            bg = C_STATS_ROW if row % 2 == 0 else 'EBF5FB'
            lc           = ws.cell(row=row, column=1, value=label)
            lc.font      = _font(bold=True, size=9, color='1A252F')
            lc.fill      = _fill(bg)
            lc.alignment = _align('left', 'center')
            lc.border    = _thin_border()

            ws.merge_cells(f'B{row}:{L}{row}')
            vc           = ws.cell(row=row, column=2, value=value)
            vc.font      = _font(bold=val_bold, size=9,
                                 color=val_text or '1A252F')
            vc.fill      = _fill(val_fill) if val_fill else _fill(bg)
            vc.alignment = _align('left', 'center')
            vc.border    = _thin_border()
            ws.row_dimensions[row].height = 20

        # ── Student Information ───────────────────────────────────────────
        _section_header(cur, 'STUDENT INFORMATION', C_STATS_BG)
        cur += 1

        dob = student.date_of_birth.strftime('%d %B %Y') if student.date_of_birth else 'N/A'
        for label, value in [
            ('Full Name',            student.full_name),
            ('Registration Number',  student.registration_number or 'N/A'),
            ('Gender',               student.get_gender_display() or 'N/A'),
            ('Date of Birth',        dob),
            ('Examination Number',   student.examination_number or 'N/A'),
        ]:
            _kv_row(cur, label, value)
            cur += 1

            # Add combination for A-Level students
        if is_alevel and student_combination:
            _kv_row(cur, 'Combination', student_combination.code, val_bold=True, val_text='1A5276')
            cur += 1

        cur += 1

        # ── Exam Information ──────────────────────────────────────────────
        _section_header(cur, 'EXAM INFORMATION', C_BRAND_MID)
        cur += 1

        for label, value in [
            ('Exam Session',  session.name),
            ('Exam Type',     session.exam_type.name),
            ('Exam Date',     session.exam_date.strftime('%d %B %Y')),
            ('Class Level',   session.class_level.name),
            ('Term',          str(session.term)),
            ('Academic Year', str(session.academic_year)),
            ('Education Level', str(educational_level)),
        ]:
            _kv_row(cur, label, value)
            cur += 1

        cur += 1

        # ── Performance Summary ───────────────────────────────────────────
        _section_header(cur, 'PERFORMANCE SUMMARY', C_ACCENT)
        cur += 1

        has_metrics = metrics is not None

        _kv_row(cur, 'Total Marks',
                f'{float(metrics.total_marks):.1f}' if has_metrics else 'N/A',
                val_bold=True, val_text='1A5276')
        cur += 1

        _kv_row(cur, 'Average Score',
                f'{float(metrics.average_marks):.1f}%' if has_metrics else 'N/A',
                val_bold=True, val_text='1A5276')
        cur += 1

        _kv_row(cur, 'Class Position',
                f'#{position.class_position}' if position and position.class_position else 'N/A',
                val_bold=True)
        cur += 1

        if position and position.stream_position:
            _kv_row(cur, 'Stream Position', f'#{position.stream_position}')
            cur += 1

        if is_secondary and has_metrics:
            div = metrics.division or 'N/A'
            div_badge, _ = _DIV_BADGE.get(str(div), _DIV_BADGE_DEFAULT)
            _kv_row(cur, 'Division', div,
                    val_bold=True, val_fill=div_badge, val_text=C_WHITE)
            cur += 1

            pts = f'{float(metrics.total_points):.1f}' if metrics.total_points else 'N/A'
            _kv_row(cur, 'Total Points', pts, val_bold=True)
            cur += 1

        else:
            # Primary / Nursery — derive overall grade from average marks
            overall_grade = 'N/A'
            if has_metrics and metrics.average_marks:
                for gs in grading_scale_list:
                    if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                        overall_grade = gs.grade
                        break
            g_fill = badge_fill.get(overall_grade)
            g_text = text_color.get(overall_grade, '000000')
            _kv_row(cur, 'Overall Grade', overall_grade,
                    val_bold=True,
                    val_fill=g_fill.fgColor.rgb if g_fill else None,
                    val_text=g_text)
            cur += 1

        cur += 2

        # ── Grading scale reference ───────────────────────────────────────
        ref_end = _grading_ref_table(ws, cur, grading_scale_list,
                                     badge_fill, text_color,
                                     last_col, educational_level)
        _footer(ws, ref_end + 1, last_col, school_meta['school_name'])
        _print_setup(ws, 12, last_col)

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 2 — SUBJECT RESULTS
    # ══════════════════════════════════════════════════════════════════════

    def _sheet_subject_results(
        self, wb, student, session, subject_results, papers_list,
        class_stats, subject_max, educational_level, level_type,
        is_secondary, grading_scale_list, badge_fill, text_color,
        row_bg_map, school_meta,
    ):
        last_col = 8 if is_secondary else 7
        ws = wb.create_sheet('Subject Results')

        sub = f'{student.full_name}  —  {session.name}'
        cur = _school_block(ws, session,
                            'SUBJECT RESULTS', sub,
                            last_col, **school_meta)

        # Quick-stats bar
        ws.row_dimensions[cur].height = 22
        n_subjects = len(subject_results)
        qs = [
            ('SUBJECTS', str(n_subjects)),
            ('SCORED',   str(sum(1 for r in subject_results if r.total_marks))),
        ]
        if is_secondary:
            grades_list = [r.grade for r in subject_results if r.grade]
            points_total = sum(float(r.points or 0) for r in subject_results)
            qs.append(('TOTAL POINTS', f'{points_total:.1f}'))
        chunk = max(1, last_col // len(qs))
        L = get_column_letter(last_col)
        for i, (lbl, val) in enumerate(qs):
            cs = i * chunk + 1
            ce = min((cs + chunk - 1) if i < len(qs) - 1 else last_col, last_col)
            if cs < ce:
                ws.merge_cells(start_row=cur, start_column=cs,
                               end_row=cur,   end_column=ce)
            c           = ws.cell(row=cur, column=cs, value=f'{lbl}: {val}')
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_SUBHDR_BG if i % 2 == 0 else C_ACCENT)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border(C_BRAND_DARK)
        cur += 1

        headers = ['SUBJECT', 'TOTAL MARKS', 'MAX MARKS', 'PERCENTAGE', 'GRADE']
        if is_secondary:
            headers.append('POINTS')
        headers += ['CLASS AVG (%)', 'CLASS POSITION']
        _col_headers(ws, cur, headers, last_col)
        data_start = cur + 1

        for ri, sr in enumerate(subject_results, data_start):
            ws.row_dimensions[ri].height = 22
            s_max = subject_max.get(sr.subject_id, 0.0)
            pct   = (float(sr.total_marks) / s_max * 100) if s_max else 0.0
            cs    = class_stats.get(sr.subject_id, {})
            c_avg = cs.get('avg', 0.0)
            c_pct = (c_avg / s_max * 100) if s_max and c_avg else 0.0
            scores = cs.get('scores', [])
            c_pos  = sum(1 for sc in scores if sc > float(sr.total_marks)) + 1
            n_cls  = cs.get('count', 0)

            grade   = sr.grade or ''
            row_bg  = row_bg_map.get(grade, 'FAFCFE')
            fc, bg  = _perf_colours(pct)

            row_data   = [sr.subject.name, float(sr.total_marks), s_max, pct, grade]
            col_aligns = ['left', 'center', 'center', 'center', 'center']

            if is_secondary:
                row_data.append(float(sr.points) if sr.points else 0.0)
                col_aligns.append('center')

            row_data  += [c_pct, f'{c_pos}/{n_cls}' if n_cls else 'N/A']
            col_aligns += ['center', 'center']

            for col, (val, aln) in enumerate(zip(row_data, col_aligns), 1):
                c = ws.cell(row=ri, column=col)

                if col == 5:   # GRADE badge
                    c.value     = val
                    c.font      = _font(bold=True, size=10,
                                        color=text_color.get(grade, C_WHITE))
                    c.fill      = badge_fill.get(grade, _fill('888888'))
                    c.alignment = _align('center', 'center')

                elif col == 4:  # PERCENTAGE — colour by performance band
                    c.value         = val
                    c.number_format = '0.0"%"'
                    c.font          = _font(bold=True, size=9, color=fc)
                    c.fill          = _fill(bg)
                    c.alignment     = _align('center', 'center')

                elif col == (6 if is_secondary else 5) + 1:  # CLASS AVG
                    c.value         = val
                    c.number_format = '0.0"%"'
                    c.font          = _font(size=9, color='1A5276')
                    c.fill          = _fill(row_bg)
                    c.alignment     = _align('center', 'center')

                elif col in (2, 3):  # MARKS columns
                    c.value         = val
                    c.number_format = '0.0'
                    c.font          = _font(bold=(col == 2), size=9, color='1A5276')
                    c.fill          = _fill(row_bg)
                    c.alignment     = _align('center', 'center')

                else:
                    c.value     = val
                    c.font      = _font(size=9)
                    c.fill      = _fill(row_bg)
                    c.alignment = _align(aln, 'center')

                c.border = _thin_border()

        last_data = data_start + len(subject_results) - 1

        # Grading scale reference
        ref_end = _grading_ref_table(
            ws, last_data + 3, grading_scale_list,
            badge_fill, text_color, last_col, educational_level,
        )
        _footer(ws, ref_end + 1, last_col, school_meta['school_name'])

        ws.column_dimensions['A'].width = 26
        for i in range(1, last_col):
            ws.column_dimensions[get_column_letter(i + 1)].width = 14

        _print_setup(ws, data_start, last_col)

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 3 — PAPER BREAKDOWN
    # ══════════════════════════════════════════════════════════════════════

    def _sheet_paper_breakdown(
        self, wb, student, session, papers_list, paper_scores,
        educational_level, grading_scale_list, badge_fill,
        text_color, school_meta,
    ):
        last_col = 7
        ws = wb.create_sheet('Paper Breakdown')

        sub = f'{student.full_name}  —  {session.name}'
        cur = _school_block(ws, session,
                            'PAPER BREAKDOWN', sub,
                            last_col, **school_meta)

        headers = ['SUBJECT', 'PAPER', 'PAPER NAME', 'MAX MARKS', 'SCORE', 'PERCENTAGE', 'GRADE']
        _col_headers(ws, cur, headers, last_col)
        data_start = cur + 1

        for ri, paper in enumerate(papers_list, data_start):
            ws.row_dimensions[ri].height = 20
            score = paper_scores.get(paper.pk)
            p_max = float(paper.max_marks)

            if score:
                marks   = float(score.marks)
                pct     = (marks / p_max * 100) if p_max else 0.0
                grade   = self._grade_from_pct(pct, grading_scale_list)
                fc, bg  = _perf_colours(pct)
            else:
                marks   = None
                pct     = None
                grade   = None
                fc, bg  = '888888', 'F5F5F5'

            row_data = [
                paper.subject.name,
                f'Paper {paper.paper_number}',
                paper.paper_name or '—',
                p_max,
                marks,
                pct,
                grade,
            ]

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=col)

                if col == 7:    # GRADE badge
                    if grade:
                        c.value     = grade
                        c.font      = _font(bold=True, size=10,
                                            color=text_color.get(grade, C_WHITE))
                        c.fill      = badge_fill.get(grade, _fill('888888'))
                    else:
                        c.value = '—'
                        c.font  = _font(size=9, color='AAAAAA', italic=True)
                        c.fill  = _fill('F5F5F5')
                    c.alignment = _align('center', 'center')

                elif col == 6:  # PERCENTAGE — performance band colour
                    if pct is not None:
                        c.value         = pct
                        c.number_format = '0.0"%"'
                        c.font          = _font(bold=True, size=9, color=fc)
                        c.fill          = _fill(bg)
                    else:
                        c.value = 'Not scored'
                        c.font  = _font(size=9, color='AAAAAA', italic=True)
                        c.fill  = _fill('F5F5F5')
                    c.alignment = _align('center', 'center')

                elif col == 5:  # SCORE
                    if marks is not None:
                        c.value         = marks
                        c.number_format = '0.0'
                        c.font          = _font(bold=True, size=10, color='1A5276')
                        c.fill          = _fill(bg)
                    else:
                        c.value = '—'
                        c.font  = _font(size=9, color='AAAAAA', italic=True)
                        c.fill  = _fill('F5F5F5')
                    c.alignment = _align('center', 'center')

                elif col == 4:  # MAX MARKS
                    c.value         = val
                    c.number_format = '0.0'
                    c.font          = _font(size=9, color='555555')
                    c.fill          = _fill('F7F9FC')
                    c.alignment     = _align('center', 'center')

                elif col == 1:  # SUBJECT
                    c.value     = val
                    c.font      = _font(bold=True, size=9)
                    c.fill      = _fill('EBF5FB')
                    c.alignment = _align('left', 'center')

                else:
                    c.value     = val
                    c.font      = _font(size=9)
                    c.fill      = _fill('F7F9FC')
                    c.alignment = _align('center', 'center')

                c.border = _thin_border()

        last_data = data_start + len(papers_list) - 1

        ref_end = _grading_ref_table(
            ws, last_data + 3, grading_scale_list,
            badge_fill, text_color, last_col, educational_level,
        )
        _footer(ws, ref_end + 1, last_col, school_meta['school_name'])

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10

        _print_setup(ws, data_start, last_col)

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 4 — PERFORMANCE ANALYSIS
    # ══════════════════════════════════════════════════════════════════════

    def _sheet_performance_analysis(
        self, wb, student, session, subject_results, papers_list,
        class_stats, subject_max, educational_level, level_type,
        grading_scale_list, badge_fill, text_color, row_bg_map,
        school_meta,
    ):
        last_col = 6
        L = get_column_letter(last_col)
        ws = wb.create_sheet('Performance Analysis')

        sub = f'{student.full_name}  —  {session.name}'
        cur = _school_block(ws, session,
                            'PERFORMANCE ANALYSIS', sub,
                            last_col, **school_meta)

        # ── A. Grade Distribution ─────────────────────────────────────────
        ws.merge_cells(f'A{cur}:{L}{cur}')
        c = ws.cell(row=cur, column=1, value='GRADE DISTRIBUTION')
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_STATS_BG)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[cur].height = 24
        cur += 1

        hdr_row = cur
        for ci, lbl in enumerate(['GRADE', 'COUNT', 'PERCENTAGE', 'BAR'], 1):
            c           = ws.cell(row=hdr_row, column=ci, value=lbl)
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_ACCENT)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border()
        ws.row_dimensions[hdr_row].height = 20
        cur += 1

        from collections import Counter
        grade_counts = Counter(r.grade for r in subject_results if r.grade)
        n_subjects   = len(subject_results) or 1

        for gs in grading_scale_list:
            g     = gs.grade
            cnt   = grade_counts.get(g, 0)
            pct_g = cnt / n_subjects * 100
            bar   = '█' * max(1, int(pct_g / 5)) if cnt else ''

            for ci, val in enumerate([g, cnt, f'{pct_g:.1f}%', bar], 1):
                c       = ws.cell(row=cur, column=ci, value=val)
                c.fill  = badge_fill.get(g, _fill('DDDDDD')) if ci == 1 else _fill('F7F9FC')
                c.font  = _font(bold=(ci == 1), size=9,
                                color=text_color.get(g, '000000') if ci == 1 else '1A252F')
                c.alignment = _align('center', 'center')
                c.border    = _thin_border()
            ws.row_dimensions[cur].height = 18
            cur += 1

        cur += 2

        # ── B. Subject vs Class Average ───────────────────────────────────
        ws.merge_cells(f'A{cur}:{L}{cur}')
        c = ws.cell(row=cur, column=1, value='SUBJECT PERFORMANCE vs CLASS AVERAGE')
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_BRAND_MID)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[cur].height = 24
        cur += 1

        for ci, lbl in enumerate(
            ['SUBJECT', 'YOUR SCORE (%)', 'CLASS AVG (%)', 'DIFFERENCE', 'STATUS', 'GRADE'], 1
        ):
            c           = ws.cell(row=cur, column=ci, value=lbl)
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_HEADER_BG)
            c.alignment = _align('center', 'center')
            c.border    = _thin_border()
        ws.row_dimensions[cur].height = 22
        cur += 1

        for sr in subject_results:
            s_max   = subject_max.get(sr.subject_id, 0.0)
            my_pct  = (float(sr.total_marks) / s_max * 100) if s_max else 0.0
            cs      = class_stats.get(sr.subject_id, {})
            c_avg   = cs.get('avg', 0.0)
            c_pct   = (c_avg / s_max * 100) if s_max and c_avg else 0.0
            diff    = my_pct - c_pct
            grade   = sr.grade or ''
            ws.row_dimensions[cur].height = 20

            row_bg = row_bg_map.get(grade, 'FAFCFE')

            if diff > 5:
                status, s_bg, s_fc = 'Above Average', '1E7E34', C_WHITE
            elif diff < -5:
                status, s_bg, s_fc = 'Below Average', 'B71C1C', C_WHITE
            else:
                status, s_bg, s_fc = 'On Average',    'E65100', C_WHITE

            row_vals = [
                sr.subject.name,
                my_pct,
                c_pct,
                diff,
                status,
                grade,
            ]

            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=cur, column=col)

                if col == 1:    # Subject name
                    c.value     = val
                    c.font      = _font(bold=True, size=9)
                    c.fill      = _fill(row_bg)
                    c.alignment = _align('left', 'center')

                elif col in (2, 3):  # Percentages
                    c.value         = val
                    c.number_format = '0.0"%"'
                    fc_p, bg_p      = _perf_colours(val)
                    c.font          = _font(bold=True, size=9, color=fc_p)
                    c.fill          = _fill(bg_p)
                    c.alignment     = _align('center', 'center')

                elif col == 4:  # Difference
                    c.value         = val
                    c.number_format = '+0.0;-0.0;0.0'
                    c.font          = _font(bold=True, size=9,
                                            color='1E7E34' if diff > 0 else 'B71C1C')
                    c.fill          = _fill(row_bg)
                    c.alignment     = _align('center', 'center')

                elif col == 5:  # Status badge
                    c.value     = val
                    c.font      = _font(bold=True, size=9, color=s_fc)
                    c.fill      = _fill(s_bg)
                    c.alignment = _align('center', 'center')

                elif col == 6:  # Grade badge
                    c.value     = grade
                    c.font      = _font(bold=True, size=10,
                                        color=text_color.get(grade, C_WHITE))
                    c.fill      = badge_fill.get(grade, _fill('888888'))
                    c.alignment = _align('center', 'center')

                c.border = _thin_border()

            cur += 1

        cur += 2

        # ── C. Summary stats block ────────────────────────────────────────
        ws.merge_cells(f'A{cur}:{L}{cur}')
        c = ws.cell(row=cur, column=1, value='COMPREHENSIVE SUMMARY')
        c.font      = _font(bold=True, size=11, color=C_WHITE)
        c.fill      = _fill(C_STATS_BG)
        c.alignment = _align('center', 'center')
        ws.row_dimensions[cur].height = 24
        cur += 1

        if subject_results:
            totals  = [float(r.total_marks) for r in subject_results]
            s_maxes = [subject_max.get(r.subject_id, 0) for r in subject_results]
            pcts    = [(t / m * 100) if m else 0 for t, m in zip(totals, s_maxes)]
            best_sr = max(subject_results, key=lambda r: float(r.total_marks))
            weak_sr = min(subject_results, key=lambda r: float(r.total_marks))
            best_s  = subject_max.get(best_sr.subject_id, 0)
            weak_s  = subject_max.get(weak_sr.subject_id, 0)
        else:
            pcts = []; best_sr = weak_sr = None; best_s = weak_s = 0

        half = last_col // 2
        summary_rows = [
            ('Subjects Attempted',  len(subject_results)),
            ('Average Score (%)',   f'{sum(pcts)/len(pcts):.1f}' if pcts else 'N/A'),
            ('Highest Subject Score',
             f'{best_sr.subject.name}: {float(best_sr.total_marks):.0f}/{best_s:.0f}' if best_sr else 'N/A'),
            ('Lowest Subject Score',
             f'{weak_sr.subject.name}: {float(weak_sr.total_marks):.0f}/{weak_s:.0f}' if weak_sr else 'N/A'),
            ('Subjects Above Class Avg', str(sum(
                1 for sr in subject_results
                if subject_max.get(sr.subject_id)
                and (float(sr.total_marks) / subject_max[sr.subject_id] * 100)
                   > (class_stats.get(sr.subject_id, {}).get('avg', 0)
                      / subject_max[sr.subject_id] * 100
                      if subject_max.get(sr.subject_id) else 0)
            ))),
        ]
        for label, val in summary_rows:
            bg = C_STATS_ROW if cur % 2 == 0 else 'EBF5FB'
            ws.merge_cells(f'A{cur}:{get_column_letter(half)}{cur}')
            lc           = ws.cell(row=cur, column=1, value=label)
            lc.font      = _font(size=9, color='1A252F')
            lc.fill      = _fill(bg)
            lc.alignment = _align('left', 'center')
            lc.border    = _thin_border()

            ws.merge_cells(f'{get_column_letter(half+1)}{cur}:{L}{cur}')
            vc           = ws.cell(row=cur, column=half + 1, value=val)
            vc.font      = _font(bold=True, size=9, color=C_BRAND_DARK)
            vc.fill      = _fill(bg)
            vc.alignment = _align('center', 'center')
            vc.border    = _thin_border()
            ws.row_dimensions[cur].height = 18
            cur += 1

        cur += 2

        # ── D. Grading scale reference ────────────────────────────────────
        ref_end = _grading_ref_table(ws, cur, grading_scale_list,
                                     badge_fill, text_color,
                                     last_col, educational_level)
        _footer(ws, ref_end + 1, last_col, school_meta['school_name'])

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 10

        _print_setup(ws, 12, last_col)

    # ── internal helper ───────────────────────────────────────────────────
    @staticmethod
    def _grade_from_pct(pct, grading_scale_list):
        for gs in grading_scale_list:
            if float(gs.min_mark) <= pct <= float(gs.max_mark):
                return gs.grade
        return grading_scale_list[-1].grade if grading_scale_list else '?'


class ExportStudentResultPDFView(ManagementRequiredMixin, View):
    """Export student results to PDF format using WeasyPrint."""
    
    def get(self, request, student_pk, session_pk):
        student = get_object_or_404(Student, pk=student_pk)
        session = get_object_or_404(ExamSession, pk=session_pk)
        
        # Get all the data (reuse from StudentResultDetailView)
        educational_level = session.class_level.educational_level
        level_type = educational_level.level_type
        is_primary_nursery = level_type in ['PRIMARY', 'NURSERY']
        is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
        is_alevel = level_type == 'A_LEVEL'
        
        enrollment = student.enrollments.filter(
            academic_year=session.academic_year,
            class_level=session.class_level
        ).first()
        
        # Get student's combination for A-Level
        student_combination = None
        if is_alevel and enrollment:
            combination_assignment = enrollment.combination_assignments.filter(is_active=True).first()
            if combination_assignment:
                student_combination = combination_assignment.combination
        
        metrics = StudentExamMetrics.objects.filter(
            student=student,
            exam_session=session
        ).first()
        
        position = StudentExamPosition.objects.filter(
            student=student,
            exam_session=session
        ).first()
        
        subject_results = StudentSubjectResult.objects.filter(
            student=student,
            exam_session=session
        ).select_related('subject').order_by('subject__name')
        
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        paper_scores = {
            ps.exam_paper_id: ps
            for ps in StudentPaperScore.objects.filter(
                student=student,
                exam_paper__exam_session=session
            ).select_related('exam_paper')
        }
        
        # Get all subject results for the session to calculate positions
        all_subject_results = StudentSubjectResult.objects.filter(
            exam_session=session
        ).select_related('subject')
        
        # Calculate subject-wise positions
        subject_positions = {}
        for result in all_subject_results:
            subj_id = result.subject_id
            if subj_id not in subject_positions:
                # Get all results for this subject
                subj_results = [r for r in all_subject_results if r.subject_id == subj_id]
                # Sort by total marks descending
                sorted_results = sorted(subj_results, key=lambda x: float(x.total_marks), reverse=True)
                # Create position map
                position_map = {}
                for idx, r in enumerate(sorted_results, 1):
                    position_map[r.student_id] = idx
                subject_positions[subj_id] = {
                    'position_map': position_map,
                    'total_students': len(sorted_results)
                }
        
        # Build subjects data with positions
        subjects_data = []
        for paper in papers:
            subject = paper.subject
            existing_subject = next((s for s in subjects_data if s['subject'].pk == subject.pk), None)
            
            if not existing_subject:
                subject_result = subject_results.filter(subject=subject).first()
                subject_max = sum(float(p.max_marks) for p in papers if p.subject_id == subject.pk)
                
                # Get position data for this subject
                position_data = subject_positions.get(subject.pk, {})
                student_position = position_data.get('position_map', {}).get(student.pk)
                total_students_in_subject = position_data.get('total_students', 0)
                
                subjects_data.append({
                    'subject': subject,
                    'papers': [],
                    'result': subject_result,
                    'total_marks': subject_result.total_marks if subject_result else None,
                    'grade': subject_result.grade if subject_result else None,
                    'percentage': (float(subject_result.total_marks) / subject_max * 100) if subject_result and subject_max > 0 else None,
                    'position': student_position,
                    'total_students': total_students_in_subject,
                    'position_display': f"{student_position}/{total_students_in_subject}" if student_position and total_students_in_subject > 0 else "—",
                })
            
            current_subject = next(s for s in subjects_data if s['subject'].pk == subject.pk)
            score = paper_scores.get(paper.pk)
            percentage = (score.marks / paper.max_marks * 100) if score and paper.max_marks > 0 else 0
            
            current_subject['papers'].append({
                'paper': paper,
                'marks': score.marks if score else None,
                'percentage': round(percentage, 2) if score else None,
                'max_marks': paper.max_marks,
                'has_score': score is not None,
            })
        
        # Calculate overall grade for Primary/Nursery
        overall_grade = None
        if is_primary_nursery and metrics and metrics.average_marks:
            grading_scale = GradingScale.objects.filter(education_level=educational_level).order_by('-min_mark')
            for gs in grading_scale:
                if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                    overall_grade = gs.grade
                    break
        
        # Prepare performance metrics for template
        performance_metrics = {
            'overall_grade': overall_grade,
        }
        
        # Prepare context for PDF template
        context = {
            'student': student,
            'session': session,
            'educational_level': educational_level,
            'level_type': level_type,
            'is_primary_nursery': is_primary_nursery,
            'is_secondary': is_secondary,
            'is_alevel': is_alevel,
            'student_combination': student_combination,
            'metrics': metrics,
            'position': position,
            'subjects_data': subjects_data,
            'performance_metrics': performance_metrics,
            'generated_date': datetime.now(),
            'school_name': request.site.name if hasattr(request, 'site') else 'School Management System',
        }
        
        # Render HTML template
        html_string = render_to_string('portal_management/students/student_result_pdf.html', context)
        
        # Configure fonts
        font_config = FontConfiguration()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f"student_results_{student.registration_number}_{session.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        try:
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config,
                presentational_hints=True,
            )
            return response
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            return JsonResponse({'success': False, 'message': str(e)}, status=500)



class ExportSessionAllStudentsPDFView(ManagementRequiredMixin, View):
    """
    Export results for ALL students in an exam session to a single PDF file.
    Each student's result appears on a separate page with full details.
    School information is retrieved from SchoolProfile model based on educational level.
    """
    
    def get(self, request, session_pk):
        session = get_object_or_404(ExamSession, pk=session_pk)
        
        # Get all enrolled students for this session's class level and academic year
        enrollments = StudentEnrollment.objects.filter(
            academic_year=session.academic_year,
            class_level=session.class_level,
            status='active'
        ).select_related('student').order_by('student__first_name', 'student__last_name')
        
        if not enrollments.exists():
            messages.warning(request, 'No students enrolled in this session.')
            return redirect('management:exam_session_list')
        
        # Get school profile for this educational level
        educational_level = session.class_level.educational_level
        school_profile = SchoolProfile.objects.get_active_profile(educational_level)
        
        # Prepare school info for template
        school_info = {
            'name': school_profile.name if school_profile else getattr(settings, 'SCHOOL_NAME', 'School Management System'),
            'address': school_profile.address if school_profile else getattr(settings, 'SCHOOL_ADDRESS', ''),
            'phone': school_profile.get_contact_phone() if school_profile else getattr(settings, 'SCHOOL_PHONE', ''),
            'email': school_profile.email if school_profile else getattr(settings, 'SCHOOL_EMAIL', ''),
            'motto': school_profile.motto if school_profile else getattr(settings, 'SCHOOL_MOTTO', ''),
            'registration_number': school_profile.registration_number if school_profile else getattr(settings, 'SCHOOL_REGISTRATION_NUMBER', ''),
            'logo': school_profile.logo.url if school_profile and school_profile.logo else None,
        }
        
        # Get grading scale for this educational level
        grading_scale = GradingScale.objects.filter(
            education_level=educational_level
        ).order_by('-min_mark')
        
        # Get division scale for secondary levels
        division_scale = None
        if educational_level.level_type in ['O_LEVEL', 'A_LEVEL']:
            division_scale = DivisionScale.objects.filter(
                education_level=educational_level
            ).order_by('min_points')
        
        # Get all metrics and positions for this session
        metrics_dict = {
            m.student_id: m for m in StudentExamMetrics.objects.filter(exam_session=session)
        }
        positions_dict = {
            p.student_id: p for p in StudentExamPosition.objects.filter(exam_session=session)
        }
        
        # Get all subject results for this session
        all_subject_results = StudentSubjectResult.objects.filter(
            exam_session=session
        ).select_related('subject', 'student')
        
        # Get all papers for this session
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        # Get all paper scores for this session
        paper_scores = {}
        for ps in StudentPaperScore.objects.filter(
            exam_paper__exam_session=session
        ).select_related('exam_paper'):
            key = f"{ps.student_id}_{ps.exam_paper_id}"
            paper_scores[key] = ps
        
        # Group subject results by student
        subject_results_by_student = {}
        for result in all_subject_results:
            if result.student_id not in subject_results_by_student:
                subject_results_by_student[result.student_id] = []
            subject_results_by_student[result.student_id].append(result)
        
        # Calculate subject-wise positions for each student
        subject_positions = self._calculate_subject_positions(all_subject_results)
        
        # Prepare data for each student
        students_data = []
        for enrollment in enrollments:
            student = enrollment.student
            
            # Get student's combination for A-Level
            student_combination = None
            if educational_level.level_type == 'A_LEVEL':
                combination_assignment = enrollment.combination_assignments.filter(is_active=True).first()
                if combination_assignment:
                    student_combination = combination_assignment.combination
            
            # Get student's metrics and position
            metrics = metrics_dict.get(student.pk)
            position = positions_dict.get(student.pk)
            
            # Get student's subject results
            student_results = subject_results_by_student.get(student.pk, [])
            
            # Build subjects data with papers
            subjects_data = self._build_subjects_data(
                student, student_results, papers, paper_scores, subject_positions, educational_level
            )
            
            # Calculate overall grade for Primary/Nursery
            overall_grade = None
            if educational_level.level_type in ['PRIMARY', 'NURSERY'] and metrics and metrics.average_marks:
                for gs in grading_scale:
                    if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                        overall_grade = gs.grade
                        break
            
            students_data.append({
                'student': student,
                'enrollment': enrollment,
                'combination': student_combination,
                'metrics': metrics,
                'position': position,
                'subjects_data': subjects_data,
                'overall_grade': overall_grade,
            })
        
        # Prepare context for PDF template
        context = {
            'session': session,
            'educational_level': educational_level,
            'level_type': educational_level.level_type,
            'school_info': school_info,
            'students_data': students_data,
            'grading_scale': grading_scale,
            'division_scale': division_scale,
            'generated_date': timezone.now(),
            'total_students': len(students_data),
        }
        
        # Render HTML template
        html_string = render_to_string('portal_management/exams/session_all_students_pdf.html', context)
        
        # Configure fonts
        font_config = FontConfiguration()
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        filename = f"exam_results_{session.name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        try:
            HTML(string=html_string).write_pdf(
                response,
                font_config=font_config,
                presentational_hints=True,
            )
            return response
        except Exception as e:
            logger.error(f'PDF generation error: {e}', exc_info=True)
            messages.error(request, f'Error generating PDF: {str(e)}')
            return redirect('management:exam_session_list')
    
    def _calculate_subject_positions(self, all_subject_results):
        """Calculate subject-wise positions for all students."""
        subject_positions = {}
        
        # Group results by subject
        results_by_subject = {}
        for result in all_subject_results:
            subj_id = result.subject_id
            if subj_id not in results_by_subject:
                results_by_subject[subj_id] = []
            results_by_subject[subj_id].append(result)
        
        # Calculate positions for each subject
        for subj_id, results in results_by_subject.items():
            # Sort by total marks descending
            sorted_results = sorted(results, key=lambda x: float(x.total_marks), reverse=True)
            
            # Create position map
            position_map = {}
            for idx, r in enumerate(sorted_results, 1):
                position_map[r.student_id] = idx
            
            subject_positions[subj_id] = {
                'position_map': position_map,
                'total_students': len(sorted_results)
            }
        
        return subject_positions
    
    def _build_subjects_data(self, student, student_results, papers, paper_scores, subject_positions, educational_level):
        """Build subject data for a student with papers and positions."""
        subjects_data = []
        
        # Group papers by subject
        papers_by_subject = {}
        for paper in papers:
            if paper.subject_id not in papers_by_subject:
                papers_by_subject[paper.subject_id] = []
            papers_by_subject[paper.subject_id].append(paper)
        
        # Process each subject
        for subject_id, subject_papers in papers_by_subject.items():
            subject = subject_papers[0].subject
            subject_result = next((r for r in student_results if r.subject_id == subject_id), None)
            
            # Calculate subject max marks
            subject_max = sum(float(p.max_marks) for p in subject_papers)
            
            # Get position data
            position_data = subject_positions.get(subject_id, {})
            student_position = position_data.get('position_map', {}).get(student.pk)
            total_students = position_data.get('total_students', 0)
            
            # Build papers data
            papers_data = []
            for paper in subject_papers:
                score_key = f"{student.pk}_{paper.pk}"
                score = paper_scores.get(score_key)
                
                papers_data.append({
                    'paper': paper,
                    'marks': score.marks if score else None,
                    'percentage': (float(score.marks) / float(paper.max_marks) * 100) if score and paper.max_marks > 0 else None,
                    'max_marks': paper.max_marks,
                    'has_score': score is not None,
                })
            
            subjects_data.append({
                'subject': subject,
                'papers': papers_data,
                'result': subject_result,
                'total_marks': subject_result.total_marks if subject_result else None,
                'grade': subject_result.grade if subject_result else None,
                'points': subject_result.points if subject_result else None,
                'percentage': (float(subject_result.total_marks) / subject_max * 100) if subject_result and subject_max > 0 else None,
                'position': student_position,
                'total_students': total_students,
                'position_display': f"{student_position}/{total_students}" if student_position and total_students > 0 else "—",
                'max_marks': subject_max,
            })
        
        # Sort subjects by name
        subjects_data.sort(key=lambda x: x['subject'].name)
        
        return subjects_data