from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import date
import logging
from collections import defaultdict
import re

from core.mixins import ManagementRequiredMixin
from core.models import (
    ExamSession, GradingScale, StudentPaperScore, Subject, SubjectExamPaper,
    Student, StudentSubjectResult, StudentExamMetrics, StudentExamPosition,
    SchoolProfile, EducationalLevel
)

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
#  Style helpers
# ──────────────────────────────────────────────────────────────

def _thin_border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=10, color='000000', italic=False, name='Arial'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hcell(ws, row, col, value, merge_end_col=None):
    """Write a header cell with bold font and centered alignment."""
    if merge_end_col and merge_end_col > col:
        # Ensure merge range is valid
        if merge_end_col <= ws.max_column:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=True, size=10, color='FFFFFF')
    c.fill = _fill('1B4F72')
    c.alignment = _align('center', 'center')
    c.border = _thin_border()
    return c


def _cell(ws, row, col, value, align='center', bold=False, color='000000'):
    """Write a data cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, size=9, color=color)
    c.alignment = _align(align, 'center')
    c.border = _thin_border()
    return c


def _sanitize_sheet_name(name):
    """Sanitize sheet name for Excel compatibility."""
    # Remove invalid characters: \ / ? * [ ] :
    name = re.sub(r'[\\/*?:\[\]]', '', name)
    # Limit to 31 characters
    if len(name) > 31:
        name = name[:31]
    # Ensure not empty
    if not name:
        name = "Results"
    return name


# ──────────────────────────────────────────────────────────────
#  Palette constants
# ──────────────────────────────────────────────────────────────

C_BRAND_DARK = '0D3349'
C_BRAND_MID = '1A5276'
C_ACCENT = '148F77'
C_HEADER_BG = '1B4F72'
C_SUBHDR_BG = '2980B9'
C_WHITE = 'FFFFFF'
C_GOLD = 'F0B429'
C_MALE_BG = 'E3F2FD'
C_FEMALE_BG = 'FCE4EC'


class ExportSessionReportView(ManagementRequiredMixin, View):
    """
    Export full session results with educational level-aware formatting.
    Handles Primary/Nursery (grades only), O-Level (division+points), 
    A-Level (combination+division+points).
    """

    def get(self, request, pk):
        session = get_object_or_404(ExamSession, pk=pk)
        
        try:
            wb = self._build_excel_report(session, request)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # Sanitize filename
            safe_name = f'results_{session.name[:30]}_{date.today()}'.replace(' ', '_').replace('/', '-')
            safe_name = re.sub(r'[\\/*?:\[\]]', '', safe_name)
            response['Content-Disposition'] = f'attachment; filename="{safe_name}.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f'ExportSessionReport error: {e}', exc_info=True)
            messages.error(request, f'Export error: {e}')
            return redirect('management:exam_session_detail', pk=pk)
    
    def _build_excel_report(self, session, request):
        """Build the Excel workbook with premium formatting."""
        # Sanitize sheet name
        sheet_name = _sanitize_sheet_name(f'{session.name[:25]} Results')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # ── Load all data ─────────────────────────────────────────────────────
        educational_level = session.class_level.educational_level
        level_type = educational_level.level_type
        is_primary_nursery = level_type in ['PRIMARY', 'NURSERY']
        is_secondary = level_type in ['O_LEVEL', 'A_LEVEL']
        is_alevel = level_type == 'A_LEVEL'
        
        # Get school profile for this educational level
        school_profile = SchoolProfile.objects.get_active_profile(educational_level)
        school_info = self._get_school_info(school_profile, educational_level)
        
        # Get all subjects for this session with their papers
        papers_by_subject = defaultdict(list)
        papers = SubjectExamPaper.objects.filter(
            exam_session=session
        ).select_related('subject').order_by('subject__name', 'paper_number')
        
        for paper in papers:
            papers_by_subject[paper.subject_id].append(paper)
        
        subjects = list(
            Subject.objects
            .filter(id__in=papers_by_subject.keys())
            .order_by('name')
            .values('id', 'name', 'short_name', 'code')
        )
        
        # Get all paper scores and calculate AVERAGE PERCENTAGE per subject
        paper_scores = defaultdict(lambda: defaultdict(list))
        
        for ps in StudentPaperScore.objects.filter(
            exam_paper__exam_session=session
        ).select_related('student', 'exam_paper'):
            student_id = ps.student_id
            subject_id = ps.exam_paper.subject_id
            max_marks = float(ps.exam_paper.max_marks)
            
            # Calculate percentage for this paper
            marks_scored = float(ps.marks)
            percentage = (marks_scored / max_marks * 100) if max_marks > 0 else 0
            
            # Store percentage for this paper
            paper_scores[student_id][subject_id].append(percentage)
        
        # Calculate subject AVERAGE percentage
        subject_averages = defaultdict(dict)
        
        for subject in subjects:
            subj_id = subject['id']
            
            for student_id, subject_scores in paper_scores.items():
                if subj_id in subject_scores:
                    percentages = subject_scores[subj_id]
                    avg_percentage = sum(percentages) / len(percentages) if percentages else 0
                    
                    subject_averages[student_id][subj_id] = {
                        'avg_percentage': round(avg_percentage, 2),
                        'avg_marks': round(avg_percentage, 2),
                        'paper_count': len(percentages),
                    }
        
        # Get all results (for grades)
        results_qs = list(
            StudentSubjectResult.objects
            .filter(exam_session=session)
            .values('student_id', 'subject_id', 'total_marks', 'grade', 'points')
        )
        
        results_by_student = defaultdict(dict)
        for row in results_qs:
            results_by_student[row['student_id']][row['subject_id']] = row
        
        # Get metrics
        metrics_qs = StudentExamMetrics.objects.filter(exam_session=session).select_related('student')
        metrics_by_student = {m.student_id: m for m in metrics_qs}
        
        # Get positions
        positions_qs = StudentExamPosition.objects.filter(exam_session=session)
        positions_by_student = {p.student_id: p for p in positions_qs}
        
        # Get student info
        student_ids = set(metrics_by_student.keys())
        student_info = {}
        for s in Student.objects.filter(id__in=student_ids).select_related('user'):
            student_info[s.id] = {
                'id': s.id,
                'first_name': s.first_name,
                'last_name': s.last_name,
                'registration_number': s.registration_number,
                'gender': s.gender,
                'gender_display': s.get_gender_display(),
            }
        
        # Get combination assignments for A-Level students
        combination_by_student = {}
        if is_alevel:
            from core.models import StudentCombinationAssignment
            assignments = StudentCombinationAssignment.objects.filter(
                student_id__in=student_ids,
                enrollment__academic_year=session.academic_year,
                is_active=True
            ).select_related('combination')
            for assignment in assignments:
                combination_by_student[assignment.student_id] = assignment.combination
        
        # Sort students by class position
        sorted_student_ids = sorted(
            metrics_by_student.keys(),
            key=lambda sid: positions_by_student.get(sid).class_position if positions_by_student.get(sid) else 9999,
        )
        
        # Calculate statistics
        total_students = len(metrics_by_student)
        students_with_results = len([s for s in metrics_by_student.values() if s.total_marks])
        total_marks_list = [float(m.total_marks) for m in metrics_by_student.values() if m.total_marks]
        
        # Gender statistics
        male_students = [sid for sid in student_ids if student_info.get(sid, {}).get('gender') == 'male']
        female_students = [sid for sid in student_ids if student_info.get(sid, {}).get('gender') == 'female']
        
        stats = {
            'total_students': total_students,
            'students_with_results': students_with_results,
            'average_total': sum(total_marks_list) / len(total_marks_list) if total_marks_list else 0,
            'highest_total': max(total_marks_list) if total_marks_list else 0,
            'lowest_total': min(total_marks_list) if total_marks_list else 0,
            'completion_rate': (students_with_results / total_students * 100) if total_students else 0,
            'male_count': len(male_students),
            'female_count': len(female_students),
        }
        
        # ── Build the worksheet ─────────────────────────────────────────────────
        # Calculate total columns first
        static_cols_count = 6 if is_alevel else 5
        subject_cols_count = len(subjects) * 2
        if is_primary_nursery:
            summary_cols_count = 4
        elif is_secondary:
            summary_cols_count = 5
        else:
            summary_cols_count = 3
        total_cols = static_cols_count + subject_cols_count + summary_cols_count
        
        # Build the worksheet with proper row counting
        current_row = self._add_school_header(ws, session, school_info, educational_level, request, total_cols)
        current_row = self._add_session_info(ws, current_row, session, stats, is_alevel, request, total_cols)
        headers_result = self._add_headers(ws, current_row, subjects, is_secondary, is_alevel, is_primary_nursery)
        current_row, subj_col_map, sum_col, col = headers_result
        header_end_row = current_row - 1
        
        if sorted_student_ids:
            current_row = self._add_data_rows(ws, current_row, sorted_student_ids, student_info, 
                                            metrics_by_student, positions_by_student, 
                                            results_by_student, subject_averages, 
                                            subjects, is_secondary, is_alevel, combination_by_student,
                                            is_primary_nursery, educational_level, subj_col_map, sum_col)
        
        self._add_footer(ws, current_row, school_info, total_cols)
        self._format_columns(ws, subjects, is_secondary, is_alevel, is_primary_nursery, header_end_row)
        
        return wb
    
    def _get_school_info(self, school_profile, educational_level):
        """Get school information from profile or fallback to settings."""
        if school_profile:
            return {
                'name': school_profile.name or 'School Management System',
                'address': school_profile.address or '',
                'phone': school_profile.get_contact_phone() or '',
                'email': school_profile.email or '',
                'motto': school_profile.motto or '',
                'registration_number': school_profile.registration_number or '',
                'website': school_profile.website or '',
                'logo_present': bool(school_profile.logo),
                'contact_person': school_profile.get_contact_name() or 'Headmaster',
                'educational_level': educational_level.name if educational_level else '',
            }
        else:
            return {
                'name': getattr(settings, 'SCHOOL_NAME', 'School Management System'),
                'address': getattr(settings, 'SCHOOL_ADDRESS', ''),
                'phone': getattr(settings, 'SCHOOL_PHONE', ''),
                'email': getattr(settings, 'SCHOOL_EMAIL', ''),
                'motto': getattr(settings, 'SCHOOL_MOTTO', ''),
                'registration_number': getattr(settings, 'SCHOOL_REGISTRATION_NUMBER', ''),
                'website': '',
                'logo_present': False,
                'contact_person': 'Headmaster',
                'educational_level': educational_level.name if educational_level else '',
            }
    
    def _add_school_header(self, ws, session, school_info, educational_level, request, total_cols):
        """Add the premium school header section."""
        row = 1
        last_col = total_cols
        
        def band(r, value, fill_hex, font_obj, align_obj=None, height=None):
            if last_col >= 1:
                ws.merge_cells(f'A{r}:{get_column_letter(last_col)}{r}')
                c = ws.cell(row=r, column=1, value=value)
                c.fill = _fill(fill_hex)
                c.font = font_obj
                c.alignment = align_obj or _align('center', 'center')
                if height:
                    ws.row_dimensions[r].height = height
            return c
        
        # Ensure we have valid values
        school_name = school_info.get('name', 'School Management System')
        school_address = school_info.get('address', '')
        school_phone = school_info.get('phone', '')
        school_email = school_info.get('email', '')
        school_motto = school_info.get('motto', '')
        school_reg_no = school_info.get('registration_number', '')
        
        band(row, '', C_BRAND_DARK, _font(), height=8)
        row += 1
        band(row, str(school_name).upper(), C_BRAND_DARK,
             _font(bold=True, size=18, color=C_GOLD),
             _align('center', 'center'), height=38)
        row += 1
        
        contact_parts = [p for p in [school_address, school_phone, school_email] if p]
        band(row, '   |   '.join(contact_parts) if contact_parts else '',
             C_BRAND_MID, _font(size=9, color='D6EAF8', italic=True),
             _align('center', 'center'), height=18)
        row += 1
        
        motto_line = ''
        if school_motto:
            motto_line += f'"{school_motto}"'
        if school_reg_no:
            motto_line += f'   Reg No: {school_reg_no}'
        band(row, motto_line, C_BRAND_MID,
             _font(size=9, color='AED6F1', italic=bool(school_motto)),
             _align('center', 'center'), height=18)
        row += 1
        
        band(row, '', C_GOLD, _font(), height=4)
        row += 1
        band(row, 'EXAMINATION RESULTS REPORT', C_ACCENT,
             _font(bold=True, size=13, color=C_WHITE),
             _align('center', 'center'), height=28)
        row += 1
        
        level_display = f'{educational_level.name.upper()}  ({educational_level.get_level_type_display()})' if educational_level else 'RESULTS'
        band(row, level_display, '0E6655',
             _font(bold=True, size=11, color='A9DFBF'),
             _align('center', 'center'), height=22)
        row += 1
        
        return row
    
    def _add_session_info(self, ws, start_row, session, stats, is_alevel, request, total_cols):
        """Add session information rows."""
        row = start_row
        last_col = total_cols
        
        if last_col < 1:
            return row
        
        stream_part = f'   │   Stream: {session.stream_class.name}' if session.stream_class else ''
        session_value = f'Session: {session.name}   │   Class: {session.class_level.name}   │   Term: {session.term}   │   Academic Year: {session.academic_year}{stream_part}'
        ws.merge_cells(f'A{row}:{get_column_letter(last_col)}{row}')
        c = ws.cell(row=row, column=1, value=session_value)
        c.font = _font(bold=True, size=9, color='1A5276')
        c.fill = _fill('D6EAF8')
        c.alignment = _align('center', 'center')
        ws.row_dimensions[row].height = 20
        row += 1
        
        info_value = f'Education Level: {session.class_level.educational_level.name}   │   Total Students: {stats["total_students"]}   │   Male: {stats["male_count"]}   │   Female: {stats["female_count"]}   │   Completion Rate: {stats["completion_rate"]:.1f}%'
        ws.merge_cells(f'A{row}:{get_column_letter(last_col)}{row}')
        c = ws.cell(row=row, column=1, value=info_value)
        c.font = _font(size=9, color='1A5276')
        c.fill = _fill('EBF5FB')
        c.alignment = _align('center', 'center')
        ws.row_dimensions[row].height = 18
        row += 1
        
        generated_by = getattr(request, 'user', None)
        generated_by_name = generated_by.get_full_name() if generated_by and hasattr(generated_by, 'get_full_name') else 'System'
        gen_value = f'Generated: {timezone.now().strftime("%A, %d %B %Y  %H:%M")}   │   By: {generated_by_name}'
        ws.merge_cells(f'A{row}:{get_column_letter(last_col)}{row}')
        c = ws.cell(row=row, column=1, value=gen_value)
        c.font = _font(size=8, color='555555', italic=True)
        c.fill = _fill('F2F3F4')
        c.alignment = _align('center', 'center')
        ws.row_dimensions[row].height = 16
        row += 1
        
        ws.merge_cells(f'A{row}:{get_column_letter(last_col)}{row}')
        ws.cell(row=row, column=1).fill = _fill('D5D8DC')
        ws.row_dimensions[row].height = 4
        row += 1
        
        return row
    
    def _add_headers(self, ws, start_row, subjects, is_secondary, is_alevel, is_primary_nursery):
        """Add column headers with educational level awareness."""
        row = start_row
        
        # Static headers
        static_headers = ['POS', 'STREAM\nPOS', 'GENDER', 'REG NO', 'STUDENT NAME']
        if is_alevel:
            static_headers.append('COMBINATION')
        
        col = 1
        for h in static_headers:
            _hcell(ws, row, col, h)
            ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            col += 1
        
        # Subject headers
        subj_col_map = {}
        for subj in subjects:
            label = subj.get('short_name') or subj.get('name', 'Subject')
            _hcell(ws, row, col, label, merge_end_col=col + 1)
            _hcell(ws, row + 1, col, 'MK')
            _hcell(ws, row + 1, col + 1, 'GR')
            subj_col_map[subj['id']] = col
            col += 2
        
        # Summary headers
        if is_primary_nursery:
            summary_headers = ['TOTAL\nMARKS', 'AVERAGE', 'GRADE', 'REMARK']
        elif is_secondary:
            summary_headers = ['TOTAL\nMARKS', 'AVERAGE', 'POINTS', 'DIVISION', 'REMARK']
        else:
            summary_headers = ['TOTAL\nMARKS', 'AVERAGE', 'REMARK']
        
        sum_col = {}
        for h in summary_headers:
            _hcell(ws, row, col, h)
            ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            sum_col[h] = col
            col += 1
        
        ws.row_dimensions[row].height = 28
        ws.row_dimensions[row + 1].height = 22
        
        return row + 2, subj_col_map, sum_col, col
    
    def _add_data_rows(self, ws, start_row, student_ids, student_info, 
                    metrics_by_student, positions_by_student, 
                    results_by_student, subject_averages,
                    subjects, is_secondary, is_alevel, combination_by_student,
                    is_primary_nursery, educational_level, subj_col_map, sum_col):
        """Add student data rows."""
        row = start_row
        
        static_cols = {'pos': 1, 'stream_pos': 2, 'gender': 3, 'reg_no': 4, 'student_name': 5}
        if is_alevel:
            static_cols['combination'] = 6
        
        for idx, student_id in enumerate(student_ids, 1):
            info = student_info.get(student_id, {})
            metrics = metrics_by_student.get(student_id)
            position = positions_by_student.get(student_id)
            results = results_by_student.get(student_id, {})
            
            full_name = f"{info.get('first_name', '')} {info.get('last_name', '')}".strip()
            if not full_name:
                full_name = '-'
            gender = info.get('gender_display', '-')
            gender_bg = C_MALE_BG if info.get('gender') == 'male' else (C_FEMALE_BG if info.get('gender') == 'female' else 'FFFFFF')
            
            class_pos = position.class_position if position else '-'
            stream_pos = position.stream_position if position else '-'
            
            # Static columns
            _cell(ws, row, static_cols['pos'], class_pos, 'center', bold=True)
            _cell(ws, row, static_cols['stream_pos'], stream_pos, 'center')
            
            gender_cell = ws.cell(row=row, column=static_cols['gender'])
            gender_cell.value = gender
            gender_cell.font = _font(size=9)
            gender_cell.fill = _fill(gender_bg)
            gender_cell.alignment = _align('center', 'center')
            gender_cell.border = _thin_border()
            
            _cell(ws, row, static_cols['reg_no'], info.get('registration_number', '-'), 'left')
            _cell(ws, row, static_cols['student_name'], full_name, 'left', bold=True)
            
            if is_alevel:
                combination = combination_by_student.get(student_id)
                _cell(ws, row, static_cols['combination'], combination.code if combination else '-', 'center')
            
            # Subject results
            for subj in subjects:
                subj_id = subj['id']
                subject_data = subject_averages.get(student_id, {}).get(subj_id, {})
                
                avg_marks = subject_data.get('avg_marks', 0)
                avg_percentage = subject_data.get('avg_percentage', 0)
                
                result = results.get(subj_id, {})
                grade = result.get('grade', '-')
                
                start_col = subj_col_map[subj_id]
                
                # Marks cell
                marks_cell = ws.cell(row=row, column=start_col)
                if avg_marks > 0:
                    marks_cell.value = avg_marks
                    marks_cell.number_format = '0.0'
                    marks_cell.font = _font(size=9)
                    if avg_percentage >= 75:
                        marks_cell.fill = _fill('E8F5E9')
                    elif avg_percentage >= 50:
                        marks_cell.fill = _fill('E3F2FD')
                    elif avg_percentage >= 30:
                        marks_cell.fill = _fill('FFF3E0')
                    else:
                        marks_cell.fill = _fill('FFEBEE')
                else:
                    marks_cell.value = '-'
                    marks_cell.font = _font(size=9, color='AAAAAA', italic=True)
                    marks_cell.fill = _fill('F9F9F9')
                marks_cell.alignment = _align('center', 'center')
                marks_cell.border = _thin_border()
                
                # Grade cell
                grade_cell = ws.cell(row=row, column=start_col + 1)
                grade_cell.value = grade
                grade_cell.font = _font(bold=True, size=9)
                grade_cell.alignment = _align('center', 'center')
                grade_cell.border = _thin_border()
                if grade != '-':
                    grade_cell.fill = _fill('E8F5E9')
            
            # Summary metrics
            if metrics:
                remark = 'PASS' if metrics.average_marks and metrics.average_marks >= 40 else 'FAIL'
                remark_fill = _fill('E8F5E9') if remark == 'PASS' else _fill('FFEBEE')
                remark_font = _font(bold=True, size=9, color='1E7E34' if remark == 'PASS' else 'B71C1C')
                
                _cell(ws, row, sum_col['TOTAL\nMARKS'], float(metrics.total_marks) if metrics.total_marks else '-', 'center', bold=True)
                
                avg_cell = ws.cell(row=row, column=sum_col['AVERAGE'])
                avg_cell.value = float(metrics.average_marks) if metrics.average_marks else '-'
                avg_cell.number_format = '0.0'
                avg_cell.font = _font(size=9)
                avg_cell.alignment = _align('center', 'center')
                avg_cell.border = _thin_border()
                
                if is_primary_nursery:
                    overall_grade = '-'
                    if metrics.average_marks:
                        grading_scale = GradingScale.objects.filter(
                            education_level=educational_level
                        ).order_by('-min_mark')
                        for gs in grading_scale:
                            if float(gs.min_mark) <= float(metrics.average_marks) <= float(gs.max_mark):
                                overall_grade = gs.grade
                                break
                    _cell(ws, row, sum_col['GRADE'], overall_grade, 'center', bold=True)
                    
                elif is_secondary:
                    _cell(ws, row, sum_col['POINTS'], float(metrics.total_points) if metrics.total_points else '-', 'center')
                    _cell(ws, row, sum_col['DIVISION'], metrics.division or '-', 'center', bold=True)
                
                remark_cell = ws.cell(row=row, column=sum_col['REMARK'])
                remark_cell.value = remark
                remark_cell.font = remark_font
                remark_cell.fill = remark_fill
                remark_cell.alignment = _align('center', 'center')
                remark_cell.border = _thin_border()
            else:
                for col_name, col_idx in sum_col.items():
                    empty_cell = ws.cell(row=row, column=col_idx)
                    empty_cell.value = '-'
                    empty_cell.font = _font(size=9, color='AAAAAA', italic=True)
                    empty_cell.alignment = _align('center', 'center')
                    empty_cell.border = _thin_border()
            
            ws.row_dimensions[row].height = 22
            row += 1
        
        return row
    
    def _add_footer(self, ws, start_row, school_info, total_cols):
        """Add footer with generation info."""
        if start_row <= 1:
            start_row = 2
        
        row = start_row + 2
        last_col = total_cols
        
        if last_col >= 1:
            ws.merge_cells(f'A{row}:{get_column_letter(last_col)}{row}')
            school_name = school_info.get('name', 'School Management System')
            footer_value = f'This report is computer-generated and is valid without a signature.  © {date.today().year} {school_name}  │  Printed: {timezone.now().strftime("%d %b %Y %H:%M")}'
            c = ws.cell(row=row, column=1, value=footer_value)
            c.font = _font(size=8, color='888888', italic=True)
            c.fill = _fill('F2F3F4')
            c.alignment = _align('center', 'center')
            ws.row_dimensions[row].height = 16
            
            ws.merge_cells(f'A{row+1}:{get_column_letter(last_col)}{row+1}')
            ws.cell(row=row + 1, column=1).fill = _fill(C_GOLD)
            ws.row_dimensions[row + 1].height = 5
    
    def _format_columns(self, ws, subjects, is_secondary, is_alevel, is_primary_nursery, header_end_row):
        """Format column widths based on content."""
        # Set column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 28
        
        if is_alevel:
            ws.column_dimensions['F'].width = 12
            col_start = 7
        else:
            col_start = 6
        
        col = col_start
        for _ in subjects:
            if col <= ws.max_column:
                ws.column_dimensions[get_column_letter(col)].width = 10
            if col + 1 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 1)].width = 8
            col += 2
        
        # Summary columns
        if col <= ws.max_column:
            ws.column_dimensions[get_column_letter(col)].width = 12
        if col + 1 <= ws.max_column:
            ws.column_dimensions[get_column_letter(col + 1)].width = 10
        
        if is_primary_nursery:
            if col + 2 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 2)].width = 8
            if col + 3 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 3)].width = 10
        elif is_secondary:
            if col + 2 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 2)].width = 8
            if col + 3 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 3)].width = 10
            if col + 4 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 4)].width = 10
        else:
            if col + 2 <= ws.max_column:
                ws.column_dimensions[get_column_letter(col + 2)].width = 10
        
        # Freeze panes - ensure header_end_row is valid
        freeze_col = get_column_letter(col_start)
        data_start_row = header_end_row + 1
        if data_start_row > 1:
            ws.freeze_panes = f'{freeze_col}{data_start_row}'
        ws.print_title_rows = f'1:{header_end_row}'
        
        # Page setup
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0